"""
visual_library_demo_v8.py
Enhanced version with full feature set (up to 5 levels deep)
Features:
- Recursive folder scanning with configurable folder_to_scan and output_file
- Hierarchical TOC with collapsible sections and hierarchical font sizes
- Folder-level comments via .comments files (title:, comments:, link: support)
- Path-based compact_view, max_view, and list_view specification
- Search includes folder names
- Full path exclusion support
- Custom heading font sizes and page font/font-size control
- Smooth scroll spy navigation
- Modern slim scrollbars (hover-reveal on sidebar)
- Border-free theme toggle button
- Underline headings support (underline_headings parameter)
- Page header (navigation_title, page_title, page_subtitle, page_description, footer_content)
- Theme config file support (theme_config_file, theme)
- All existing features preserved
"""

from pathlib import Path
from collections import defaultdict
import csv
import json
import re
from typing import Dict, List, Any, Optional, Union, Tuple


def create_site(
    folder_to_scan: str,
    output_file: str = None,
    side_bar_width: str = "250px",
    page_color: str = None,
    toc_color: str = None,
    toc_bg_color: str = None,
    modal_background_color: str = None,
    light_mode_colors: Dict[str, str] = None,
    dark_mode_colors: Dict[str, str] = None,
    default_theme: str = "light",
    image_extensions: List[str] = None,
    exclude_folders: List[str] = None,
    compact_view: List[str] = None,
    max_view: List[str] = None,
    hide_icons: List[str] = None,
    max_view_thumb_size: int = 320,
    compact_thumb_size: int = 40,
    card_image_height: int = None,
    toc_link_gap: int = 1,
    compact_column_ratio: List[int] = None,
    max_depth: int = 5,
    heading_font_sizes: Dict[int, str] = None,
    page_font: str = None,
    page_font_size: str = None,
    comment_title_font_size: str = "12px",
    comment_body_font_size: str = "12px",
    navigation_title: str = "Visual Library",
    page_title: str = None,
    page_subtitle: str = None,
    page_description: str = None,
    footer_content: str = None,
    theme_config_file: str = None,
    theme_sheet: str = "Theme_Config",
    theme: str = None,
    underline_headings: Dict[str, Any] = None,
    nav_font_sizes: Dict[int, str] = None,
    store_content: bool = False,
    sidebar_collapsible: bool = False,
    rename_headings: Dict[str, str] = None,
):
    """
    Automatically generates an HTML gallery page by scanning a folder structure recursively.
    Supports nested folders up to max_depth levels.

    .comments file format (place a file named '.comments' in any scanned folder):
        title:    Short display title — shown in compact-view title column and card caption.
                  If a link: is also set, the title becomes a hyperlink.
        link:     URL or relative file/folder path — wraps the title or filename in <a>.
        keywords: Space- or comma-separated extra search terms appended to each item's
                  search index so items surface even when the keyword isn't in the filename.
                  Example:  keywords: revenue, kpi, quarterly
        comments: Free-form description text (may continue on subsequent lines).
                  Shown in the third column of compact view and as a section description.

    Keyboard shortcuts (built into the generated HTML):
        F        Focus the search box.
        T        Toggle light / dark theme.
        N        Focus the navigation sidebar (then ↑/↓ to move between links).
        M        Focus the main content area (default on load; use for keyboard scrolling).
        ↑ / ↓   Navigate visible sidebar links (only when a sidebar link is focused).
        Esc      Clear search box / close image modal (returns focus to main content).
        ?        Click the ? button in the sidebar header to see the shortcut cheatsheet.

    Args:
        folder_to_scan (str): Path to the folder to scan recursively for images.
            The folder can have any name; subfolders become hierarchy levels in the
            sidebar and headings on the page.
        output_file (str): Full path (including filename) for the generated HTML file.
            Defaults to <folder_to_scan>/visual_library.html when not set.
            If output_file is outside folder_to_scan and store_content=False,
            image paths will be broken — use store_content=True for portability.
        side_bar_width (str): CSS width of the fixed left sidebar. Default: "250px".
        page_color (str): Override body background colour (rarely needed; prefer
            light_mode_colors/dark_mode_colors).
        toc_color (str): Override sidebar link colour.
        toc_bg_color (str): Override sidebar background colour.
        modal_background_color (str): Override image-modal backdrop colour.
        light_mode_colors (dict): Full colour scheme for light mode. Keys:
            bg, text, sidebar_bg, accent, border, hover, modal_bg.
        dark_mode_colors (dict): Full colour scheme for dark mode. Same keys.
        default_theme (str): Starting theme — "light" or "dark". Default: "light".
            The user's last choice is persisted in localStorage.
        image_extensions (list): File extensions treated as the primary thumbnail.
            Default includes .jpg .jpeg .png .gif .bmp .svg .webp .jfif .tiff .tif.
        exclude_folders (list): Folder names or relative paths to skip entirely.
            Exact name match:  "img"
            Path match:        "slidejs/img"
        compact_view (list): Path patterns for folders rendered as compact table rows
            (thumbnail | icons | comments) instead of a card grid.
            Supports exact paths "Category/Sub", wildcards "*/reference", "bar/*".
            Matching is case-insensitive.
        max_view (list): Path patterns for folders whose card grid uses wider thumbnails
            (max_view_thumb_size px minimum instead of 250 px). Same pattern syntax.
        hide_icons (list): Path patterns for folders where the SVG icon strip is hidden —
            both the folder-explore icon and any associated-file-extension icons are
            omitted from compact-view rows *and* card-view file-list bars, giving a
            cleaner look.  Accepts the same wildcard syntax as compact_view:
              "Category/Sub"   — exact path
              "*/reference"    — any folder named "reference"
              "bar/*"          — everything under "bar"
            Matching is case-insensitive.  Default: [] (icons shown everywhere).
        max_view_thumb_size (int): Min-width in px for max_view gallery cards. Default: 320.
        compact_thumb_size (int): Thumbnail pixel size (width) in compact-view rows. Default: 40.
        card_image_height (int|None): Fixed px height for card-view images; None = natural. Default: None.
        toc_link_gap (int): Vertical gap in px between sidebar navigation links. Default: 1.
        compact_column_ratio (list): Three ints [thumb%, icons%, comments%] controlling
            compact-row column widths. Must sum to ~100. Default: [30, 20, 50].
        max_depth (int): Maximum folder nesting levels to scan. Default: 5.
        heading_font_sizes (dict): Font sizes per heading level (1–6).
            Example: {{1: "32px", 2: "28px", 3: "24px", 4: "20px", 5: "18px", 6: "16px"}}
            Defaults: 28 / 24 / 20 / 18 / 16 / 14 px.
        page_font (str): CSS font-family applied to the entire page body.
            Example: "Arial" or "'Inter', sans-serif". Default: system-ui stack.
        page_font_size (str): CSS font-size for the body. Example: "14px". Default: "14px".
        comment_title_font_size (str): Font size for the title line in .comments blocks
            (compact-view title column and card caption link). Default: "12px".
        comment_body_font_size (str): Font size for the description body in .comments
            blocks. Default: "12px".
        navigation_title (str): Text shown in the sidebar header above the search box.
            Default: "Visual Library".
        page_title (str): Optional large title rendered at the top of the main content.
        page_subtitle (str): Optional subtitle rendered below page_title.
        page_description (str): Optional HTML description rendered below page_subtitle.
        footer_content (str): Optional HTML rendered in a footer bar at the bottom
            of the main content.
        theme_config_file (str|Path): Path to an .xlsm/.xlsx workbook containing a
            Theme_Config sheet. When provided with theme=, colours are loaded from it.
        theme_sheet (str): Sheet name inside theme_config_file. Default: "Theme_Config".
            Expected columns: Test_ID, Theme_Name, primary, text, bg, slide_bg,
            bg_light, border, border_muted, content_bg, muted.
            Test_ID matches the theme= value; Theme_Name must be "light" or "dark".
        theme (str): Theme ID to look up in the Test_ID column. Requires theme_config_file.
        underline_headings (dict): Per-level bottom border drawn under section headings.
            Keys are "h1"–"h6"; values are option dicts with any of:
              line_width (str)  default "1px"
              line_color (str)  default "var(--border-color)"
              line_style (str)  default "solid"
            Example: {{"h2": {{}}, "h3": {{"line_color": "#555", "line_width": "2px"}}}}
        store_content (bool): When True every image and linked asset is embedded as a
            base64 data URI so the HTML file is fully self-contained and works from any
            location. When False (default) all paths stay relative to folder_to_scan —
            output_file must live inside or adjacent to the scanned folder.

    Returns:
        str: Absolute path to the generated HTML file, or None on error.
    """

    # Set default values
    if light_mode_colors is None:
        light_mode_colors = {
            "bg": "#ffffff",
            "text": "#333333",
            "sidebar_bg": "#f8f9fa",
            "accent": "#007bff",
            "border": "#dee2e6",
            "hover": "#f1f3f5",
            "modal_bg": "rgba(0,0,0,0.8)",
        }

    if dark_mode_colors is None:
        dark_mode_colors = {
            "bg": "#1e1e1e",
            "text": "#e0e0e0",
            "sidebar_bg": "#2d2d2d",
            "accent": "#4a9eff",
            "border": "#444444",
            "hover": "#383838",
            "modal_bg": "rgba(0,0,0,0.9)",
        }

    if image_extensions is None:
        image_extensions = [
            ".jpg",
            ".jpeg",
            ".png",
            ".gif",
            ".bmp",
            ".svg",
            ".webp",
            ".jfif",
            ".tiff",
            ".tif",
        ]

    if exclude_folders is None:
        exclude_folders = []

    if compact_view is None:
        compact_view = []

    if max_view is None:
        max_view = []

    if hide_icons is None:
        hide_icons = []

    if compact_column_ratio is None:
        compact_column_ratio = [30, 20, 50]

    # Normalize rename_headings
    _rename_lookup = {}
    if rename_headings:
        for _k, _v in rename_headings.items():
            if _k is None:
                continue
            _rename_lookup[str(_k).strip().lower()] = _v

    def resolve_display_name(key, current_path, default_display):
        """ Return overriden heading text from rename_headings, else default. """
        if not _rename_lookup:
            return default_display
        candidates = []
        if key:
            candidates.append(str(key))
        if default_display:
            candidates.append(str(default_display))
        if current_path is not None:
            path_parts = list(current_path) + ([key] if key else [])
            if path_parts:
                candidates.append("/".join(path_parts))
                candidates.append("\\".join(path_parts))
        for cand in candidates:
            hit = _rename_lookup.get(cand.strip().lower())
            if hit is not None:
                return hit
        return default_display

    # ── Theme config loading (theme= + theme_config_file= take priority) ──────
    if theme is not None and theme_config_file is not None:
        try:
            import io
            import openpyxl

            _path = Path(theme_config_file)
            if not _path.exists():
                print(f"❌ theme_config_file not found: {_path}")
            else:
                with open(_path, "rb") as _fh:
                    _file_bytes = io.BytesIO(_fh.read())
                _wb = openpyxl.load_workbook(
                    _file_bytes, read_only=True, data_only=True, keep_vba=False
                )
                if theme_sheet not in _wb.sheetnames:
                    print(
                        f"❌ Sheet '{theme_sheet}' not found. "
                        f"Available: {_wb.sheetnames}"
                    )
                else:
                    _ws = _wb[theme_sheet]
                    _rows = list(_ws.iter_rows(values_only=True))
                    _header = [
                        str(c).strip().lower() if c is not None else ""
                        for c in _rows[0]
                    ]

                    def _col(name):
                        key = name.strip().lower()
                        if key not in _header:
                            raise KeyError(f"Column '{name}' not found in {theme_sheet}")
                        return _header.index(key)

                    _idx = {
                        "test_id":      _col("test_id"),
                        "theme_name":   _col("theme_name"),
                        "primary":      _col("primary"),
                        "text":         _col("text"),
                        "muted":        _col("muted"),
                        "content_bg":   _col("content_bg"),
                        "slide_bg":     _col("slide_bg"),
                        "bg":           _col("bg"),
                        "bg_light":     _col("bg_light"),
                        "border":       _col("border"),
                        "border_muted": _col("border_muted"),
                    }

                    _light_row = _dark_row = None
                    for _row in _rows[1:]:
                        _tid = (
                            str(_row[_idx["test_id"]]).strip()
                            if _row[_idx["test_id"]] is not None else ""
                        )
                        _tname = (
                            str(_row[_idx["theme_name"]]).strip().lower()
                            if _row[_idx["theme_name"]] is not None else ""
                        )
                        if _tid == str(theme).strip():
                            if _tname == "light":
                                _light_row = _row
                            elif _tname == "dark":
                                _dark_row = _row

                    if _light_row is None and _dark_row is None:
                        print(f"❌ theme '{theme}' not found in '{theme_sheet}'")
                    else:
                        def _v(row, key):
                            val = row[_idx[key]]
                            return str(val).strip() if val is not None else ""

                        def _build_colors(row):
                            return {
                                "bg":         _v(row, "bg"),
                                "text":       _v(row, "text"),
                                "sidebar_bg": _v(row, "slide_bg"),
                                "accent":     _v(row, "primary"),
                                "border":     _v(row, "border"),
                                "hover":      _v(row, "bg_light"),
                                "modal_bg":   "rgba(0,0,0,0.85)",
                            }

                        if _light_row is not None:
                            light_mode_colors = {**light_mode_colors, **_build_colors(_light_row)}
                        if _dark_row is not None:
                            dark_mode_colors = {**dark_mode_colors, **_build_colors(_dark_row)}
                        _modes = []
                        if _light_row: _modes.append("light")
                        if _dark_row:  _modes.append("dark")
                        print(f"✅ Theme '{theme}' loaded ({' + '.join(_modes)} modes)")
        except ImportError:
            print("❌ openpyxl is required for theme_config_file — pip install openpyxl")
        except Exception as _te:
            import traceback
            print(f"❌ theme_config_file error: {_te}")
            traceback.print_exc()
    elif theme is not None and theme_config_file is None:
        print("⚠️  theme= provided but theme_config_file= is missing — using default colors")

    # Set default heading font sizes if not provided
    if heading_font_sizes is None:
        heading_font_sizes = {
            1: "28px",
            2: "24px",
            3: "20px",
            4: "18px",
            5: "16px",
            6: "14px",
        }

    # Generate CSS for heading font sizes (plain string, injected via concatenation not f-string)
    heading_css = ""
    for level, size in heading_font_sizes.items():
        heading_css += f"\n        h{level} {{\n            font-size: {size};\n        }}"

    # Generate underline_headings CSS
    def _build_underline_css(uh):
        """Generate border-bottom CSS for each configured heading level."""
        if not uh:
            return ""
        css_lines = []
        for level, opts in uh.items():
            level  = level.lower().strip()
            width  = opts.get("line_width",  "1px")
            color  = opts.get("line_color",  "var(--border-color)")
            style_ = opts.get("line_style",  "solid")
            border = f"{width} {style_} {color}"
            css_lines.append(
                f"        {level}.section-heading {{\n"
                f"            border-bottom: {border};\n"
                f"            padding-bottom: 10px;\n"
                f"            margin-bottom: 20px;\n"
                f"        }}"
            )
        return "\n\n".join(css_lines)

    _underline_css = _build_underline_css(underline_headings)

    # ── Nav font sizes (per hierarchy level) ─────────────────────────────
    _default_nav_sizes = {0: "13px", 1: "12px", 2: "11.5px", 3: "11px", 4: "10.5px", 5: "10.5px"}
    _nav_sizes = {**_default_nav_sizes, **(nav_font_sizes or {})}

    def _build_nav_font_css(sizes):
        lines = []
        for level, size in sizes.items():
            weight = "700" if level == 0 else ("600" if level == 1 else ("500" if level == 2 else "400"))
            opacity = f"\n            opacity: {max(0.65, 1.0 - max(0, level - 2) * 0.08):.2f};" if level >= 3 else ""
            lines.append(
                f"        .toc-item[data-level=\"{level}\"] > .toc-header-item > .toc-link,\n"
                f"        .toc-item[data-level=\"{level}\"] > .toc-link {{\n"
                f"            font-size: {size};\n"
                f"            font-weight: {weight};{opacity}\n"
                f"        }}"
            )
        return "\n\n".join(lines)

    _nav_font_css = _build_nav_font_css(_nav_sizes)

    # Page font overrides
    _body_font_family = page_font if page_font else (
        "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif"
    )
    _body_font_size = page_font_size if page_font_size else "14px"

    # max_view patterns (same matching logic as compact_view)
    max_view_patterns = [pattern.lower() for pattern in max_view]

    # Normalize compact patterns for case-insensitive matching
    compact_patterns = [pattern.lower() for pattern in compact_view]

    # Normalize hide_icons patterns (same syntax as compact_view / max_view)
    hide_icons_patterns = [pattern.lower() for pattern in hide_icons]

    # Normalize exclude patterns - store both exact folder names and path patterns
    exclude_patterns = []
    exclude_folder_names = set()
    for pattern in exclude_folders:
        pattern_lower = pattern.lower().replace("\\", "/")
        if "/" in pattern_lower:
            exclude_patterns.append(pattern_lower)
        else:
            exclude_folder_names.add(pattern_lower)

    # Map extension → CSS class name for file icons
    file_colors = {
        ".py": ("python", "#3776ab"),
        ".r": ("r", "#276dc3"),
        ".R": ("r", "#276dc3"),
        ".bas": ("basic", "#8b4513"),
        ".csv": ("csv", "#28a745"),
        ".xlsx": ("csv", "#15b63b"),
        ".xlsm": ("csv", "#11c93c"),
        ".jpg": ("jpg", "#dc3545"),
        ".jpeg": ("jpg", "#dc3545"),
        ".txt": ("txt", "#6c757d"),
        ".json": ("json", "#17a2b8"),
        ".xml": ("xml", "#fd7e14"),
        ".html": ("html", "#e83e8c"),
        ".js": ("js", "#ffc107"),
        ".css": ("css", "#20c997"),
        ".md": ("md", "#6f42c1"),
        ".sql": ("sql", "#e36209"),
    }

    # SVG icon definitions
    FILE_SVG_PATHS = {
        "python": '<path d="M9.86 2c-1.04.01-2.03.09-2.9.26C4.7 2.7 4.32 3.64 4.32 5.24V7h5.6v.67H2.67C1.2 7.67 0 8.67 0 11.33c0 2.68 1.47 3.34 3.33 3.34H4.8v-1.6c0-1.87 1.6-3.74 3.6-3.74h5.6c1.6 0 2.67-1.07 2.67-2.67V5.24c0-1.6-1.07-2.67-2.67-3.01C13.02 2.09 12 2 10.93 2H9.86zM8.8 3.33c.53 0 .93.4.93.94s-.4.93-.93.93c-.54 0-.94-.4-.94-.93s.4-.94.94-.94zM16 8v1.47c0 1.87-1.6 3.6-3.6 3.6H6.8C5.2 13.07 4 14.4 4 16v2.77c0 1.6 1.07 2.4 2.67 2.77.53.13 1.07.2 1.6.2 1.6 0 3.2-.67 3.2-2.24V18H6.4v-.67H16c1.47 0 3.2-1.07 3.2-3.33 0-2.28-1.73-3.34-3.2-3.34H14.4V8H16zm-3.2 8.67c.53 0 .93.4.93.93s-.4.94-.93.94c-.54 0-.94-.41-.94-.94s.4-.93.94-.93z"/>',
        "r": '<path d="M4 2h16a2 2 0 0 1 2 2v16a2 2 0 0 1-2 2H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2zm3 4v12h2v-4h3l2.5 4h2.3L14.2 14A4 4 0 0 0 11 7H7zm2 2h2a2 2 0 0 1 0 4H9V8z"/>',
        "basic": '<path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8l-6-6zm4 18H6V4h7v5h5v11z"/>',
        "csv": '<path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8l-6-6zm4 18H6V4h7v5h5v11zm-8-7h8v2h-8zm0 4h4v2h-4zm0-8h8v2h-8z"/>',
        "jpg": '<path d="M21 19V5c0-1.1-.9-2-2-2H5c-1.1 0-2 .9-2 2v14c0 1.1.9 2 2 2h14c1.1 0 2-.9 2-2zM8.5 13.5l2.5 3.01L14.5 12l4.5 6H5l3.5-4.5z"/>',
        "txt": '<path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8l-6-6zm4 18H6V4h7v5h5v11zm-9-4h6v2H9zm0-4h8v2H9zm0-4h3v2H9z"/>',
        "json": '<path d="M5 3h2v2H5v5a2 2 0 0 1-2 2 2 2 0 0 1 2 2v5h2v2H5c-1.07-.27-2-.9-2-2v-4a2 2 0 0 0-2-2H0v-2h1a2 2 0 0 0 2-2V5a2 2 0 0 1 2-2m14 0c1.07.27 2 .9 2 2v4a2 2 0 0 0 2 2h1v2h-1a2 2 0 0 0-2 2v4a2 2 0 0 1-2 2h-2v-2h2v-5a2 2 0 0 1 2-2 2 2 0 0 1-2-2V5h-2V3h2M12 15a1 1 0 0 1 1 1 1 1 0 0 1-1 1 1 1 0 0 1-1-1 1 1 0 0 1 1-1m-4 0a1 1 0 0 1 1 1 1 1 0 0 1-1 1 1 1 0 0 1-1-1 1 1 0 0 1 1-1m8 0a1 1 0 0 1 1 1 1 1 0 0 1-1 1 1 1 0 0 1-1-1 1 1 0 0 1 1-1z"/>',
        "xml": '<path d="M12 2L2 7l10 5 10-5-10-5zM2 17l10 5 10-5M2 12l10 5 10-5"/>',
        "html": '<path d="M12 17.56L16.07 16.43L16.62 10.33H9.38L9.2 8.3H16.8L17 6.31H7L7.56 12.32H14.45L14.22 14.9L12 15.5L9.78 14.9L9.64 13.24H7.64L7.93 16.43L12 17.56M4.07 3H19.93L18.5 19.2L12 21L5.5 19.2L4.07 3Z"/>',
        "js": '<path d="M3 3h18v18H3V3zm16 16V5H5v14h14zM7.5 15.5l1.41-1.41c.29.29.7.46 1.09.46.68 0 1-.36 1-1.46V8h2v5.12c0 2.17-1.17 3.13-2.89 3.13-.79 0-1.85-.31-2.61-1.07v-.68zm5.71-1.07c.85.5 1.96.85 3.09.85 1.29 0 2.2-.65 2.2-1.72 0-.98-.54-1.56-1.87-2.1l-.65-.28c-1.85-.79-3.07-1.78-3.07-3.88 0-1.93 1.47-3.4 3.77-3.4 1.64 0 2.82.57 3.67 2.06l-2.01 1.29c-.44-.79-.92-1.1-1.66-1.1-.75 0-1.23.5-1.23 1.1 0 .77.47 1.08 1.56 1.56l.65.28c2.18.93 3.41 1.89 3.41 4.04 0 2.32-1.82 3.58-4.27 3.58-2.4 0-3.95-.81-4.71-2.28l2.12-1"/>',
        "css": '<path d="M5 3l-.65 3.34h13.59L17.5 8.5H3.92l-.66 3.33h13.59l-.76 3.81-3.77 1.3-3.78-1.3.26-1.31H6.18l-.52 2.6L12 19.24l6.35-2.14L20 3H5z"/>',
        "md": '<path d="M20.56 18H3.44C2.65 18 2 17.37 2 16.59V7.41C2 6.63 2.65 6 3.44 6h17.12C21.35 6 22 6.63 22 7.41v9.18c0 .78-.65 1.41-1.44 1.41zM6 14.59v-3.76L8 13l2-2.17v3.76h1.5V9.41H10L8 11.56 6 9.41H4.5v5.18H6zm11.44-5.18l-2.08 2.61 2.08 2.57H15.5l-1.14-1.42-1.14 1.42H11.5l2.08-2.57-2.08-2.61H13.5l.86 1.07.86-1.07h2.22z"/>',
        "sql": '<path d="M12 3C7.58 3 4 4.79 4 7v10c0 2.21 3.58 4 8 4s8-1.79 8-4V7c0-2.21-3.58-4-8-4zm6 14c0 .5-2.13 2-6 2s-6-1.5-6-2v-2.23c1.61.78 3.72 1.23 6 1.23s4.39-.45 6-1.23V17zm0-4.55c-1.3.95-3.58 1.55-6 1.55s-4.7-.6-6-1.55V9.64c1.47.83 3.61 1.36 6 1.36s4.53-.53 6-1.36v2.81zM12 9C8.13 9 6 7.5 6 7s2.13-2 6-2 6 1.5 6 2-2.13 2-6 2z"/>',
        "file": '<path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8l-6-6zm4 18H6V4h7v5h5v11z"/>',
    }

    # Expand/collapse icons
    EXPAND_SVG = '<path d="M19 13h-6v6h-2v-6H5v-2h6V5h2v6h6v2z"/>'
    COLLAPSE_SVG = '<path d="M19 13H5v-2h14v2z"/>'

    FOLDER_SVG_PATH = '<path d="M12 2a9 9 0 0 0-9 9v11l3-3 3 3 3-3 3 3 3-3 3 3V11a9 9 0 0 0-9-9m7 15.17-1-1-1.41 1.42L15 19.17l-1.59-1.58L12 16.17l-1.41 1.42L9 19.17l-1.59-1.58L6 16.17l-1 1V11c0-3.86 3.14-7 7-7s7 3.14 7 7v6.17M11 10c0 1.11-.89 2-2 2s-2-.89-2-2 .9-2 2-2 2 .9 2 2m6 0c0 1.11-.89 2-2 2s-2-.89-2-2 .9-2 2-2 2 .9 2 2Z"/>'

    MOON_SVG = '<path d="M12 3a9 9 0 1 0 9 9c0-.46-.04-.92-.1-1.36a5.389 5.389 0 0 1-4.4 2.26 5.403 5.403 0 0 1-3.14-9.8c-.44-.06-.9-.1-1.36-.1z"/>'
    SUN_SVG = '<path d="M12 7a5 5 0 0 1 5 5 5 5 0 0 1-5 5 5 5 0 0 1-5-5 5 5 0 0 1 5-5m0-2a7 7 0 0 0-7 7 7 7 0 0 0 7 7 7 7 0 0 0 7-7 7 7 0 0 0-7-7M2 11h2v2H2v-2m18 0h2v2h-2v-2M11 2h2v2h-2V2m0 18h2v2h-2v-2M4.22 3.93l1.42 1.42-1.42 1.41-1.41-1.41 1.41-1.42m15.14 13.3 1.41 1.41-1.41 1.42-1.42-1.42 1.42-1.41M4.22 19.07l-1.41-1.42 1.41-1.41 1.42 1.41-1.42 1.42M19.36 5.36l-1.42-1.42 1.42-1.41 1.41 1.42-1.41 1.41z"/>'

    def should_exclude_folder(folder_path: Path, root_path: Path) -> bool:
        """Check if a folder should be excluded based on name or path."""
        folder_name = folder_path.name.lower()

        # Check exact folder name match
        if folder_name in exclude_folder_names:
            return True

        # Get the relative path from root
        try:
            rel_path = (
                str(folder_path.relative_to(root_path)).replace("\\", "/").lower()
            )

            # Check each exclude pattern
            for pattern in exclude_patterns:
                pattern_lower = pattern.lower()

                # Exact path match
                if pattern_lower == rel_path:
                    return True

                # Pattern matches this folder as a parent
                if rel_path.startswith(pattern_lower + "/"):
                    return True

                # Pattern matches this folder exactly or as part of path
                if pattern_lower in rel_path.split("/"):
                    return True

        except ValueError:
            pass

        return False

    def read_csv_preview(file_path, max_rows=5):
        """Read the first few rows of a CSV file for preview"""
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                sample = f.read(1024)
                f.seek(0)

                sniffer = csv.Sniffer()
                delimiter = ","
                try:
                    delimiter = sniffer.sniff(sample).delimiter
                except Exception:
                    delimiter = ","

                reader = csv.reader(f, delimiter=delimiter)
                rows = []
                for i, row in enumerate(reader):
                    if i >= max_rows:
                        break
                    row = [
                        str(cell)[:50] + "..." if len(str(cell)) > 50 else str(cell)
                        for cell in row
                    ]
                    rows.append(row)

                return rows
        except Exception as e:
            return [["Error reading file:", str(e)]]

    def read_comments_file(file_path):
        """Parse a .comments file returning dict with title/comments/link/keywords keys.

        Supported keys (case-insensitive, each on its own line):
            title:    Short display title shown in compact view / card caption
            link:     URL or file path — wraps the title/caption in a hyperlink
            keywords: Space- or comma-separated extra search terms appended to
                      data-search so items surface even when the keyword isn't
                      in the filename or folder name
            comments: Free-form description text (may span multiple lines)
        """
        result = {"title": "", "comments": "", "link": "", "keywords": ""}
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
                lines = content.split("\n")
                in_comments = False
                for line in lines:
                    ll = line.lower()
                    if ll.startswith("title:"):
                        result["title"] = line[6:].strip()
                        in_comments = False
                    elif ll.startswith("link:"):
                        result["link"] = line[5:].strip()
                        in_comments = False
                    elif ll.startswith("keywords:"):
                        kw_raw = line[9:].strip()
                        # normalise: replace commas with spaces, collapse whitespace
                        result["keywords"] = " ".join(kw_raw.replace(",", " ").split())
                        in_comments = False
                    elif ll.startswith("comments:"):
                        result["comments"] = line[9:].strip()
                        in_comments = True
                    elif in_comments:
                        result["comments"] += "\n" + line.rstrip()
        except Exception:
            pass
        return result

    def read_folder_comments(folder_path):
        """Check for a .comments file in the folder and parse it."""
        comments_file = Path(folder_path) / ".comments"
        if comments_file.exists():
            return read_comments_file(comments_file)
        return None

    def matches_compact_pattern(
        folder_path: str, root_path: Path, patterns: List[str]
    ) -> bool:
        """Check if a folder path matches any compact view pattern."""
        if not folder_path:
            return False

        rel_path = folder_path.replace("\\", "/").lower()

        for pattern in patterns:
            pattern_lower = pattern.lower()

            # Exact match
            if pattern_lower == rel_path:
                return True

            # Wildcard pattern for any folder name anywhere
            if pattern_lower.startswith("*/"):
                folder_name = pattern_lower[2:]
                if rel_path.endswith("/" + folder_name) or rel_path == folder_name:
                    return True

            # Wildcard at the end
            if pattern_lower.endswith("/*"):
                prefix = pattern_lower[:-2]
                if rel_path.startswith(prefix + "/") or rel_path == prefix:
                    return True

            # Simple wildcard anywhere in the path
            if "*" in pattern_lower:
                pattern_regex = pattern_lower.replace(".", "\\.").replace("*", ".*")
                if re.match(pattern_regex + "$", rel_path):
                    return True

        return False

    def make_file_svg(class_name, size=16, extra_class=""):
        """Return an inline SVG for a given file-type class name."""
        path_d = FILE_SVG_PATHS.get(class_name, FILE_SVG_PATHS["file"])
        cls = f"file-type-svg {extra_class}".strip()
        return (
            f'<svg class="{cls}" data-type="{class_name}" '
            f'viewBox="0 0 24 24" width="{size}" height="{size}" '
            f'xmlns="http://www.w3.org/2000/svg">{path_d}</svg>'
        )

    def make_folder_svg(size=14):
        return (
            f'<svg class="folder-svg" viewBox="0 0 24 24" '
            f'width="{size}" height="{size}" xmlns="http://www.w3.org/2000/svg">'
            f"{FOLDER_SVG_PATH}</svg>"
        )

    def file_icon_link(file_info, item, file_colors, csv_previews):
        """Return the <a> tag for a single associated file (SVG icon)."""
        ext = file_info["extension"]
        class_name = file_colors.get(ext, ("file", ""))[0]

        if file_info["name"].startswith(item["basename"]):
            display_name = ext if ext else file_info["name"]
        else:
            display_name = file_info["name"]

        svg_icon = make_file_svg(class_name)
        title = f"{display_name} ({file_info['name']})"

        link = (
            f'<a href="{file_info["path"]}" class="file-link {class_name}" '
            f'target="_blank" data-file-path="{file_info["path"]}" title="{title}">'
            f"{svg_icon}"
        )
        if ext == ".csv" and file_info["path"] in csv_previews:
            link += """
                        <div class="csv-preview"><div class="csv-table-container"></div></div>"""
        link += "</a>"
        return link

    def scan_folder_recursive(
        folder_path: Path,
        root_path: Path,
        current_depth: int = 0,
        folder_comments: dict = None,
    ) -> Dict[str, Any]:
        """Recursively scan folders and return nested structure."""
        if current_depth >= max_depth:
            return {}

        result = {}
        items_by_basename = defaultdict(list)
        comments_by_basename = {}
        folder_level_comments = folder_comments or read_folder_comments(folder_path)

        try:
            for file_path in folder_path.iterdir():
                if file_path.is_file():
                    basename = file_path.stem
                    extension = file_path.suffix.lower()

                    if extension == ".comments":
                        comments_by_basename[basename] = read_comments_file(file_path)
                        continue

                    file_info = {
                        "name": file_path.name,
                        "extension": extension,
                        "path": str(file_path.relative_to(root_path)).replace(
                            "\\", "/"
                        ),
                    }
                    items_by_basename[basename].append(file_info)
        except PermissionError:
            pass

        gallery_items = []
        for basename, files in items_by_basename.items():
            image_file = None
            other_files = []

            for file_info in files:
                if file_info["extension"] in image_extensions:
                    if image_file is None or image_extensions.index(
                        file_info["extension"]
                    ) < image_extensions.index(image_file["extension"]):
                        if image_file:
                            other_files.append(image_file)
                        image_file = file_info
                    else:
                        other_files.append(file_info)
                else:
                    other_files.append(file_info)

            if image_file:
                other_files.sort(key=lambda x: x["extension"])
                folder_relative_path = str(folder_path.relative_to(root_path)).replace(
                    "\\", "/"
                )

                gallery_items.append(
                    {
                        "basename": basename,
                        "image": image_file,
                        "other_files": other_files,
                        "folder_path": folder_relative_path,
                        "comments": comments_by_basename.get(basename, None),
                    }
                )

        gallery_items.sort(key=lambda x: x["basename"])

        if gallery_items:
            result["_items"] = gallery_items

        if folder_level_comments:
            result["_comments"] = folder_level_comments

        try:
            for subfolder in folder_path.iterdir():
                if subfolder.is_dir():
                    if should_exclude_folder(subfolder, root_path):
                        print(
                            f"  📁 Excluding folder: {subfolder.relative_to(root_path)}"
                        )
                        continue

                    subfolder_result = scan_folder_recursive(
                        subfolder, root_path, current_depth + 1
                    )
                    if subfolder_result:
                        result[subfolder.name] = subfolder_result
        except PermissionError:
            pass

        return result

    def sentence_case(name):
        """Convert snake_case or any name to sentence case."""
        words = name.replace("_", " ").replace("-", " ").split()
        if not words:
            return name
        return words[0].capitalize() + (
            " " + " ".join(words[1:]) if len(words) > 1 else ""
        )

    def generate_toc_html(
        node: Dict[str, Any],
        current_path: List[str],
        level: int = 0,
        parent_id: str = None,
    ) -> Tuple[str, List[str]]:
        """Generate hierarchical TOC HTML with minimal, consistent alignment."""
        html = ""
        folder_ids = []

        for key, value in sorted(node.items()):
            if key in ["_items", "_comments"]:
                continue

            if not value.get("_items") and not any(
                k not in ["_items", "_comments"] for k in value.keys()
            ):
                continue

            folder_id = (
                "-".join(current_path + [key])
                .lower()
                .replace(" ", "-")
                .replace("_", "-")
            )
            folder_ids.append(folder_id)
            display_name = resolve_display_name(key, current_path, sentence_case(key))

            has_children = any(k not in ["_items", "_comments"] for k in value.keys())

            base_indent = 3
            level_indent = level * 15
            # Total icon footprint: margin-left(4) + icon(10) + margin-right(2) = 16px
            # Plus toc-link's own padding-left(4px) inside the row = 20px
            icon_offset = 20

            if has_children:
                html += f'''
        <div class="toc-item" data-level="{level}" data-nav-id="{folder_id}">
            <div class="toc-header-item" style="padding-left: {base_indent + level_indent}px;" onclick="handleNavRowClick(event, '{folder_id}')">
                <svg class="toggle-icon" id="toggle-icon-{folder_id}" viewBox="0 0 24 24" width="10" height="10">
                    {COLLAPSE_SVG}
                </svg>
                <a href="#{folder_id}" class="toc-link">{display_name}</a>
            </div>
            <div class="toc-children" id="toc-children-{folder_id}">'''

                children_html, child_ids = generate_toc_html(
                    value, current_path + [key], level + 1, folder_id
                )
                html += children_html
                folder_ids.extend(child_ids)

                html += """
            </div>
        </div>"""
            else:
                html += f'''
        <div class="toc-item" data-level="{level}" data-nav-id="{folder_id}">
            <a href="#{folder_id}" class="toc-link" style="padding-left: {base_indent + level_indent + icon_offset}px;">{display_name}</a>
        </div>'''

        return html, folder_ids

    def generate_content_html(
        node: Dict[str, Any],
        current_path: List[str],
        level: int,
        csv_previews: Dict,
        compact_patterns: List[str],
        max_view_patterns: List[str],
        hide_icons_patterns: List[str],
        root_path: Path,
        compact_thumb_size: int,
        compact_column_ratio: List[int],
        card_image_height: Optional[int],
        file_colors: Dict,
    ) -> str:
        """Generate HTML content for a folder and its subfolders recursively."""
        html = ""

        items = node.get("_items", [])
        folder_comments = node.get("_comments")

        if items:
            folder_id = (
                "-".join(current_path).lower().replace(" ", "-").replace("_", "-")
                if current_path
                else "root"
            )
            if current_path:
                _default_disp = sentence_case(current_path[-1])
                display_name = resolve_display_name(current_path[-1], current_path[:-1], _default_disp)
            else:
                display_name = resolve_display_name("", [], "Root")

            heading_level = min(level + 1, 6)
            html += f'\n<h{heading_level} id="{folder_id}" data-section="{folder_id}" class="section-heading">'
            html += display_name
            if folder_comments and folder_comments.get("title"):
                html += f' <span class="section-title-comment">{folder_comments["title"]}</span>'
            html += f"</h{heading_level}>\n"

            if folder_comments and folder_comments.get("comments"):
                html += f'<div class="section-description" data-section="{folder_id}">{folder_comments["comments"]}</div>\n'

            folder_rel_path = "/".join(current_path) if current_path else ""
            is_compact = matches_compact_pattern(
                folder_rel_path, root_path, compact_patterns
            )
            is_max = matches_compact_pattern(
                folder_rel_path, root_path, max_view_patterns
            )
            is_hide_icons = matches_compact_pattern(
                folder_rel_path, root_path, hide_icons_patterns
            )

            if is_compact:
                html += f'<div class="compact-list" data-section="{folder_id}">\n'

                for item in items:
                    image_info = item["image"]
                    alt_text = sentence_case(item["basename"])
                    folder_path = item["folder_path"]
                    comments = item.get("comments")

                    all_files = (
                        [image_info["name"]]
                        + [f["name"] for f in item["other_files"]]
                        + ["folder"]
                    )
                    _kw = (comments or {}).get("keywords", "")
                    _title_kw = (comments or {}).get("title", "")
                    search_data = " ".join(
                        all_files + [folder_rel_path] + [item["basename"]]
                        + ([_kw] if _kw else []) + ([_title_kw] if _title_kw else [])
                    ).lower()

                    has_comments = comments and (
                        comments.get("title") or comments.get("comments")
                    )
                    row_class = (
                        "compact-row" if has_comments else "compact-row no-comments"
                    )

                    html += f'''
            <div class="{row_class}" data-search="{search_data}" data-section="{folder_id}">
                <img class="compact-thumb" src="{image_info["path"]}" alt="{alt_text}"
                     onclick="openModal('{image_info["path"]}', '{alt_text}')">'''

                    if not is_hide_icons:
                        html += f'''
                <div class="compact-icons">
                    <a href="{folder_path}" class="file-link folder-link" target="_blank" title="Open folder">
                        {make_folder_svg(16)}
                    </a>'''

                        for file_info in item["other_files"]:
                            html += file_icon_link(
                                file_info, item, file_colors, csv_previews
                            )

                        html += "\n                </div>"

                    if has_comments:
                        # Build title — wrap in link if comments has a link:
                        _link = comments.get("link", "").strip()
                        if comments.get("title"):
                            _title_text = comments["title"]
                            if _link:
                                title_html = f'<div class="compact-comments-title"><a href="{_link}" target="_blank" class="comments-link">{_title_text}</a></div>'
                            else:
                                title_html = f'<div class="compact-comments-title">{_title_text}</div>'
                        else:
                            title_html = ""
                        body_text = (
                            comments.get("comments", "")
                            .replace("<", "&lt;")
                            .replace(">", "&gt;")
                        )
                        html += f"""
                <div class="compact-comments">
                    {title_html}
                    <div class="compact-comments-body">{body_text}</div>
                </div>"""
                    else:
                        html += '\n                <div class="compact-comments"></div>'

                    html += "\n            </div>"

                html += "</div>\n"

            else:
                # Normal card view — max_view uses wider minmax
                _minmax = f"{max_view_thumb_size}px" if is_max else "250px"
                _extra_class = " max-view-container" if is_max else ""
                html += f'<div class="gallery-container{_extra_class}" style="grid-template-columns: repeat(auto-fill, minmax({_minmax}, 1fr));" data-section="{folder_id}">\n'

                for item in items:
                    image_info = item["image"]
                    alt_text = item["basename"].replace("_", " ").title()
                    folder_path = item["folder_path"]
                    comments = item.get("comments")

                    all_files = (
                        [image_info["name"]]
                        + [f["name"] for f in item["other_files"]]
                        + ["folder"]
                    )
                    _kw = (comments or {}).get("keywords", "")
                    _title_kw = (comments or {}).get("title", "")
                    search_data = " ".join(
                        all_files + [folder_rel_path] + [item["basename"]]
                        + ([_kw] if _kw else []) + ([_title_kw] if _title_kw else [])
                    ).lower()

                    height_style = (
                        f"height: {card_image_height}px;" if card_image_height else ""
                    )

                    # Caption: use comments title or filename; wrap in link if comments has link:
                    _cap_link = (comments or {}).get("link", "").strip()
                    _cap_title = (comments or {}).get("title", "").strip()
                    _cap_text = _cap_title if _cap_title else image_info["name"]
                    if _cap_link:
                        caption_html = f'<a href="{_cap_link}" target="_blank" class="comments-link caption-link">{_cap_text}</a>'
                    else:
                        caption_html = _cap_text

                    html += f'''
            <div class="gallery-item" data-search="{search_data}" data-section="{folder_id}">
                <div class="image-container" style="{height_style}" onclick="openModal('{image_info["path"]}', '{alt_text}')">
                    <img src="{image_info["path"]}" alt="{alt_text}" style="width: 100%;">
                </div>
                <div class="image-caption">
                    {caption_html}
                </div>'''

                    if not is_hide_icons:
                        html += f'''
                <div class="file-list">
                    <a href="{folder_path}" class="file-link folder-link" target="_blank" title="Open folder">
                        {make_folder_svg(14)}
                    </a>'''

                        for file_info in item["other_files"]:
                            html += file_icon_link(
                                file_info, item, file_colors, csv_previews
                            )

                        html += """
                </div>"""

                    html += """
            </div>"""

                html += "</div>\n"

        for key, value in sorted(node.items()):
            if key in ["_items", "_comments"]:
                continue

            if not value.get("_items") and not any(
                k not in ["_items", "_comments"] for k in value.keys()
            ):
                continue

            html += generate_content_html(
                value,
                current_path + [key],
                level + 1,
                csv_previews,
                compact_patterns,
                max_view_patterns,
                hide_icons_patterns,
                root_path,
                compact_thumb_size,
                compact_column_ratio,
                card_image_height,
                file_colors,
            )

        return html

    def generate_html(
        gallery_data: Dict[str, Any],
        csv_previews: Dict,
        root_path: Path,
        compact_patterns: List[str],
        max_view_patterns: List[str],
        hide_icons_patterns: List[str],
    ) -> str:
        """Generate complete HTML with hierarchical structure."""

        toc_html, folder_ids = generate_toc_html(gallery_data, [])

        main_content_html = generate_content_html(
            gallery_data,
            [],
            0,
            csv_previews,
            compact_patterns,
            max_view_patterns,
            hide_icons_patterns,
            root_path,
            compact_thumb_size,
            compact_column_ratio,
            card_image_height,
            file_colors,
        )

        thumb_height = round(compact_thumb_size * 9 / 16)

        # Pre-compute optional page header and footer HTML
        if page_title:
            _subtitle_html = f'<div class="page-header-subtitle">{page_subtitle}</div>' if page_subtitle else ""
            _desc_html = f'<div class="page-header-description">{page_description}</div>' if page_description else ""
            page_header_html = f'<div class="page-header"><div class="page-header-title">{page_title}</div>{_subtitle_html}{_desc_html}</div>'
        else:
            page_header_html = ""
        page_footer_html = f'<div class="page-footer">{footer_content}</div>' if footer_content else ""

        collapse_js = f"""
        let collapsedStates = new Set();

        function initCollapsible() {{
            const allFolders = {json.dumps(folder_ids)};
            allFolders.forEach(folderId => {{
                const children = document.getElementById(`toc-children-${{folderId}}`);
                const icon = document.getElementById(`toggle-icon-${{folderId}}`);
                if (children && icon) {{
                    children.style.display = 'block';
                    icon.innerHTML = `{COLLAPSE_SVG}`;
                }}
            }});
        }}

        function toggleSection(folderId) {{
            const children = document.getElementById(`toc-children-${{folderId}}`);
            const icon = document.getElementById(`toggle-icon-${{folderId}}`);

            if (children && icon) {{
                if (children.style.display === 'none') {{
                    children.style.display = 'block';
                    icon.innerHTML = `{COLLAPSE_SVG}`;
                    collapsedStates.delete(folderId);
                }} else {{
                    children.style.display = 'none';
                    icon.innerHTML = `{EXPAND_SVG}`;
                    collapsedStates.add(folderId);
                }}
                try {{
                    localStorage.setItem('toc-collapsed', JSON.stringify(Array.from(collapsedStates)));
                }} catch(e) {{
                    console.log('Could not save to localStorage');
                }}
            }}
        }}

        // Row click handler: if the click landed on the <a> link, let it navigate
        // (smooth scroll via the DOMContentLoaded listener). If it landed anywhere
        // else on the row (icon, padding), toggle the section.
        function handleNavRowClick(event, folderId) {{
            if (event.target.closest('a.toc-link')) {{
                // Let the anchor's own click handler (smooth scroll) fire — don't toggle
                return;
            }}
            event.preventDefault();
            toggleSection(folderId);
        }}

        function restoreCollapsedStates() {{
            try {{
                const saved = localStorage.getItem('toc-collapsed');
                if (saved) {{
                    const collapsed = JSON.parse(saved);
                    collapsed.forEach(folderId => {{
                        const children = document.getElementById(`toc-children-${{folderId}}`);
                        const icon = document.getElementById(`toggle-icon-${{folderId}}`);
                        if (children && icon) {{
                            children.style.display = 'none';
                            icon.innerHTML = `{EXPAND_SVG}`;
                            collapsedStates.add(folderId);
                        }}
                    }});
                }}
            }} catch(e) {{
                console.log('Could not restore from localStorage');
            }}
        }}

        function collapseAll() {{
            const allFolders = {json.dumps(folder_ids)};
            allFolders.forEach(folderId => {{
                const children = document.getElementById(`toc-children-${{folderId}}`);
                const icon = document.getElementById(`toggle-icon-${{folderId}}`);
                if (children && icon) {{
                    children.style.display = 'none';
                    icon.innerHTML = `{EXPAND_SVG}`;
                    collapsedStates.add(folderId);
                }}
            }});
            try {{ localStorage.setItem('toc-collapsed', JSON.stringify(Array.from(collapsedStates))); }} catch(e) {{}}
        }}

        function expandAll() {{
            const allFolders = {json.dumps(folder_ids)};
            allFolders.forEach(folderId => {{
                const children = document.getElementById(`toc-children-${{folderId}}`);
                const icon = document.getElementById(`toggle-icon-${{folderId}}`);
                if (children && icon) {{
                    children.style.display = 'block';
                    icon.innerHTML = `{COLLAPSE_SVG}`;
                    collapsedStates.delete(folderId);
                }}
            }});
            try {{ localStorage.setItem('toc-collapsed', JSON.stringify([])); }} catch(e) {{}}
        }}

        function initScrollSpy() {{
            const sections = document.querySelectorAll('.section-heading');
            const navLinks = document.querySelectorAll('.toc-link');
            const navDots = document.querySelectorAll('.sidebar-nav-dot);
            
            if (sections.length === 0) return;
            
            const observer = new IntersectionObserver(
                (entries) => {{
                    entries.forEach(entry => {{
                        const id = entry.target.getAttribute('id');
                        
                        if (entry.isIntersecting) {{
                            // Don't apply active highlight while user is searching
                            const searchTerm = document.getElementById('gallery-search').value.trim();
                            if (searchTerm) return;

                            navLinks.forEach(link => {{
                                link.classList.remove('active');
                            }});
                            
                            const activeLink = document.querySelector(`.toc-link[href="#${{id}}"]`);
                            if (activeLink) {{
                                activeLink.classList.add('active');
                                activeLink.scrollIntoView({{
                                    block: 'nearest',
                                    behavior: 'smooth'
                                }});
                            }}

                            // Update nav dots
                            navDots.forEach(function(d) {{ d.classList.remove('active'); }});
                            var activeDot = document.querySelector('.sidebar-nav-dot[data-dot-section="' + id + '"]');
                            if (activeDot) activeDot.classList.add('active');
                        }}
                    }});
                }},
                {{
                    root: null,
                    rootMargin: '-30% 0px -60% 0px',
                    threshold: 0
                }}
            );
            
            sections.forEach(section => observer.observe(section));
        }}
        """

        html_template = f"""<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{page_title if page_title else navigation_title}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        :root {{
            --bg-color: {light_mode_colors["bg"]};
            --text-color: {light_mode_colors["text"]};
            --sidebar-bg: {light_mode_colors["sidebar_bg"]};
            --accent-color: {light_mode_colors["accent"]};
            --border-color: {light_mode_colors["border"]};
            --hover-color: {light_mode_colors["hover"]};
            --modal-bg: {light_mode_colors["modal_bg"]};
        }}

        [data-theme="dark"] {{
            --bg-color: {dark_mode_colors["bg"]};
            --text-color: {dark_mode_colors["text"]};
            --sidebar-bg: {dark_mode_colors["sidebar_bg"]};
            --accent-color: {dark_mode_colors["accent"]};
            --border-color: {dark_mode_colors["border"]};
            --hover-color: {dark_mode_colors["hover"]};
            --modal-bg: {dark_mode_colors["modal_bg"]};
        }}

        body {{
            font-family: {_body_font_family};
            font-size: {_body_font_size};
            background-color: var(--bg-color);
            color: var(--text-color);
            padding-left: {'40px' if sidebar_collapsible else side_bar_width};
            min-height: 100vh;
            transition: background-color 0.3s, color 0.3s;
        }}

        body::-webkit-scrollbar {{
            display: none;
        }}

        body {{
            -ms-overflow-style: none;
            scrollbar-width: none;
        }}

        /* ── Modern slim scrollbars (global) ── */
        * {{
            scrollbar-width: thin;
            scrollbar-color: rgba(128,128,128,0.35) transparent;
        }}
        *::-webkit-scrollbar {{ width: 6px; height: 6px; }}
        *::-webkit-scrollbar-track {{ background: transparent; }}
        *::-webkit-scrollbar-thumb {{
            background: rgba(128,128,128,0.35);
            border-radius: 3px;
        }}
        *::-webkit-scrollbar-thumb:hover {{
            background: rgba(128,128,128,0.6);
        }}
        [data-theme="dark"] *::-webkit-scrollbar-thumb {{
            background: rgba(200,200,200,0.25);
        }}
        [data-theme="dark"] *::-webkit-scrollbar-thumb:hover {{
            background: rgba(200,200,200,0.5);
        }}
        [data-theme="dark"] * {{
            scrollbar-color: rgba(200,200,200,0.25) transparent;
        }}

        .toc-sidebar {{
            position: fixed;
            top: 0;
            left: 0;
            width: {side_bar_width};
            height: 100vh;
            background-color: var(--sidebar-bg);
            padding: 20px 10px;
            box-shadow: 2px 0 5px rgba(0, 0, 0, 0.1);
            overflow-y: auto;
            z-index: 1000;
            transition: background-color 0.3s;
            /* Hover-reveal scrollbar for sidebar */
            scrollbar-width: thin;
            scrollbar-color: transparent transparent;
            transition: scrollbar-color 0.3s ease;
        }}

        .toc-sidebar:hover {{
            scrollbar-color: rgba(120, 120, 120, 0.3) transparent;
        }}

        .toc-sidebar::-webkit-scrollbar {{
            width: 8px;
        }}

        .toc-sidebar::-webkit-scrollbar-track {{
            background: transparent;
        }}

        .toc-sidebar::-webkit-scrollbar-thumb {{
            background: transparent;
            border-radius: 10px;
        }}

        .toc-sidebar:hover::-webkit-scrollbar-thumb {{
            background: rgba(120, 120, 120, 0.3);
        }}

        .toc-sidebar:hover::-webkit-scrollbar-thumb:hover {{
            background: rgba(120, 120, 120, 0.5);
        }}

        [data-theme="dark"] .toc-sidebar:hover {{
            scrollbar-color: rgba(200, 200, 200, 0.2) transparent;
        }}

        [data-theme="dark"] .toc-sidebar:hover::-webkit-scrollbar-thumb {{
            background: rgba(200, 200, 200, 0.2);
        }}

        [data-theme="dark"] .toc-sidebar:hover::-webkit-scrollbar-thumb:hover {{
            background: rgba(200, 200, 200, 0.35);
        }}


        /* — collapsible sidebar: floating nav rail + popover panel — */
        .sidebar-collapsed-strip {{
            display: {'flex' if sidebar_collapsible else 'none'};
            position: fixed;
            top: 0;
            left: 0;
            width: 40px;
            height: 100vh;
            flex-direction: column;
            align-items: center;
            padding-top: 14px;
            z-index: 1001;
            pointer-events: none;
        }}

        .sidebar-collapsed-strip > * {{
            pointer-events: auto;
        }}

        .sidebar-contents-badge {{
            background: var(--sidebar-bg);
            color: var(--text-color);
            font-size: 9px;
            font-weight: 700;
            letter-spacing: 1.2px;
            text-transform: uppercase;
            padding: 6px 4px;
            border-radius: 4px;
            cursor: pointer;
            user-select: none;
            writing-mode: vertical-rl;
            text-orientation: mixed;
            transform: rotate(180deg);
            border: 1px solid var(--border-color);
            opacity: 0.85;
            transition: opacity 0.2s, background 0.2s;
            margin-bottom: 8px;
        }}

        .sidebar-contents-badge:hover {{
            opacity: 1;
            background: var(--hover-color);
        }}

        .sidebar-nav-dots {{
            display: flex;
            flex-direction: column;
            align-items: center;
            gap: 6px;
            margin-top: 4px;
        }}

        .sidebar-nav-dot {{
            width: 7px;
            height: 7px;
            border-radius: 50%;
            background: var(--border-color);
            cursor: pointer;
            transition: background 0.2s, transform 0.2s;
            position: relative;
        }}

        .sidebar-nav-dot:hover {{
            background: var(--accent-color);
            transform: scale(1.4);
        }}

        .sidebar-nav-dot.active {{
            background: var(--accent-color);
            transform: scale(1.3);
        }}

        .sidebar-nav-dot-tooltip {{
            display: none;
            position: absolute;
            left: 18px;
            top: 50%;
            transform: translateY(-50%);
            background: var(--sidebar-bg);
            color: var(--text-color);
            border: 1px solid var(--border-color);
            border-radius: 4px;
            padding: 3px 8px;
            font-size: 11px;
            white-space: nowrap;
            pointer-events: none;
            z-index: 1003;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
        }}

        .sidebar-nav-dot:hover .sidebar-nav-dot-tooltip {{
            display: block;
        }}

        .sidebar-backdrop {{
            display: none;
            position: fixed;
            inset: 0;
            z-index: 999;
        }}

        .sidebar-backdrop.active {{
            display: block;
        }}

        .sidebar.collapsible {{
            position: fixed;
            top: 10px;
            left: 44px;
            width: min({side_bar_width}, calc(100vw - 60px));
            height: auto;
            max-height: calc(100vh - 20px);
            border-radius: 10px;
            border: 1px solid var(--border-color);
            box-shadow: 0 8px 32px rgba(0,0,0,0.18);
            transform: scale(0.95);
            opacity: 0;
            pointer-events: none;
            z-index: 1002;
            transition: transform 0.2s cubic-bezier(0.4,0,0.2,1), opacity 0.2s;
            overflow-y: auto;
        }}

        .sidebar.collapsible.open {{
            transform: scale(1);
            opacity: 1;
            pointer-events: auto;
        }}
                
        .toc-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
        }}

        .toc-title {{
            font-size: 14px;
            font-weight: bold;
            color: var(--accent-color);
        }}

        /* ── Search box with buttons below ── */
        .search-box {{
            margin-bottom: 12px;
        }}

        .search-wrap {{
            position: relative;
            display: flex;
            align-items: center;
        }}

        .search-input {{
            flex: 1;
            min-width: 0;
            width: 100%;
            padding: 7px 26px 7px 10px;
            border: 1px solid var(--border-color);
            border-radius: 4px;
            font-size: 12px;
            background-color: var(--bg-color);
            color: var(--text-color);
            transition: all 0.2s;
        }}

        .search-input:focus {{
            outline: none;
            border-color: var(--accent-color);
            box-shadow: 0 0 0 2px rgba(0, 123, 255, 0.12);
        }}

        .search-input::placeholder {{
            color: #999;
            font-size: 11px;
        }}

        .search-clear {{
            position: absolute;
            right: 6px;
            background: none;
            border: none;
            cursor: pointer;
            color: var(--text-color);
            font-size: 11px;
            opacity: 0;
            padding: 2px 3px;
            border-radius: 3px;
            transition: opacity 0.15s;
            line-height: 1;
            z-index: 1;
        }}
        .search-clear.visible {{ opacity: 0.5; }}
        .search-clear:hover {{ opacity: 1 !important; }}

        /* ── Icon group: below the search input ── */
        .search-icon-group {{
            display: flex;
            align-items: center;
            gap: 2px;
            flex-shrink: 0;
            margin-top: 4px;
        }}

        .search-icon-btn {{
            background: none;
            border: none;
            color: var(--text-color);
            cursor: pointer;
            padding: 4px 5px;
            border-radius: 4px;
            opacity: 0.5;
            transition: opacity 0.15s, background 0.15s;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            line-height: 1;
        }}
        .search-icon-btn:hover {{
            opacity: 1;
            background: var(--hover-color);
        }}

        /* ── Help tooltip — opens below the search row ── */
        .help-tooltip {{
            display: none;
            background: var(--sidebar-bg);
            border: 1px solid var(--border-color);
            border-radius: 6px;
            padding: 8px 10px;
            margin-top: 6px;
            font-size: 11px;
            color: var(--text-color);
        }}
        .help-tooltip.visible {{ display: block; }}
        .help-row {{
            display: flex;
            align-items: center;
            gap: 5px;
            margin: 3px 0;
        }}
        .help-muted {{ opacity: 0.55; }}
        kbd {{
            display: inline-block;
            padding: 1px 5px;
            border: 1px solid var(--border-color);
            border-radius: 3px;
            font-size: 10px;
            font-family: monospace;
            background: var(--hover-color);
            color: var(--text-color);
            line-height: 1.5;
        }}

        .search-count {{
            font-size: 10px;
            color: var(--text-color);
            margin-top: 4px;
            display: block;
            opacity: 0.6;
        }}

        .toc-links {{
            display: flex;
            flex-direction: column;
            gap: {toc_link_gap}px;
        }}

        .toc-item {{
            margin: 0;
            width: 100%;
        }}

        .toc-item.nav-hidden {{
            display: none;
        }}

        .toc-header-item {{
            display: flex;
            align-items: center;
            cursor: pointer;
            user-select: none;
            border-radius: 4px;
            width: 100%;
            transition: background-color 0.15s;
        }}

        .toc-header-item:hover {{
            background-color: var(--hover-color);
        }}

        .toggle-icon {{
            fill: var(--text-color);
            opacity: 0.5;
            width: 10px;
            height: 10px;
            margin-left: 4px;
            margin-right: 2px;
            flex-shrink: 0;
            transition: opacity 0.15s;
        }}

        .toc-header-item:hover .toggle-icon {{
            opacity: 0.9;
        }}

        .toc-children {{
            margin-left: 0;
        }}

        .toc-link {{
            color: var(--text-color);
            text-decoration: none;
            padding: 6px 10px 6px 4px;
            border-radius: 4px;
            border-left: 2px solid transparent;
            display: block;
            transition: all 0.15s;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            flex: 1;
            min-width: 0;
        }}

        /* Hierarchical nav font sizes — controlled by nav_font_sizes parameter */
        __NAV_FONT_CSS_PLACEHOLDER__

        .toc-link:hover {{
            color: var(--accent-color);
        }}

        .toc-header-item:hover .toc-link {{
            color: var(--accent-color);
        }}

        .toc-link.active {{
            color: var(--accent-color);
            font-weight: 600;
            border-left-color: var(--accent-color);
        }}

        .main-content {{
            padding: 20px 30px 30px 30px;
            min-height: 100vh;
        }}

        .gallery-container {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(250px, 1fr));
            gap: 20px;
            padding: 10px;
            max-width: 1400px;
            margin: 0 auto 30px auto;
        }}

        .gallery-item {{
            position: relative;
            border-radius: 8px;
            overflow: hidden;
            background-color: var(--sidebar-bg);
            border: 1px solid var(--border-color);
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
            transition: transform 0.3s ease, box-shadow 0.3s ease, opacity 0.3s ease;
        }}

        .gallery-item.hidden {{
            display: none;
        }}

        .gallery-item:hover {{
            transform: translateY(-5px);
            box-shadow: 0 4px 16px rgba(0, 0, 0, 0.2);
        }}

        .image-container {{
            position: relative;
            width: 100%;
            cursor: pointer;
            background-color: var(--bg-color);
            overflow: hidden;
        }}

        .gallery-item img {{
            width: 100%;
            height: 100%;
            object-fit: cover;
            display: block;
        }}

        .image-caption {{
            padding: 12px;
            background-color: var(--sidebar-bg);
            font-size: 12px;
            color: var(--text-color);
            text-align: center;
            font-weight: 500;
            word-break: break-word;
            border-top: 1px solid var(--border-color);
        }}

        .file-list {{
            padding: 10px;
            background-color: var(--sidebar-bg);
            border-top: 1px solid var(--border-color);
            display: flex;
            flex-wrap: wrap;
            gap: 5px;
        }}

        .file-link {{
            display: inline-flex;
            align-items: center;
            justify-content: center;
            padding: 4px;
            color: var(--text-color);
            text-decoration: none;
            border-radius: 4px;
            font-size: 12px;
            transition: opacity 0.2s ease, background 0.2s ease;
            position: relative;
        }}

        .file-link:hover {{
            background-color: var(--hover-color);
            opacity: 1;
        }}

        .file-type-svg {{
            fill: var(--text-color);
            opacity: 0.55;
            transition: opacity 0.2s, fill 0.2s;
            display: block;
        }}

        .file-link:hover .file-type-svg {{
            opacity: 1;
        }}

        .file-link.python:hover .file-type-svg {{ fill: #3776ab; opacity: 1; }}
        .file-link.r:hover .file-type-svg {{ fill: #276dc3; opacity: 1; }}
        .file-link.html:hover .file-type-svg {{ fill: #e83e8c; opacity: 1; }}
        .file-link.js:hover .file-type-svg {{ fill: #f0c000; opacity: 1; }}
        .file-link.css:hover .file-type-svg {{ fill: #20c997; opacity: 1; }}
        .file-link.csv:hover .file-type-svg {{ fill: #28a745; opacity: 1; }}
        .file-link.json:hover .file-type-svg {{ fill: #17a2b8; opacity: 1; }}
        .file-link.xml:hover .file-type-svg {{ fill: #fd7e14; opacity: 1; }}
        .file-link.txt:hover .file-type-svg {{ fill: #6c757d; opacity: 1; }}
        .file-link.md:hover .file-type-svg {{ fill: #6f42c1; opacity: 1; }}
        .file-link.sql:hover .file-type-svg {{ fill: #e36209; opacity: 1; }}
        .file-link.jpg:hover .file-type-svg {{ fill: #dc3545; opacity: 1; }}

        .file-link.folder-link {{
            display: inline-flex;
            align-items: center;
            gap: 4px;
        }}

        .folder-svg {{
            fill: var(--text-color);
            opacity: 0.55;
            transition: fill 0.2s, opacity 0.2s;
        }}

        .file-link.folder-link:hover .folder-svg {{
            fill: #c23899;
            opacity: 1;
        }}

        .compact-list {{
            max-width: 1400px;
            margin: 0 auto 24px auto;
            padding: 0 10px;
            border: 1px solid var(--border-color);
            border-radius: 8px;
            overflow: hidden;
        }}

        .compact-row {{
            display: grid;
            grid-template-columns: {compact_thumb_size + 10}px {compact_column_ratio[1]}% 1fr;
            align-items: center;
            gap: 12px;
            padding: 6px 12px;
            border-bottom: 1px solid var(--border-color);
            transition: background-color 0.15s;
        }}

        .compact-row.no-comments {{
            grid-template-columns: {compact_thumb_size + 10}px {compact_column_ratio[1]}% 1fr;
        }}

        .compact-row:last-child {{
            border-bottom: none;
        }}

        .compact-row:hover {{
            background-color: var(--hover-color);
        }}

        .compact-row.hidden {{
            display: none;
        }}

        .compact-thumb {{
            width: {compact_thumb_size}px;
            height: {thumb_height}px;
            object-fit: cover;
            border-radius: 4px;
            border: 1px solid var(--border-color);
            display: block;
            cursor: pointer;
        }}

        .compact-comments {{
            font-size: __COMMENT_BODY_SIZE__;
            color: var(--text-color);
            line-height: 1.5;
            overflow: hidden;
            padding-left: 8px;
        }}

        .compact-comments-title {{
            font-weight: 600;
            margin-bottom: 3px;
            color: var(--accent-color);
            font-size: __COMMENT_TITLE_SIZE__;
        }}

        .compact-comments-body {{
            opacity: 0.8;
            white-space: pre-line;
            font-size: __COMMENT_BODY_SIZE__;
        }}

        .compact-icons {{
            display: flex;
            align-items: center;
            gap: 2px;
            flex-shrink: 0;
        }}

        .csv-preview {{
            display: none;
            position: absolute;
            bottom: 100%;
            left: 50%;
            transform: translateX(-50%);
            background-color: var(--sidebar-bg);
            border: 1px solid var(--border-color);
            border-radius: 6px;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.3);
            z-index: 1500;
            max-width: 600px;
            max-height: 400px;
            overflow: auto;
            margin-bottom: 8px;
        }}

        .csv-preview table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 11px;
            font-family: monospace;
        }}

        .csv-preview th,
        .csv-preview td {{
            border: 1px solid var(--border-color);
            padding: 6px 8px;
            text-align: left;
            max-width: 150px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            background-color: var(--bg-color);
            color: var(--text-color);
        }}

        .csv-preview th {{
            background-color: var(--sidebar-bg);
            font-weight: bold;
            position: sticky;
            top: 0;
            color: var(--accent-color);
        }}

        .file-link.csv:hover .csv-preview {{
            display: block;
        }}

        .modal {{
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background-color: var(--modal-bg);
            z-index: 2000;
            transition: background-color 0.3s;
        }}

        .modal-content {{
            position: relative;
            max-width: 90vw;
            max-height: 90vh;
            margin: 5% auto;
            padding: 20px;
            background-color: var(--sidebar-bg);
            border-radius: 8px;
            box-shadow: 0 0 20px rgba(0, 0, 0, 0.5);
        }}

        .modal img {{
            max-width: 100%;
            max-height: 75vh;
            display: block;
            margin: 0 auto;
            border-radius: 4px;
        }}

        .close-btn {{
            position: absolute;
            top: 15px;
            right: 20px;
            cursor: pointer;
            font-size: 32px;
            color: var(--text-color);
            transition: opacity 0.3s ease;
            z-index: 1;
        }}

        .close-btn:hover {{
            opacity: 0.7;
        }}

        h1, h2, h3, h4, h5, h6 {{
            color: var(--accent-color);
            margin: 40px auto 20px auto;
            max-width: 1400px;
            text-transform: capitalize;
            scroll-margin-top: 20px;
            padding-left: 10px;
        }}

        /* First heading on the page should sit flush with the top padding,
           not add an extra 40px gap above it */
        .main-content > h1:first-child,
        .main-content > h2:first-child,
        .main-content > h3:first-child,
        .main-content > h4:first-child,
        .main-content > h5:first-child,
        .main-content > h6:first-child,
        .page-header + h1,
        .page-header + h2,
        .page-header + h3,
        .page-header + h4,
        .page-header + h5,
        .page-header + h6 {{
            margin-top: 0;
        }}
        
        __HEADING_CSS_PLACEHOLDER__

        __UNDERLINE_CSS_PLACEHOLDER__

        /* link: hyperlinks in captions and compact titles */
        .comments-link {{
            color: var(--accent-color);
            text-decoration: none;
            transition: opacity 0.2s;
        }}
        .comments-link:hover {{
            opacity: 0.75;
            text-decoration: underline;
        }}
        .caption-link {{
            display: block;
        }}

        /* Page header / footer */
        .page-header {{
            max-width: 1400px;
            margin: 0 auto 30px auto;
            padding: 0 10px 18px 10px;
            border-bottom: 1px solid var(--border-color);
        }}
        .page-header-title {{
            font-size: 26px;
            font-weight: 700;
            color: var(--accent-color);
            margin-top: 0;
            margin-bottom: 4px;
        }}
        .page-header-subtitle {{
            font-size: 14px;
            color: var(--text-color);
            opacity: 0.75;
            margin-bottom: 8px;
        }}
        .page-header-description {{
            font-size: 13px;
            color: var(--text-color);
            opacity: 0.65;
            line-height: 1.6;
        }}
        .page-footer {{
            max-width: 1400px;
            margin: 40px auto 0 auto;
            padding: 16px 10px;
            border-top: 1px solid var(--border-color);
            font-size: 12px;
            color: var(--text-color);
            opacity: 0.6;
            text-align: center;
        }}

        .section-heading.hidden,
        .section-description.hidden,
        .gallery-container.hidden,
        .compact-list.hidden {{
            display: none;
        }}

        .section-description {{
            max-width: 1400px;
            margin: -10px auto 20px auto;
            padding: 10px 15px;
            background-color: var(--hover-color);
            border-left: 3px solid var(--accent-color);
            border-radius: 4px;
            font-size: 14px;
            color: var(--text-color);
            white-space: pre-line;
        }}

        .section-title-comment {{
            font-size: 14px;
            font-weight: normal;
            color: var(--text-color);
            margin-left: 10px;
            opacity: 0.7;
        }}

        @media (max-width: 768px) {{
            body {{
                padding-left: {'40px' if sidebar_collapsible else '0'};
                padding-top: {'0' if sidebar_collapsible else '60px'};
            }}

            .toc-sidebar:not(.collapsible) {{
                width: 100%;
                height: auto;
                position: relative;
            }}

            .main-content {{
                padding: 20px;
            }}

            .gallery-container {{
                grid-template-columns: 1fr;
            }}

            .csv-preview {{
                max-width: 300px;
                left: 0;
                transform: none;
            }}
        }}

    </style>
</head>
<body data-theme="{default_theme}">

    <script>
        const csvData = {json.dumps(csv_previews)};
    </script>

""" + (f"""    <div class="sidebar-backdrop" id="sidebar-backdrop" onclick="closeSidebar()"></div>
    <div class="sidebar-collapsed-strip" id="sidebar-strip">
        <span class="sidebar-contents-badge" onclick="toggleSidebar()">Index</span>
        <div class="sidebar-nav-dots" id="sidebar-nav-dots"></div>
    </div>
""" if sidebar_collapsible else "") + f"""    <div class="sidebar{' collapsible' if sidebar_collapsible else ''}">
        <div class="sidebar-header">
            <div class="sidebar-title">{navigation_title}</div>""" + (f"""
            <button class="sidebar-hamburger" onclick="closeSidebar()" title="Close navigation (S)" style="margin-left:auto;background:none;border:none;color:var(--text-color);cursor:pointer;padding:4px 6px;border-radius:4px;opacity:0.5;transition:opacity 0.15s;">
                <svg width="16" height="16" viewBox="0 0 24 24" fill="currentColor"><path d="M19 6.41L17.59 5 12 10.59 6.41 5 5 6.41 10.59 12 5 17.59 6.41 19 12 13.41 17.59 19 19 17.59 13.41 12z"/></svg>
            </button>""" if sidebar_collapsible else "") + """
        </div>
        <div class="search-box">
            <div class="search-wrap">
                <input type="text"
                       class="search-input"
                       id="gallery-search"
                       placeholder="Search… (F)"
                       title="Press F to focus · Esc to clear · ↑↓ navigate"
                       oninput="filterGallery()">
                <button class="search-clear" id="search-clear" onclick="clearSearch()" title="Clear search (Esc)">✕</button>
            </div>
            <div class="search-icon-group">
                <button class="search-icon-btn" title="Expand all" onclick="expandAll()">
                    <svg width="14" height="14" viewBox="0 0 24 24" fill="currentColor">
                        <path d="M19 13h-6v6h-2v-6H5v-2h6V5h2v6h6v2z"/>
                    </svg>
                </button>
                <button class="search-icon-btn" title="Collapse all" onclick="collapseAll()">
                    <svg width="14" height="14" viewBox="0 0 24 24" fill="currentColor">
                        <path d="M19 13H5v-2h14v2z"/>
                    </svg>
                </button>
                <button class="search-icon-btn" id="themeToggleBtn" title="Toggle light / dark theme (T)" onclick="toggleTheme()">
                    <svg id="theme-icon" width="14" height="14" viewBox="0 0 24 24" fill="currentColor">
                        <path d="M12 7a5 5 0 0 1 5 5 5 5 0 0 1-5 5 5 5 0 0 1-5-5 5 5 0 0 1 5-5m0-2a7 7 0 0 0-7 7 7 7 0 0 0 7 7 7 7 0 0 0 7-7 7 7 0 0 0-7-7M2 11h2v2H2v-2m18 0h2v2h-2v-2M11 2h2v2h-2V2m0 18h2v2h-2v-2M4.22 3.93l1.42 1.42-1.42 1.41-1.41-1.41 1.41-1.42m15.14 13.3 1.41 1.41-1.41 1.42-1.42-1.42 1.42-1.41M4.22 19.07l-1.41-1.42 1.41-1.41 1.42 1.41-1.42 1.42M19.36 5.36l-1.42-1.42 1.42-1.41 1.41 1.42-1.41 1.41z"/>
                    </svg>
                </button>
                <button class="search-icon-btn" id="helpBtn" title="Keyboard shortcuts" onclick="toggleHelp()">
                    <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                        <path d="M8.5 8.5A3.5 3.5 0 1 1 12 12v1"/>
                        <path d="M12 17h.01"/>
                    </svg>
                </button>
            </div>
            <div class="help-tooltip" id="helpTooltip">
                <div class="help-row"><kbd>F</kbd> Focus search box</div>
                <div class="help-row"><kbd>↑</kbd><kbd>↓</kbd> Navigate sidebar links</div>
                <div class="help-row"><kbd>T</kbd> Toggle light / dark</div>
                <div class="help-row"><kbd>Esc</kbd> Clear search / close</div>
                <div class="help-row"><kbd>N</kbd> Focus navigation panel</div>
                <div class="help-row"><kbd>M</kbd> Focus main content</div>
            </div>
            <span class="search-count" id="search-count"></span>
        </div>
        <div class="toc-links">
            {toc_html}
        </div>
    </div>

    <div class="main-content" id="main-content-area" tabindex="-1" style="outline:none;">
        {page_header_html}
        {main_content_html}
        {page_footer_html}
    </div>

    <div class="modal" id="imageModal" onclick="closeModal()">
        <div class="modal-content" onclick="event.stopPropagation()">
            <span class="close-btn" onclick="closeModal()">×</span>
            <img id="modal-image" src="" alt="Modal Image">
        </div>
    </div>

    <script>
        {collapse_js}

        function openSidebar() {{
            var sb = document.querySelector('.sidebar.collapsible');
            var bd = document.getElementById('sidebar-backdrop');
            if (sb) sb.classList.add('open');
            if (bd) bd.classList.add('active');
        }}

        function closeSidebar() {{
            var sb = document.querySelector('.sidebar.collapsible');
            var bd = document.getElementById('sidebar-backdrop');
            if (sb) sb.classList.remove('open');
            if (bd) bd.classList.remove('active');
        }}

        function toggleSidebar() {{
            var sb = document.querySelector('.sidebar.collapsible');
            if (!sb) return;
            if (sb.classList.contains('open')) closeSidebar(); else openSidebar();
        }}

        // Build nav dots from H2 sections
        (function() {{
            var dotsContainer = document.getElementById('sidebar-nav-dots');
            if (!dotsContainer) return;
            var sections = document.querySelectorAll('.section[data-section-id]');
            sections.forEach(function(sec) {{
                var id = sec.getAttribute('data-section-id');
                var title = sec.querySelector('.section-title');
                var label = title ? (title.querySelector('.section-title-text') || title).textContent.trim() : id;
                var dot = document.createElement('div');
                dot.className = 'sidebar-nav-dot';
                dot.setAttribute('data-dot-section', id);
                dot.title = label;
                dot.innerHTML = '<span class="sidebar-nav-dot-tooltip">' + label + '</span>';
                dot.addEventListener('click', function() {{
                    sec.scrollIntoView({{ behavior: 'smooth', block: 'start' }});
                }});
                dotsContainer.appendChild(dot);
            }});
        }})();        
        
        function toggleTheme() {{
            const body = document.body;
            const currentTheme = body.getAttribute('data-theme');
            const newTheme = currentTheme === 'light' ? 'dark' : 'light';
            body.setAttribute('data-theme', newTheme);
            localStorage.setItem('gallery-theme', newTheme);
            updateThemeIcon(newTheme);
        }}

        function updateThemeIcon(theme) {{
            const icon = document.getElementById('theme-icon');
            if (!icon) return;
            if (theme === 'dark') {{
                // Show moon icon in dark mode
                icon.innerHTML = '<path d="M12 3a9 9 0 1 0 9 9c0-.46-.04-.92-.1-1.36a5.389 5.389 0 0 1-4.4 2.26 5.403 5.403 0 0 1-3.14-9.8c-.44-.06-.9-.1-1.36-.1z"/>';
            }} else {{
                // Show sun icon in light mode
                icon.innerHTML = '<path d="M12 7a5 5 0 0 1 5 5 5 5 0 0 1-5 5 5 5 0 0 1-5-5 5 5 0 0 1 5-5m0-2a7 7 0 0 0-7 7 7 7 0 0 0 7 7 7 7 0 0 0 7-7 7 7 0 0 0-7-7M2 11h2v2H2v-2m18 0h2v2h-2v-2M11 2h2v2h-2V2m0 18h2v2h-2v-2M4.22 3.93l1.42 1.42-1.42 1.41-1.41-1.41 1.41-1.42m15.14 13.3 1.41 1.41-1.41 1.42-1.42-1.42 1.42-1.41M4.22 19.07l-1.41-1.42 1.41-1.41 1.42 1.41-1.42 1.42M19.36 5.36l-1.42-1.42 1.42-1.41 1.41 1.42-1.41 1.41z"/>';
            }}
        }}

        const savedTheme = localStorage.getItem('gallery-theme');
        if (savedTheme) {{
            document.body.setAttribute('data-theme', savedTheme);
            updateThemeIcon(savedTheme);
        }} else {{
            updateThemeIcon('{default_theme}');
        }}

        function openModal(imagePath, altText) {{
            const modal = document.getElementById('imageModal');
            const modalImage = document.getElementById('modal-image');
            modalImage.src = imagePath;
            modalImage.alt = altText;
            modal.style.display = 'block';
        }}

        function closeModal() {{
            document.getElementById('imageModal').style.display = 'none';
        }}

        document.addEventListener('keydown', function(event) {{
            const tag     = document.activeElement.tagName;
            const inInput = tag === 'INPUT' || tag === 'TEXTAREA';

            if (event.key === 'Escape') {{
                if (document.getElementById('imageModal').style.display === 'block') {{
                    closeModal();
                }} else if (inInput) {{
                    clearSearch();
                    document.activeElement.blur();
                    document.getElementById('main-content-area').focus();
                }}
                return;
            }}

            // F — focus search box only; Esc exits
            if (event.key === 'f' || event.key === 'F') {{
                if (!inInput) {{
                    event.preventDefault();
                    const inp = document.getElementById('gallery-search');
                    inp.focus(); inp.select();
                }}
                // When already in the search box, let F type normally
                return;
            }}

            // T — toggle theme (only when not typing)
            if ((event.key === 't' || event.key === 'T') && !inInput) {{
                event.preventDefault();
                toggleTheme();
                return;
            }}

            // S — toggle collapsible sidebar
            if ((event.key === 's' || event.key === 'S') && !inInput) {{
                event.preventDefault();
                toggleSidebar();
                return;
            }}
            
            // N — focus navigation sidebar
            if ((event.key === 'n' || event.key === 'N') && !inInput) {{
                event.preventDefault();
                const links = Array.from(
                    document.querySelectorAll('.toc-item:not(.nav-hidden) .toc-link')
                );
                if (links.length) links[0].focus();
                return;
            }}

            // M — focus main content (for keyboard scrolling)
            if ((event.key === 'm' || event.key === 'M') && !inInput) {{
                event.preventDefault();
                document.getElementById('main-content-area').focus();
                return;
            }}

            // Arrow keys — navigate visible nav items ONLY when a toc-link is already focused
            const focusedIsTocLink = document.activeElement.classList.contains('toc-link');
            if (focusedIsTocLink && (event.key === 'ArrowDown' || event.key === 'ArrowUp')) {{
                event.preventDefault();
                const links = Array.from(
                    document.querySelectorAll('.toc-item:not(.nav-hidden) .toc-link')
                );
                if (!links.length) return;
                const idx = links.indexOf(document.activeElement);
                if (event.key === 'ArrowDown') {{
                    links[idx + 1 < links.length ? idx + 1 : 0].focus();
                }} else {{
                    links[idx - 1 >= 0 ? idx - 1 : links.length - 1].focus();
                }}
            }}
        }});

        function clearSearch() {{
            const input = document.getElementById('gallery-search');
            input.value = '';
            // Remove any stale active highlights
            document.querySelectorAll('.toc-link.active').forEach(link => {{
                link.classList.remove('active');
            }});
            filterGallery();
        }}

        function toggleHelp() {{
            document.getElementById('helpTooltip').classList.toggle('visible');
        }}

        document.addEventListener('click', function(e) {{
            const btn = document.getElementById('helpBtn');
            const tt  = document.getElementById('helpTooltip');
            if (btn && tt && !btn.contains(e.target) && !tt.contains(e.target)) {{
                tt.classList.remove('visible');
            }}
        }});

        function filterGallery() {{
            const searchInput = document.getElementById('gallery-search');
            const searchTerm  = searchInput.value.toLowerCase().trim();
            const clearBtn    = document.getElementById('search-clear');
            const allItems    = document.querySelectorAll('.gallery-item, .compact-row');
            const allSections = document.querySelectorAll(
                '.section-heading, .section-description, .gallery-container, .compact-list'
            );
            const searchCount = document.getElementById('search-count');

            // Show / hide the ✕ clear button
            clearBtn.classList.toggle('visible', !!searchTerm);

            let visibleCount = 0;
            const totalCount = allItems.length;
            const visibleSections = new Set();

            // ── Filter page items ──────────────────────────────────────────────
            allItems.forEach(item => {{
                const searchData = (item.getAttribute('data-search') || '');
                const sectionId  = item.getAttribute('data-section');
                const matches    = !searchTerm || searchData.includes(searchTerm);
                item.classList.toggle('hidden', !matches);
                if (matches) {{
                    if (sectionId) visibleSections.add(sectionId);
                    visibleCount++;
                }}
            }});

            // Show/hide section wrappers
            allSections.forEach(section => {{
                const sectionId = section.getAttribute('data-section');
                const show = !searchTerm || (sectionId && visibleSections.has(sectionId));
                section.classList.toggle('hidden', !show);
            }});

            // ── Filter nav links to only show sections that have results ───────
            const allNavItems = document.querySelectorAll('.toc-item[data-nav-id]');
            allNavItems.forEach(navItem => {{
                const navId = navItem.getAttribute('data-nav-id');
                const selfVisible = !searchTerm || visibleSections.has(navId);
                const childNavIds = Array.from(
                    navItem.querySelectorAll('[data-nav-id]')
                ).map(el => el.getAttribute('data-nav-id'));
                const childVisible = childNavIds.some(id => visibleSections.has(id));
                navItem.classList.toggle('nav-hidden', !(selfVisible || childVisible));
            }});

            searchCount.textContent = searchTerm
                ? `${{visibleCount}} of ${{totalCount}} shown`
                : '';
        }}

        document.addEventListener('DOMContentLoaded', function() {{
            initCollapsible();
            restoreCollapsedStates();
            initializeCsvPreviews();
            initScrollSpy();

            // Default keyboard scroll focus on main content area
            const mainArea = document.getElementById('main-content-area');
            if (mainArea) mainArea.focus();

            document.querySelectorAll('.toc-link').forEach(link => {{
                link.addEventListener('click', function(e) {{
                    e.preventDefault();
                    const targetId = this.getAttribute('href');
                    const targetElement = document.querySelector(targetId);
                    if (targetElement) {{
                        targetElement.scrollIntoView({{
                            behavior: 'smooth',
                            block: 'start'
                        }});
                        closeSidebar();
                    }}
                }});
            }});
        }});

        function initializeCsvPreviews() {{
            const csvLinks = document.querySelectorAll('.file-link.csv[data-file-path]');

            csvLinks.forEach(link => {{
                const filePath = link.getAttribute('data-file-path');
                const previewDiv = link.querySelector('.csv-preview .csv-table-container');

                if (previewDiv && csvData[filePath]) {{
                    const csvRows = csvData[filePath];
                    const table = document.createElement('table');

                    csvRows.forEach((row, index) => {{
                        const tr = document.createElement('tr');
                        row.forEach(cell => {{
                            const td = document.createElement(index === 0 ? 'th' : 'td');
                            td.textContent = cell;
                            td.title = cell;
                            tr.appendChild(td);
                        }});
                        table.appendChild(tr);
                    }});

                    previewDiv.appendChild(table);
                }}
            }});
        }}
    </script>
</body>
</html>"""

        # ── Post-process: substitute placeholders that couldn't be inside the f-string
        # because they contain literal { } characters (CSS rules, var() references, etc.)
        html_template = (
            html_template
            .replace("__HEADING_CSS_PLACEHOLDER__",   heading_css)
            .replace("__UNDERLINE_CSS_PLACEHOLDER__", _underline_css)
            .replace("__NAV_FONT_CSS_PLACEHOLDER__",  _nav_font_css)
            .replace("__COMMENT_TITLE_SIZE__",         comment_title_font_size)
            .replace("__COMMENT_BODY_SIZE__",          comment_body_font_size)
        )

        return html_template

    try:
        root_path = Path(folder_to_scan).resolve()
        if not root_path.exists():
            raise FileNotFoundError(f"Folder '{folder_to_scan}' does not exist")

        print(f"🔍 Scanning folder recursively: {folder_to_scan}")

        gallery_data = scan_folder_recursive(root_path, root_path)

        if not gallery_data or (
            not gallery_data.get("_items")
            and not any(k not in ["_items", "_comments"] for k in gallery_data.keys())
        ):
            print("No image files found!")
            return None

        csv_previews = {}

        def collect_csv_previews(node: Dict[str, Any], root_path: Path):
            for item in node.get("_items", []):
                for file_info in item["other_files"]:
                    if file_info["extension"] == ".csv":
                        file_path = root_path / file_info["path"]
                        if file_path.exists():
                            csv_previews[file_info["path"]] = read_csv_preview(
                                file_path
                            )

            for key, value in node.items():
                if key not in ["_items", "_comments"]:
                    collect_csv_previews(value, root_path)

        collect_csv_previews(gallery_data, root_path)

        print("🌿 Generating HTML...")
        html_content = generate_html(
            gallery_data, csv_previews, root_path, compact_patterns, max_view_patterns, hide_icons_patterns
        )

        if output_file:
            output_path = Path(output_file).resolve()
            output_path.parent.mkdir(parents=True, exist_ok=True)
            if not store_content and output_path.parent != root_path:
                print(
                    "⚠️  output_file is outside folder_to_scan — image paths will be "
                    "broken unless you use store_content=True to embed images as base64."
                )
        else:
            output_path = root_path / "visual_library.html"

        # ── store_content: embed all images as base64 data URIs ──────────────
        if store_content:
            print("📦 Embedding images as base64 (store_content=True)...")
            import base64
            import mimetypes

            _embed_count = 0
            _skip_count  = 0

            def _embed_src(match):
                nonlocal _embed_count, _skip_count
                attr  = match.group(1)   # 'src' or 'href'
                quote = match.group(2)   # ' or "
                path  = match.group(3)
                # Skip already-embedded, external URLs, and anchors
                if path.startswith("data:") or path.startswith("http") or path.startswith("#"):
                    return match.group(0)
                # Resolve relative to root_path
                abs_path = (root_path / path).resolve()
                if not abs_path.exists():
                    _skip_count += 1
                    return match.group(0)
                mime, _ = mimetypes.guess_type(str(abs_path))
                if not mime:
                    _skip_count += 1
                    return match.group(0)
                with open(abs_path, "rb") as _fh:
                    b64 = base64.b64encode(_fh.read()).decode("ascii")
                _embed_count += 1
                return f'{attr}={quote}data:{mime};base64,{b64}{quote}'

            # Embed src="..." (images) and href="..." (linked files)
            html_content = re.sub(
                r'(src|href)=(["\'])([^"\'#][^"\']*)\2',
                _embed_src,
                html_content,
            )
            print(f"   ✅ Embedded {_embed_count} assets, skipped {_skip_count}")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_content)

        def count_items(node):
            count = len(node.get("_items", []))
            for key, value in node.items():
                if key not in ["_items", "_comments"]:
                    count += count_items(value)
            return count

        def count_folders(node, depth=0):
            folders = 0
            for key, value in node.items():
                if key not in ["_items", "_comments"]:
                    folders += 1
                    folders += count_folders(value, depth + 1)
            return folders

        total_items = count_items(gallery_data)
        total_folders = count_folders(gallery_data)

        print("✅ Gallery created successfully!")
        print(f"- {total_folders} folders scanned (up to {max_depth} levels deep)")
        print(f"- {total_items} image sets")
        print(f"- {len(csv_previews)} CSV files with preview")
        print(f"- 🔍 Search enabled for all items (includes folder names)")
        print(f"- 📁 Folder links, comments (title/comments/link) supported")
        print(f"- 🗂️ Collapsible sidebar TOC with hierarchical font sizes")
        print(f"- 🎯 Active navigation highlighting via IntersectionObserver")
        print(f"- 📐 Custom heading font sizes: {heading_font_sizes}")
        if max_view:
            print(f"- 🔭 max_view patterns: {max_view} (thumb minmax: {max_view_thumb_size}px)")
        if hide_icons:
            print(f"- 🚫 hide_icons patterns: {hide_icons}")
        if underline_headings:
            print(f"- 〰️  Underline headings: {list(underline_headings.keys())}")
        if page_font or page_font_size:
            print(f"- 🔤 Page font: {_body_font_family} / {_body_font_size}")
        if theme and theme_config_file:
            print(f"- 🎨 Theme: '{theme}' from sheet '{theme_sheet}'")
        if store_content:
            print(f"- 📦 store_content=True — all assets embedded as base64")
        print(f"- Output: {output_path}")

        return str(output_path)

    except Exception as e:
        print(f"Error creating gallery: {e}")
        import traceback

        traceback.print_exc()
        return None