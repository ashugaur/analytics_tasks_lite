# %% convert_markdown_to_html

## Dependencies
from pathlib import Path
import re
import base64
from datetime import datetime
from bs4 import BeautifulSoup


def load_theme_config(theme_config_file, theme_id):
    """
    Load light + dark color dicts from an Excel (.xlsm/.xlsx) Theme_Config sheet.

    Expected sheet name  : 'Theme_Config'
    Expected columns     : Test_ID, Theme_Name, primary, text, muted, light,
                           content_bg, slide_bg, header_border, bg_dark, bg,
                           bg_light, text_muted, highlight, border, border_muted,
                           secondary, danger, warning, success, info, Notes

    Column → CSS-variable mapping
    ─────────────────────────────
    primary        → accent          (buttons, links, active nav, reference border)
    text           → text
    bg             → bg
    content_bg     → code_bg         (code block background)
    slide_bg       → sidebar_bg      (sidebar / panel background)
    bg_light       → hover           (hover highlight), reference_bg
    border         → border
    border_muted   → (informational only — not currently mapped)
    muted          → reference_border (subtle accent line on reference blocks)

    Args:
        theme_config_file (str | Path): UNC or local path to the .xlsm / .xlsx file.
        theme_id          (str)        : Value to match in the Test_ID column (e.g. 'test_1').

    Returns:
        tuple[dict, dict]: (light_mode_colors, dark_mode_colors)
                           Each dict has keys: bg, text, sidebar_bg, accent,
                           border, hover, code_bg, reference_bg, reference_border.
        Returns (None, None) if the theme_id is not found or the file cannot be read.
    """
    try:
        import openpyxl
    except ImportError:
        print("❌ load_theme_config: openpyxl is required — pip install openpyxl")
        return None, None

    try:
        import io

        path = Path(theme_config_file)
        if not path.exists():
            print(f"❌ load_theme_config: file not found — {path}")
            return None, None

        # Read entire file into memory so openpyxl never holds an OS-level
        # lock on the .xlsm — the file stays editable while open in Excel.
        with open(path, "rb") as fh:
            file_bytes = io.BytesIO(fh.read())

        wb = openpyxl.load_workbook(
            file_bytes, read_only=True, data_only=True, keep_vba=False
        )

        sheet_name = "Theme_Config"
        if sheet_name not in wb.sheetnames:
            print(
                f"❌ load_theme_config: sheet '{sheet_name}' not found. "
                f"Available: {wb.sheetnames}"
            )
            return None, None

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("❌ load_theme_config: sheet is empty")
            return None, None

        # Build header index (case-insensitive, strip whitespace)
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        def col(name):
            key = name.strip().lower()
            if key not in header:
                raise KeyError(f"Column '{name}' not found in Theme_Config header")
            return header.index(key)

        # Required column indices
        idx = {
            "test_id": col("test_id"),
            "theme_name": col("theme_name"),
            "primary": col("primary"),
            "text": col("text"),
            "muted": col("muted"),
            "content_bg": col("content_bg"),
            "slide_bg": col("slide_bg"),
            "bg": col("bg"),
            "bg_light": col("bg_light"),
            "border": col("border"),
            "border_muted": col("border_muted"),
        }

        light_row = dark_row = None
        for row in rows[1:]:
            tid = (
                str(row[idx["test_id"]]).strip()
                if row[idx["test_id"]] is not None
                else ""
            )
            tname = (
                str(row[idx["theme_name"]]).strip().lower()
                if row[idx["theme_name"]] is not None
                else ""
            )
            if tid == str(theme_id).strip():
                if tname == "light":
                    light_row = row
                elif tname == "dark":
                    dark_row = row

        if light_row is None and dark_row is None:
            print(
                f"❌ load_theme_config: theme_id '{theme_id}' not found in Theme_Config"
            )
            return None, None

        def v(row, key):
            """Safe cell value as string, empty string if missing."""
            val = row[idx[key]]
            return str(val).strip() if val is not None else ""

        def build_colors(row):
            primary = v(row, "primary")
            text = v(row, "text")
            muted = v(row, "muted")
            content_bg = v(row, "content_bg")
            slide_bg = v(row, "slide_bg")
            bg = v(row, "bg")
            bg_light = v(row, "bg_light")
            border = v(row, "border")
            border_muted = v(row, "border_muted")
            return {
                "bg": bg,
                "text": text,
                "sidebar_bg": slide_bg,
                "accent": primary,
                "border": border,
                "hover": bg_light,
                "code_bg": content_bg,
                "reference_bg": bg_light,
                "reference_border": muted,
            }

        light_colors = build_colors(light_row) if light_row is not None else None
        dark_colors = build_colors(dark_row) if dark_row is not None else None

        mode_label = []
        if light_colors:
            mode_label.append("light")
        if dark_colors:
            mode_label.append("dark")
        print(
            f"✅ load_theme_config: loaded theme '{theme_id}' "
            f"({' + '.join(mode_label)} modes)"
        )
        return light_colors, dark_colors

    except Exception as e:
        import traceback

        print(f"❌ load_theme_config error: {e}")
        traceback.print_exc()
        return None, None


def convert_markdown_to_html(
    markdown_file,
    output_file="convert_markdown_to_html.html",
    page_title="Documentation Report",
    page_subtitle=None,
    page_description=None,
    footer_content=None,
    sidebar_width="280px",
    default_theme="light",
    light_mode_colors=None,
    dark_mode_colors=None,
    navigation_title="📚 Navigation",
    include_code_blocks=True,
    code_blocks_collapsed=False,
    style_output_blocks=True,
    default_embed_height=500,
    theme_config_file=None,
    theme=None,
    report_header=None,
    report_footer=None,
    default_code_height="200px",
    monospace_output_heuristic=True,
    underline_headings=None,
    nav_font_sizes=None,
    heading_font_sizes=None,
    page_font=None,
    page_font_size=None,
    sidebar_collapsible=True,
    format_file=None,
    format_sheet_name="Formats",
):
    """
    Convert a Python-generated markdown file to a standalone HTML report.
    Handles specific patterns: ## headers, ### subheaders, code blocks,
    tab-indented reference text, and images.

    Args:
        markdown_file (str): Path to the markdown file
        output_file (str): Name of the output HTML file
        page_title (str): Main title for the report
        page_subtitle (str): Optional subtitle
        page_description (str): Optional description text for header
        footer_content (str): Optional footer content (HTML allowed)
        sidebar_width (str): Width of the sidebar (default: "280px")
        default_theme (str): Default theme - "light" or "dark"
        light_mode_colors (dict): Custom light mode colors (overridden by theme= if provided)
        dark_mode_colors (dict): Custom dark mode colors (overridden by theme= if provided)
        navigation_title (str): Title for the navigation sidebar
        include_code_blocks (bool): Whether to include code blocks in output
        code_blocks_collapsed (bool): Whether code blocks start collapsed
        style_output_blocks (bool): Whether to style output blocks with border/background (default: True)
        default_embed_height (int): Default height in px for ##-add-html embedded charts (default: 500)
        theme_config_file (str | Path): UNC/local path to the .xlsm/.xlsx Theme_Config workbook.
                           When provided together with theme=, colors are loaded from the
                           'Theme_Config' sheet and override light_mode_colors / dark_mode_colors.
        theme (str): Test_ID value from the Theme_Config sheet to apply (e.g. 'test_1').
                     Requires theme_config_file to also be set.
        report_header (dict): Optional structured header block rendered above the content area.
                     Supported keys:
                       "title"       – large H1 heading
                       "subtitle"    – smaller H2 below the title
                       "metadata"    – dict of key/value pairs rendered as label: value rows
                                       e.g. {"Author": "Alice", "Version": "1.0"}
                       "description" – one or more paragraphs (separate with \\n\\n)
                     When provided this replaces the default page_title/page_subtitle/
                     page_description header block entirely.
        report_footer (dict): Optional structured footer block rendered below the content area.
                     Supported keys:
                       "title"   – small heading inside the footer
                       "content" – body text (separate paragraphs with \\n\\n)
                     When provided this replaces the default footer_content/timestamp block.
        default_code_height (str): CSS max-height for code block pre elements (default: '200px').
                     Set to 'none' to disable height capping.
        underline_headings (dict): Optional per-level separator line drawn ABOVE headings.
                     Keys are heading levels as strings: "h2", "h3", "h4", "h5", "h6".
                     Each value is a dict with optional keys:
                       "line_width"  – CSS border thickness, e.g. "1px" (default "1px")
                       "line_color"  – CSS color value, e.g. "#ffffff" or "var(--border-color)"
                                       (default "var(--border-color)")
                       "line_style"  – CSS border-style, e.g. "solid", "dashed", "dotted"
                                       (default "solid")
                     Example — solid line above H2, dashed above H3:
                       underline_headings={
                           "h2": {"line_width": "1px", "line_color": "#ffffff", "line_style": "solid"},
                           "h3": {"line_width": "1px", "line_color": "var(--border-color)", "line_style": "dashed"},
                       }
                     Levels not listed get no separator line.
        nav_font_sizes (dict): Optional per-depth font sizes for sidebar navigation links,
                     giving visual hierarchy to headings vs. subheadings.
                     Keys are heading depth integers (2 = H2 section, 3 = H3, …, 6 = H6).
                     Values are CSS font-size strings.
                     Example:
                       nav_font_sizes={2: "13px", 3: "12px", 4: "11px", 5: "11px", 6: "10px"}
                     Default when not set: H2 = 13px bold, H3+ = 12px normal (existing behaviour).
        heading_font_sizes (dict): Optional font sizes for H2–H6 section headings in the main
                     content area. Keys are integers 2–6; values are CSS font-size strings.
                     Example: heading_font_sizes={2: "28px", 3: "22px", 4: "18px"}
                     If not provided, the stylesheet defaults are used.
        page_font (str): CSS font-family applied to the entire page body.
                     Example: "Arial" or "'Inter', sans-serif". Default: system-ui stack.
        page_font_size (str): CSS font-size for the body. Example: "15px". Default: "14px".
        monospace_output_heuristic (bool): When True (default), any indented reference block
                     that contains a line of 20+ repeated '=' or '-' characters is automatically
                     rendered as a <pre class="model-output"> block so that fixed-width model
                     summaries (statsmodels, sklearn, scipy etc.) preserve character alignment.
                     Set to False to disable auto-detection and rely solely on explicit
                     ##-pre-start / ##-pre-end markers in the source .py file.

    Keyboard shortcuts (built into every generated HTML page):
        F        Focus the search box.
        T        Toggle light / dark theme from anywhere on the page.
        N        Focus the navigation sidebar (then ↑/↓ to move between links).
        M        Focus the main content area (default on load; use for keyboard scrolling).
        ↑ / ↓   Navigate visible sidebar links (only when a sidebar link is focused).
        Esc      Clear search / close any open modal (returns focus to main content).
        ?        Click the ? button in the sidebar header to show this cheatsheet.

    Returns:
        str: Path to the generated HTML file
    """

    # ── Load column formats from Excel sheet ────────────────────────────
    _global_column_formats = {}
    _override_column_formats = {}
    if format_file is not None:
        try:
            import openpyxl as _xl_fmt
            import io as _io_fmt

            _fmt_path = Path(format_file)
            if _fmt_path.exists():
                with open(_fmt_path, "rb") as _fh:
                    _buf = _io_fmt.BytesIO(_fh.read())
                _wb = _xl_fmt.load_workbook(_buf, read_only=True, data_only=True, keep_vba=False)
                if format_sheet_name in _wb.sheetnames:
                    _ws = _wb[format_sheet_name]
                    _rows = list(_ws.iter_rows(values_only=True))
                    if _rows:
                        _hdr = [str(c).strip().lower() if c else "" for c in _rows[0]]
                        _ci = {name: _hdr.index(name) for name in _hdr if name}
                        _loaded_g = 0
                        _loaded_o = 0
                        for _r in _rows[1:]:
                            if not _r or not _r[_ci.get("column_pattern", 0)]:
                                continue
                            
                            _col_pat = str(_r[_ci["column_pattern"]]).strip().lower()
                            _fmt_str = str(_r[_ci["format_string"]]).strip() if _r[_ci.get("format_string", 1)] else ""
                            _scope = str(_r[_ci.get("scope", 2)]).strip().lower() if _r[_ci.get("scope", 2)] else "global"
                            _tbl_pat = str(_r[_ci.get("table_pattern", 3)]).strip().lower() if _ci.get("table_pattern") is not None and _r[_ci["table_pattern"]] else ""
                            
                            if not _fmt_str:
                                continue
                            
                            if _scope == "override" and _tbl_pat:
                                _override_column_formats[(_tbl_pat, _col_pat)] = _fmt_str
                                _loaded_o += 1
                            else:
                                _global_column_formats[_col_pat] = _fmt_str
                                _loaded_g += 1
                        
                        print(f"✅ Loaded formats from '{format_sheet_name}': {_loaded_g} global, {_loaded_o} override")
                    _wb.close()
                else:
                    print(f"⚠️  format_sheet_name '{format_sheet_name}' not found in {_fmt_path.name}")
            else:
                print(f"⚠️  format_file not found: {_fmt_path}")
        except Exception as _e:
            print(f"⚠️  Failed to load format config: {_e}")

    # ── Theme config loading (theme= + theme_config_file= take priority) ──────
    if theme is not None and theme_config_file is not None:
        _light, _dark = load_theme_config(theme_config_file, theme)
        if _light is not None:
            light_mode_colors = _light
        if _dark is not None:
            dark_mode_colors = _dark
    elif theme is not None and theme_config_file is None:
        print(
            "⚠️  theme= provided but theme_config_file= is missing — using default colors"
        )

    # Set default colors
    if light_mode_colors is None:
        light_mode_colors = {
            "bg": "#ffffff",
            "text": "#333333",
            "sidebar_bg": "#f8f9fa",
            "accent": "#007bff",
            "border": "#dee2e6",
            "hover": "#f1f3f5",
            "code_bg": "#f5f5f5",
            "reference_bg": "#fffbeb",
            "reference_border": "#fbbf24",
        }

    if dark_mode_colors is None:
        dark_mode_colors = {
            "bg": "#1e1e1e",
            "text": "#e0e0e0",
            "sidebar_bg": "#2d2d2d",
            "accent": "#4a9eff",
            "border": "#444444",
            "hover": "#383838",
            "code_bg": "#2a2a2a",
            "reference_bg": "#2d2516",
            "reference_border": "#d97706",
        }

    def image_to_base64(image_path):
        """Convert image to base64 data URI"""
        try:
            with open(image_path, "rb") as img_file:
                encoded = base64.b64encode(img_file.read()).decode("utf-8")
                ext = image_path.suffix.lower()
                mime_type = {
                    ".png": "image/png",
                    ".jpg": "image/jpeg",
                    ".jpeg": "image/jpeg",
                    ".gif": "image/gif",
                    ".svg": "image/svg+xml",
                }.get(ext, "image/png")
                return f"data:{mime_type};base64,{encoded}"
        except Exception as e:
            print(f"Warning: Could not encode image {image_path}: {e}")
            return ""

    def embed_html_chart(html_path, height, chart_index):
        """
        Embed an external HTML chart file inline in the report.

        Follows the same contract as slidejs.extract_chart_components:
        - Chart HTML uses #container as the root div (from slidejs_chart_template.html)
        - Inline <script> is extracted and wrapped in an IIFE with try/catch
        - Container div ID is namespaced to chart_N_container to prevent clashes
        - CDN <script src="..."> tags are re-emitted before the IIFE so libraries load first
        - CSS is scoped simply: styles are wrapped with #chart_N_container { ... }
            (no brace-walking parser needed — chart styles are minimal by design)
        - window.addEventListener('resize', ...) is removed; the IIFE handles resize
            via myChart.resize() in a setTimeout after init
        - window.setChartTheme is detected and registered so the report's theme
            toggle can call it when switching light/dark mode

        Args:
            html_path  : absolute or relative path to the chart .html file
            height     : embed height in pixels (int)
            chart_index: zero-based counter used for namespacing

        Returns:
            str: self-contained HTML snippet ready to drop into the report body
        """
        try:
            html_path = Path(html_path)
            if not html_path.exists():
                msg = f"⚠ Chart not found: {html_path}"
                print(f"❌ embed_html_chart: {msg}")
                return f'<div style="color:red;padding:1rem;border:1px solid red;">{msg}</div>'

            print(f"\n  📊 embed_html_chart [{chart_index}]: {html_path.name}")

            raw = html_path.read_text(encoding="utf-8")
            soup = BeautifulSoup(raw, "html.parser")

            container_id = f"chart_{chart_index}_container"

            # ── 1. CDN script tags ─────────────────────────────────────────────
            cdns = []
            for tag in soup.find_all("script", src=True):
                cdns.append(tag["src"])
                print(f"    📦 CDN: {tag['src']}")

            is_d3 = any("d3" in c.lower() for c in cdns)
            is_echarts = any("echarts" in c.lower() for c in cdns)

            if is_d3:
                print("    🔵 D3 chart detected")
            elif is_echarts:
                print("    🟢 ECharts chart detected")

            # ── 2. Inline scripts ──────────────────────────────────────────────
            inline_scripts = []
            for tag in soup.find_all("script", src=False):
                if tag.string and tag.string.strip():
                    inline_scripts.append(tag.string)

            if not inline_scripts:
                print("    ⚠️  No inline scripts found — embedding as static HTML")
                # Fall back: just return the body content as-is inside a wrapper
                body = soup.find("body")
                inner = str(body) if body else raw
                return (
                    f'<div id="{container_id}" '
                    f'style="width:100%;height:{height}px;overflow:hidden;margin:1.5rem 0;">'
                    f"\n{inner}\n</div>"
                )

            combined = "\n\n// --- next block ---\n\n".join(inline_scripts)

            # ── 3. Container ID substitution (same 6 patterns as slidejs) ─────
            # All chart templates use #container / getElementById('container') etc.
            id_patterns = [
                (
                    r"getElementById\s*\(\s*['\"]container['\"]\s*\)",
                    f"getElementById('{container_id}')",
                ),
                (
                    r"getElementById\s*\(\s*['\"]chart['\"]\s*\)",
                    f"getElementById('{container_id}')",
                ),
                (
                    r"getElementById\s*\(\s*['\"]main['\"]\s*\)",
                    f"getElementById('{container_id}')",
                ),
                (
                    r"querySelector\s*\(\s*['\"]#container['\"]\s*\)",
                    f"querySelector('#{container_id}')",
                ),
                (
                    r"querySelector\s*\(\s*['\"]#chart['\"]\s*\)",
                    f"querySelector('#{container_id}')",
                ),
                (
                    r"querySelector\s*\(\s*['\"]#main['\"]\s*\)",
                    f"querySelector('#{container_id}')",
                ),
                (
                    r'd3\.select\s*\(\s*["\']#container["\']\s*\)',
                    f'd3.select("#{container_id}")',
                ),
                (
                    r'd3\.select\s*\(\s*["\']#chart["\']\s*\)',
                    f'd3.select("#{container_id}")',
                ),
                (
                    r'd3\.select\s*\(\s*["\']#main["\']\s*\)',
                    f'd3.select("#{container_id}")',
                ),
            ]

            script = combined
            for pattern, replacement in id_patterns:
                script = re.sub(pattern, replacement, script)

            # ── 4. Custom tooltip scoping (mirrors slidejs) ────────────────────
            has_custom_tooltip = (
                "customTooltip" in script
                or "custom-tooltip" in script
                or "custom-tooltip-target" in script
            )
            tooltip_id = None
            if has_custom_tooltip:
                tooltip_id = f"customTooltip_{chart_index}"
                script = re.sub(
                    r"getElementById\s*\(\s*['\"]customTooltip['\"]\s*\)",
                    f"getElementById('{tooltip_id}')",
                    script,
                )
                script = re.sub(
                    r"querySelector\s*\(\s*['\"]#customTooltip['\"]\s*\)",
                    f"querySelector('#{tooltip_id}')",
                    script,
                )
                print(f"    💬 Custom tooltip → {tooltip_id}")

            # ── 5. Remove window resize listeners (IIFE handles resize itself) ─
            lines_buf = script.split("\n")
            filtered = []
            skip = False
            depth = 0

            for line in lines_buf:
                if (
                    "window.addEventListener('resize'" in line
                    or 'window.addEventListener("resize"' in line
                ):
                    # Single-line handler?
                    if line.count("(") == line.count(")") and ");" in line:
                        print("    🧹 Removed single-line resize handler")
                        continue
                    else:
                        skip = True
                        depth = line.count("{") - line.count("}")
                        print("    🧹 Removed multi-line resize handler")
                        continue

                if skip:
                    depth += line.count("{") - line.count("}")
                    if depth <= 0 and line.strip() in ["});", "};"]:
                        skip = False
                        depth = 0
                    continue

                filtered.append(line)

            script = "\n".join(filtered).strip()

            # ── 6. Detect theme support ────────────────────────────────────────
            has_theme = "window.setChartTheme" in script
            if has_theme:
                print(
                    f"    🎨 Theme support detected → will register as setChartTheme_{chart_index}"
                )
                # Rename to a namespaced function so multiple charts don't clash
                script = script.replace(
                    "window.setChartTheme",
                    f"window.setChartTheme_{chart_index}",
                )

            # ── 7. CSS — simple container-scoped styles ────────────────────────
            style_blocks = []
            for style_tag in soup.find_all("style"):
                if style_tag.string and style_tag.string.strip():
                    # Scope body/html rules to the container; leave @-rules alone
                    css = style_tag.string
                    css = re.sub(r"\bbody\b", f"#{container_id}", css)
                    # Re-scope [data-theme] rules back to body — theme attribute lives on <body>,
                    # not the chart container, so the scoped selector would never match.
                    css = re.sub(
                        rf"#{re.escape(container_id)}\[data-theme", "[data-theme", css
                    )
                    css = re.sub(r"\bhtml\b", f"#{container_id}", css)
                    # Remove overflow:hidden from #container so labels aren't clipped
                    css = re.sub(
                        r"(#container\s*\{[^}]*?)overflow\s*:\s*hidden\s*;",
                        r"\1overflow: visible;",
                        css,
                        flags=re.DOTALL,
                    )
                    style_blocks.append(css)

            scoped_css = ""
            if style_blocks:
                scoped_css = (
                    f"<style>\n"
                    f"#{container_id} {{ width:100%; height:{height}px; overflow:visible; }}\n"
                    + "\n".join(style_blocks)
                    + f"\n</style>\n"
                )
            else:
                scoped_css = (
                    f"<style>\n"
                    f"#{container_id} {{ width:100%; height:{height}px; overflow:visible; }}\n"
                    f"</style>\n"
                )

            # ── 8. Custom tooltip div (if needed) ─────────────────────────────
            tooltip_div = ""
            if tooltip_id:
                tooltip_div = f'<div id="{tooltip_id}" style="display:none;position:absolute;"></div>\n'

            # ── 9. Wrap script in IIFE with try/catch + post-init resize ───────
            indented = "\n".join("        " + l for l in script.split("\n"))

            # Register theme function on the report's global registry so the
            # report's light/dark toggle can call all chart theme functions at once.
            theme_registration = ""
            if has_theme:
                theme_registration = (
                    f"\n        // Register theme function with report\n"
                    f"        if (!window._chartThemeFns) window._chartThemeFns = [];\n"
                    f"        window._chartThemeFns.push(window.setChartTheme_{chart_index});\n"
                )

            resize_observer = (
                f"\n        // ResizeObserver: re-render when container width changes\n"
                f"        // Handles D3 charts (re-calls render fn) and ECharts (.resize())\n"
                f"        (function() {{\n"
                f"            if (typeof ResizeObserver === 'undefined') return;\n"
                f"            var _ro_el = document.getElementById('{container_id}');\n"
                f"            if (!_ro_el) return;\n"
                f"            var _ro_lastW = _ro_el.clientWidth;\n"
                f"            var _ro_timer = null;\n"
                f"            var _ro = new ResizeObserver(function(entries) {{\n"
                f"                var newW = entries[0].contentRect.width;\n"
                f"                if (Math.abs(newW - _ro_lastW) < 2) return;\n"
                f"                _ro_lastW = newW;\n"
                f"                clearTimeout(_ro_timer);\n"
                f"                _ro_timer = setTimeout(function() {{\n"
                f"                    try {{\n"
                f"                        if (typeof myChart !== 'undefined' && myChart && myChart.resize) {{\n"
                f"                            myChart.resize(); return;\n"
                f"                        }}\n"
                f"                        var _fns = ['createSankeyChart','createChart','drawChart',\n"
                f"                                    'renderChart','draw','render','init'];\n"
                f"                        for (var _fi = 0; _fi < _fns.length; _fi++) {{\n"
                f"                            if (typeof window[_fns[_fi]] === 'function') {{\n"
                f"                                window[_fns[_fi]](); return;\n"
                f"                            }}\n"
                f"                        }}\n"
                f"                    }} catch(e) {{ console.warn('Chart {chart_index} resize redraw:', e); }}\n"
                f"                }}, 150);\n"
                f"            }});\n"
                f"            _ro.observe(_ro_el);\n"
                f"        }})();\n"
            )

            iife = (
                f"// ===== CHART {chart_index} START ({container_id}) =====\n"
                f"(function() {{\n"
                f"    try {{\n"
                f"{indented}\n"
                f"{theme_registration}"
                f"\n        // Post-init resize to ensure correct dimensions\n"
                f"        setTimeout(function() {{\n"
                f"            try {{\n"
                f"                var _c = document.getElementById('{container_id}');\n"
                f"                if (_c && typeof myChart !== 'undefined' && myChart.resize) {{\n"
                f"                    myChart.resize();\n"
                f"                }}\n"
                f"            }} catch(e) {{ console.warn('Chart {chart_index} resize:', e); }}\n"
                f"        }}, 300);\n"
                f"{resize_observer}"
                f"\n"
                f"    }} catch(err) {{\n"
                f"        console.error('❌ Chart {chart_index} error:', err);\n"
                f"    }}\n"
                f"}})();\n"
                f"// ===== CHART {chart_index} END =====\n"
            )

            # ── 10. CDN tags (deduplication handled by caller if needed) ───────
            cdn_tags = "\n".join(f'<script src="{cdn}"></script>' for cdn in cdns)

            # ── 11. Assemble final snippet ─────────────────────────────────────
            # Use overflow:visible + extra bottom padding so axis/stage labels
            # that render outside the SVG drawing area are never clipped.
            label_clearance = 60  # px — enough for two lines of axis text
            snippet = (
                f'<div class="embedded-chart" '
                f'style="width:100%;min-height:{height}px;padding-bottom:{label_clearance}px;'
                f'overflow-x:clip;overflow-y:visible;margin:1.5rem 0;position:relative;">\n'
                f"{scoped_css}"
                f"{cdn_tags}\n"
                f'<div id="{container_id}" style="width:100%;max-width:100%;height:{height}px;overflow:visible;"></div>\n'
                f"{tooltip_div}"
                f"<script>\n{iife}\n</script>\n"
                f"</div>"
            )

            print(f"    ✅ Chart {chart_index} embedded ({len(snippet):,} chars)")
            return snippet

        except Exception as exc:
            import traceback

            print(f"❌ embed_html_chart error: {exc}")
            traceback.print_exc()
            return f'<div style="color:red;padding:1rem;">⚠ Error embedding chart {chart_index}: {exc}</div>'

    def embed_datatable(data_path, kwargs_str, table_index):
        """
        Read a data file and return a fully self-contained HTML datatable snippet
        ready to drop into the report body.

        Accepts any file pandas can read: .parquet, .csv, .xlsx, .json.
        kwargs_str is the raw options string from the EMBED_DATATABLE placeholder,
        e.g. 'title="Word Frequency" max_rows=100 row_density=compact freeze_columns=word,freq'

        Supported kwargs (all optional):
            title            str   — heading shown above the table (default: file stem)
            max_rows         int   — rows visible on initial load (default: 50)
            stored_rows      int   — rows embedded for in-page search (default: 500)
            row_density      str   — normal | compact | ultracompact (default: normal)
            freeze_columns   str   — comma-separated column names to highlight
            table_font_size  str   — CSS font-size (default: 11px)
        """
        import json as _json
        import gzip as _gzip
        import base64 as _base64

        try:
            data_path = Path(data_path)
            if not data_path.exists():
                msg = f"⚠ Data file not found: {data_path}"
                print(f"❌ embed_datatable: {msg}")
                return f'<div style="color:red;padding:1rem;border:1px solid red;">{msg}</div>'

            print(f"\n  📊 embed_datatable [{table_index}]: {data_path.name}")

            # ── 1. Load data ───────────────────────────────────────────────────
            ext = data_path.suffix.lower()
            try:
                import pandas as _pd

                if ext == ".parquet":
                    df = _pd.read_parquet(data_path)
                elif ext == ".csv":
                    df = _pd.read_csv(data_path)
                elif ext in (".xlsx", ".xls"):
                    df = _pd.read_excel(data_path)
                elif ext == ".json":
                    df = _pd.read_json(data_path)
                else:
                    df = _pd.read_csv(data_path)  # best-effort fallback
            except Exception as e:
                msg = f"⚠ Could not read {data_path.name}: {e}"
                print(f"❌ embed_datatable: {msg}")
                return f'<div style="color:red;padding:1rem;border:1px solid red;">{msg}</div>'

            total_rows = len(df)
            total_cols = len(df.columns)

            # ── 2. Parse kwargs string ─────────────────────────────────────────
            # Handles: title="My Table" max_rows=100 row_density=compact freeze_columns=a,b
            kw = {}
            if kwargs_str:
                # Extract quoted title first
                title_m = re.search(r'title=["\']([^"\']+)["\']', kwargs_str)
                if title_m:
                    kw["title"] = title_m.group(1)
                # Then bare key=value pairs (unquoted values)
                for m in re.finditer(r'(\w+)=([^\s"\']+)', kwargs_str):
                    if m.group(1) != "title":
                        kw[m.group(1)] = m.group(2)

            title = kw.get("title", data_path.stem.replace("_", " ").title())
            max_rows = int(kw.get("max_rows", 50))
            stored_rows = int(kw.get("stored_rows", 500))
            row_density = kw.get("row_density", "normal")
            freeze_cols = [
                c.strip() for c in kw.get("freeze_columns", "").split(",") if c.strip()
            ]
            font_size = kw.get("table_font_size", "11px")
            font_family = "'SF Mono', Monaco, 'Inconsolata', 'Fira Code', monospace"

            density_settings = {
                "normal": {"th": "12px 8px", "td": "10px 12px"},
                "compact": {"th": "8px 6px", "td": "4px 6px"},
                "ultracompact": {"th": "4px 4px", "td": "2px 4px"},
            }
            ds = density_settings.get(row_density, density_settings["normal"])

            actual_stored = min(stored_rows, total_rows)
            actual_max_rows = min(max_rows, actual_stored)
            freeze_set = set(c.lower() for c in freeze_cols)
            query_id = f"dt_{table_index}"

            # ── 3. Compress full dataset ───────────────────────────────────────
            payload = {
                "columns": [str(c) for c in df.columns],
                "data": [
                    [
                        None
                        if (isinstance(v, float) and _pd.isna(v))
                        else (v.item() if hasattr(v, "item") else v)
                        for v in row
                    ]
                    for row in df.itertuples(index=False)
                ],
            }
            raw_bytes = _json.dumps(payload, default=str).encode("utf-8")
            compressed = _gzip.compress(raw_bytes, compresslevel=6)
            b64 = _base64.b64encode(compressed).decode("ascii")
            cols_js = "[" + ", ".join(f'"{c}"' for c in df.columns) + "]"

            # ── 4. Analyse columns for display hints ──────────────────────────
            # Classify each column as numeric-float, numeric-int, or text, and
            # assign a CSS width / white-space policy so the table looks good
            # without any manual rounding in the source data.
            _FLOAT_SIG = 4  # significant decimal places to show for floats
            col_meta = {}  # col_name -> {is_float, is_int, width_style, wrap}
            for col in df.columns:
                s = df[col].dropna()
                is_float = _pd.api.types.is_float_dtype(df[col])
                is_int = _pd.api.types.is_integer_dtype(df[col])
                # Estimate how wide the rendered text is by sampling up to 50 rows
                sample_len = (
                    s.head(50).apply(lambda v: len(str(v))).max() if len(s) else 0
                )
                if is_float:
                    col_meta[str(col)] = {
                        "is_float": True,
                        "is_int": False,
                        "width_style": "width:90px;max-width:110px;",
                        "wrap": False,
                    }
                elif is_int:
                    col_meta[str(col)] = {
                        "is_float": False,
                        "is_int": True,
                        "width_style": "width:60px;max-width:80px;",
                        "wrap": False,
                    }
                elif sample_len <= 30:
                    col_meta[str(col)] = {
                        "is_float": False,
                        "is_int": False,
                        "width_style": "max-width:200px;",
                        "wrap": True,
                    }
                else:
                    # Long text column — give it room to breathe but cap it
                    col_meta[str(col)] = {
                        "is_float": False,
                        "is_int": False,
                        "width_style": "max-width:320px;",
                        "wrap": True,
                    }

            def _fmt_cell(val, col_name):
                """Format a cell value for display, respecting format_file config."""
                meta = col_meta.get(str(col_name), {})
                if val is None or (isinstance(val, float) and _pd.isna(val)):
                    return ""
                # Check format_file overrides/globals
                col_lower = str(col_name).lower()
                tbl_lower = data_path.stem.lower()
                _ofk = (tbl_lower, col_lower)
                _fmt = _override_column_formats.get(_ofk) or _global_column_formats.get(col_lower)
                if _fmt:
                    try:
                        if _fmt.startswith("%"):
                            return _pd.to_datetime(val).strftime(_fmt)
                        else:
                            return _fmt.format(val)
                    except Exception:
                        pass
                if meta.get("is_float") and isinstance(val, float):
                    # Show up to _FLOAT_SIG decimal places, strip trailing zeros
                    return f"{val:.{_FLOAT_SIG}f}".rstrip("0").rstrip(".")
                return str(val)

            # ── 5. Build table HTML rows ───────────────────────────────────────
            tbody_html = ""
            df_head = df.head(actual_stored)
            for idx, row in df_head.iterrows():
                row_class = ' class="dt-hidden"' if idx >= actual_max_rows else ""
                tbody_html += f'<tr data-row-id="{idx}"{row_class}>'
                for col, val in zip(df_head.columns, row):
                    col_lower = str(col).lower()
                    fa = ' data-freeze="true"' if col_lower in freeze_set else ""
                    meta = col_meta.get(str(col), {})
                    ws = (
                        "white-space:normal;"
                        if meta.get("wrap")
                        else "white-space:nowrap;"
                    )
                    col_width = meta.get("width_style", "")
                    cell = _fmt_cell(val, col)
                    # Truncate long text for display
                    if len(cell) > 120:
                        escaped = (
                            cell.replace("&", "&amp;")
                            .replace('"', "&quot;")
                            .replace("<", "&lt;")
                            .replace(">", "&gt;")
                        )
                        tbody_html += (
                            f'<td data-col="{col_lower}"{fa} class="dt-trunc" '
                            f'style="{ws}{col_width}" '
                            f'data-full="{escaped}" onclick="dtExpand(this)" title="Click to expand">'
                            f'{cell[:120]}<span style="color:var(--accent-color);font-weight:bold;">…</span></td>'
                        )
                    else:
                        tbody_html += (
                            f'<td data-col="{col_lower}"{fa} '
                            f'style="{ws}{col_width}">{cell}</td>'
                        )
                tbody_html += "</tr>\n"

            # ── 6. Build header row ────────────────────────────────────────────
            thead_html = "<tr>"
            for col in df.columns:
                col_lower = str(col).lower()
                fa = ' data-freeze="true"' if col_lower in freeze_set else ""
                thead_html += (
                    f'<th data-col="{col_lower}"{fa} '
                    f'style="white-space:nowrap;">'
                    f'{col}<span class="dt-sort-ind">⇅</span></th>'
                )
            thead_html += "</tr>"

            row_info = f"Showing {actual_max_rows:,} of {total_rows:,} rows"
            search_info = f"all {total_rows:,}"

            snippet = f"""
<div class="dt-wrapper" id="dt-wrap-{query_id}" style="margin:1.5rem 0;">
<script>
window.dtDataB64_{query_id} = '{b64}';
window.dtCols_{query_id}    = {cols_js};
window.dtTotalRows_{query_id} = {total_rows};
</script>
<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:8px;border-bottom:1px solid var(--border-color);padding-bottom:6px;">
  <span style="font-weight:700;font-size:14px;color:var(--accent-color);">{title}</span>
  <span style="font-size:11px;opacity:0.5;">{total_rows:,} rows · {total_cols} cols</span>
</div>
<div style="display:flex;gap:12px;margin-bottom:8px;">
  <div style="flex:1;display:flex;align-items:center;gap:6px;border-bottom:1px solid var(--border-color);">
    <input type="text" style="flex:1;padding:4px 8px;border:none;background:transparent;color:var(--text-color);font-size:12px;"
           placeholder="▼ Search {search_info} rows..." onkeyup="dtSearchRows('{query_id}',{actual_max_rows})">
    <span id="dt-rcount-{query_id}" style="font-size:11px;opacity:0.5;white-space:nowrap;">{row_info}</span>
  </div>
  <div style="flex:1;display:flex;align-items:center;gap:6px;border-bottom:1px solid var(--border-color);">
    <input type="text" style="flex:1;padding:4px 8px;border:none;background:transparent;color:var(--text-color);font-size:12px;"
           placeholder="🔍 Filter columns..." onkeyup="dtFilterCols('{query_id}')">
    <span id="dt-ccount-{query_id}" style="font-size:11px;opacity:0.5;white-space:nowrap;">Showing {total_cols} of {total_cols} cols</span>
  </div>
</div>
<div style="overflow-x:auto;overflow-y:auto;max-height:500px;border:1px solid var(--border-color);border-radius:6px;
            scrollbar-width:thin;scrollbar-color:rgba(120,120,120,0.3) transparent;">
<table id="dt-{query_id}" style="border-collapse:collapse;font-family:{font_family};font-size:{font_size};">
<thead id="dt-head-{query_id}" style="position:sticky;top:0;z-index:10;background:var(--sidebar-bg);">
{thead_html}
</thead>
<tbody id="dt-body-{query_id}">
{tbody_html}
</tbody>
</table>
</div>
</div>
<style>
#dt-{query_id} th {{
    text-align:left;font-weight:600;cursor:pointer;user-select:none;
    padding:8px 20px 8px 8px;
    border-bottom:2px solid var(--border-color);background:var(--sidebar-bg);
    white-space:nowrap;
}}
#dt-{query_id} th:hover {{ background:var(--hover-color); }}
#dt-{query_id} td {{ padding:{ds["td"]}; border-bottom:1px solid var(--border-color); }}
#dt-{query_id} tbody tr:hover {{ background:var(--hover-color); }}
#dt-{query_id} .dt-sort-ind {{ font-size:10px;margin-left:4px;opacity:0.3; }}
#dt-{query_id} th[data-freeze="true"] {{ background:var(--accent-color);color:#fff; }}
#dt-{query_id} td[data-freeze="true"] {{ background:var(--hover-color);font-weight:500; }}
#dt-{query_id} tr.dt-hidden {{ display:none; }}
#dt-{query_id} td.dt-trunc {{ cursor:pointer;max-width:280px; }}
#dt-{query_id} td.dt-trunc:hover {{ outline:1px solid var(--accent-color);background:var(--hover-color); }}
#dt-{query_id} th.dt-col-hidden, #dt-{query_id} td.dt-col-hidden {{ display:none; }}
</style>
<script>
(function() {{
    // Decompress full dataset once on load
    var _b64 = window.dtDataB64_{query_id};
    if (!_b64) return;
    var binStr = atob(_b64);
    var bytes  = new Uint8Array(binStr.length);
    for (var i = 0; i < binStr.length; i++) bytes[i] = binStr.charCodeAt(i);
    var ds = new DecompressionStream('gzip');
    var w  = ds.writable.getWriter();
    w.write(bytes); w.close();
    new Response(ds.readable).text().then(function(text) {{
        var p = JSON.parse(text);
        var cols = p.columns;
        window.dtFullData_{query_id} = p.data.map(function(row) {{
            var obj = {{}};
            for (var i = 0; i < cols.length; i++) obj[cols[i]] = row[i];
            return obj;
        }});
        window.dtFullCols_{query_id} = cols;
        console.log('✅ datatable ready: {query_id}', window.dtFullData_{query_id}.length, 'rows');
    }});

    // Make sortable
    var table = document.getElementById('dt-{query_id}');
    if (!table) return;
    Array.from(table.querySelectorAll('thead th')).forEach(function(th, ci) {{
        th.addEventListener('click', function() {{
            var tbody = table.querySelector('tbody');
            var rows  = Array.from(tbody.querySelectorAll('tr'));
            var asc   = this.getAttribute('data-sort') !== 'asc';
            Array.from(table.querySelectorAll('thead th')).forEach(function(h) {{ h.removeAttribute('data-sort'); }});
            this.setAttribute('data-sort', asc ? 'asc' : 'desc');
            var ind = this.querySelector('.dt-sort-ind');
            if (ind) {{ ind.textContent = asc ? '⮝' : '⮟'; ind.style.opacity = '1'; }}
            rows.sort(function(a, b) {{
                var ca = a.children[ci] ? a.children[ci].textContent.trim() : '';
                var cb = b.children[ci] ? b.children[ci].textContent.trim() : '';
                var na = parseFloat(ca.replace(/[^0-9.-]/g, '')), nb = parseFloat(cb.replace(/[^0-9.-]/g, ''));
                if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
                return asc ? ca.localeCompare(cb) : cb.localeCompare(ca);
            }});
            rows.forEach(function(r) {{ tbody.appendChild(r); }});
        }});
    }});
}})();
</script>"""

            print(
                f"    ✅ Datatable {table_index} embedded ({total_rows:,} rows, {len(snippet):,} chars)"
            )
            return snippet

        except Exception as exc:
            import traceback

            print(f"❌ embed_datatable error: {exc}")
            traceback.print_exc()
            return f'<div style="color:red;padding:1rem;">⚠ Error embedding datatable {table_index}: {exc}</div>'

    def slugify(title, _seen={}):
        """
        Convert an arbitrary heading title into a safe, unique HTML id.

        Steps:
          1. Unicode-normalise to NFKD and drop combining marks (strips accents,
             decomposes ligatures) so accented letters become ASCII equivalents.
          2. Encode to ASCII with 'ignore' to drop any remaining non-ASCII code-
             points (emojis, CJK glyphs, box-drawing chars, etc.).
          3. Lowercase and collapse every run of non-alphanumeric characters to a
             single hyphen, then strip leading/trailing hyphens.
          4. Fall back to 'section' if the result is empty (all-emoji title, etc.).
          5. Append a numeric suffix (-2, -3, …) when the same slug appears more
             than once in the document so every anchor stays unique.

        The _seen dict is intentionally NOT shared across calls to convert_markdown_to_html
        — it is reset by passing a fresh default-arg dict each call via the wrapper below.
        """
        import unicodedata

        # Normalise → strip combining marks → drop non-ASCII
        nfkd = unicodedata.normalize("NFKD", title)
        ascii_bytes = nfkd.encode("ascii", "ignore")
        text = ascii_bytes.decode("ascii").lower()
        # Replace every run of non-alphanumeric chars with a hyphen
        text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
        if not text:
            text = "section"
        # Uniqueness suffix
        if text not in _seen:
            _seen[text] = 1
            slug = text
        else:
            _seen[text] += 1
            slug = f"{text}-{_seen[text]}"
        return slug

    # Reset the seen-slugs registry for each document conversion
    # by patching the default-arg dict in place.
    slugify.__defaults__[0].clear()  # clears _seen between calls

    def parse_markdown(md_content, base_path):
        """
        Parse markdown content with specific patterns:
        - ## for main sections
        - ### for subsections
        - ``` for code blocks
        - 4-space/tab indented text for reference blocks
        - ![alt](path) for images
        - <!-- EMBED_HTML: <path> | height=<px> --> for inline chart embedding
        """
        sections = []
        current_section = None
        current_subsection = None
        subsection_stack = []  # stack of (depth, node) — enables correct sibling/child nesting
        embed_chart_counter = [0]  # mutable so inner branches can increment it

        embed_html_pat = re.compile(
            r"<!--\s*EMBED_HTML:\s*(.+?)(?:\s*\|\s*height=(\d+))?\s*-->",
            re.IGNORECASE,
        )
        # Fallback: catches ##-add-html lines that leaked into markdown as raw text
        # e.g.  ##-add-html-"C:\path\chart.html" height=400
        #       -add-html-"C:\path\chart.html" height=400   (after ## was stripped)
        raw_add_html_pat = re.compile(
            r'(?:##)?-add-html-["\'](.+?)["\'](?:\s+height=(\d+))?',
            re.IGNORECASE,
        )
        # <!-- EMBED_DATATABLE: path | kwargs --> produced by ##-add-datatables directive
        embed_datatable_pat = re.compile(
            r"<!--\s*EMBED_DATATABLE:\s*(.+?)(?:\s*\|\s*(.*?))?\s*-->",
            re.IGNORECASE,
        )
        # Fallback: catches ##-add-datatables lines that leaked into markdown as raw text
        raw_add_datatables_pat = re.compile(
            r'-add-datatables-["\'](.+?)["\'](?:\s+(.*))?$',
            re.IGNORECASE,
        )
        embed_datatable_counter = [0]

        # ── pre-output marker patterns (from ##-pre-start / ##-pre-end) ──────
        pre_start_pat = re.compile(r"<!--\s*PRE_OUTPUT_START\s*-->", re.IGNORECASE)
        pre_end_pat = re.compile(r"<!--\s*PRE_OUTPUT_END\s*-->", re.IGNORECASE)
        in_pre_block = [False]  # mutable so inner branches can toggle it

        def _is_model_output(text):
            """Heuristic: True if text looks like fixed-width model output.
            Triggered by any line that is 20+ chars of only '=' or '-'."""
            if not monospace_output_heuristic:
                return False
            for line in text.splitlines():
                stripped = line.strip()
                if len(stripped) >= 20 and all(c in ("=", "-") for c in stripped):
                    return True
            return False

        lines = md_content.split("\n")
        i = 0

        while i < len(lines):
            line = lines[i]

            # Pre-check: handle all directive markers before heading detection fires.
            # Covers both the clean <!-- COMMENT --> form and the leaked ### -directive
            # form that appears when a stale .ipynb is reused without rebuild_ipynb=True.
            _line_stripped = line.strip().lstrip("#").strip().lower()

            # ##-pre-start / -pre-start → enter pre-output mode
            if pre_start_pat.search(line) or _line_stripped == "-pre-start":
                in_pre_block[0] = True
                i += 1
                continue
            # ##-pre-end / -pre-end → exit pre-output mode
            if pre_end_pat.search(line) or _line_stripped == "-pre-end":
                in_pre_block[0] = False
                i += 1
                continue
            # ##-run-skip-start/end → these mark code that runs but output is hidden;
            # nothing to render in HTML, just swallow the marker line silently.
            if _line_stripped in ("-run-skip-start", "-run-skip-end"):
                i += 1
                continue
            # ##-skip-start/end → code+output fully excluded; swallow silently.
            if _line_stripped in ("-skip-start", "-skip-end"):
                i += 1
                continue

            # Pre-check: catch ##-add-datatables / ### -add-datatables lines
            # BEFORE heading detection so they're never treated as subsection titles.
            # This handles the case where rebuild_ipynb was not used and the old .ipynb
            # still has the directive as a raw ### heading line.
            _pre_dt = embed_datatable_pat.search(line) or raw_add_datatables_pat.search(
                line
            )
            if _pre_dt and current_section and "-add-datatables-" in line:
                _dt_path = _pre_dt.group(1).strip()
                _dt_kwargs = (_pre_dt.group(2) or "").strip()
                _dt_html = embed_datatable(
                    _dt_path, _dt_kwargs, embed_datatable_counter[0]
                )
                embed_datatable_counter[0] += 1
                target = current_subsection if current_subsection else current_section
                if "content" not in target:
                    target["content"] = []
                target["content"].append({"type": "html", "content": _dt_html})
                i += 1
                continue

            # Inline chart embedding  <!-- EMBED_HTML: path | height=px -->
            # Also catches raw ##-add-html lines that survived into the markdown
            m = embed_html_pat.search(line) or raw_add_html_pat.search(line)
            if m and current_section:
                chart_path = m.group(1).strip()
                chart_height = int(m.group(2)) if m.group(2) else default_embed_height
                chart_html = embed_html_chart(
                    chart_path, chart_height, embed_chart_counter[0]
                )
                embed_chart_counter[0] += 1

                target = current_subsection if current_subsection else current_section
                if "content" not in target:
                    target["content"] = []
                target["content"].append({"type": "html", "content": chart_html})
                i += 1
                continue

            # Inline datatable embedding  <!-- EMBED_DATATABLE: path | kwargs -->
            # Also catches raw ##-add-datatables lines that survived into the markdown
            m_dt = embed_datatable_pat.search(line) or raw_add_datatables_pat.search(
                line
            )
            if m_dt and current_section:
                dt_path = m_dt.group(1).strip()
                dt_kwargs = (m_dt.group(2) or "").strip()
                dt_html = embed_datatable(
                    dt_path, dt_kwargs, embed_datatable_counter[0]
                )
                embed_datatable_counter[0] += 1

                target = current_subsection if current_subsection else current_section
                if "content" not in target:
                    target["content"] = []
                target["content"].append({"type": "html", "content": dt_html})
                i += 1
                continue

            # Main section header (##)
            if line.startswith("## "):
                # Save previous section
                if current_section:
                    sections.append(current_section)

                title = line[3:].strip()
                current_section = {
                    "title": title,
                    "id": slugify(title),
                    "subsections": [],
                    "content": [],
                }
                current_subsection = None
                subsection_stack = []
                i += 1
                continue

            # Subsection headers (### through ######)
            # heading_depth: ### = 3, #### = 4, ##### = 5, ###### = 6
            elif (
                line.startswith("### ")
                or line.startswith("#### ")
                or line.startswith("##### ")
                or line.startswith("###### ")
            ):
                if current_section:
                    # Determine depth: count leading '#' chars
                    heading_depth = len(line) - len(line.lstrip("#"))
                    title = line[heading_depth:].strip()
                    # Safety: if the "heading" is actually an EMBED_HTML placeholder
                    # (can happen if the notebook was built before the ##-directive fix),
                    # treat it as a chart embed rather than a subsection heading.
                    m_sub = embed_html_pat.search(title) or raw_add_html_pat.search(
                        title
                    )
                    if m_sub:
                        chart_path = m_sub.group(1).strip()
                        chart_height = (
                            int(m_sub.group(2))
                            if m_sub.group(2)
                            else default_embed_height
                        )
                        chart_html = embed_html_chart(
                            chart_path, chart_height, embed_chart_counter[0]
                        )
                        embed_chart_counter[0] += 1
                        target = (
                            current_subsection
                            if current_subsection
                            else current_section
                        )
                        if "content" not in target:
                            target["content"] = []
                        target["content"].append(
                            {"type": "html", "content": chart_html}
                        )
                    else:
                        m_sub_dt = embed_datatable_pat.search(
                            title
                        ) or raw_add_datatables_pat.search(title)
                        if m_sub_dt:
                            dt_path = m_sub_dt.group(1).strip()
                            dt_kwargs = (m_sub_dt.group(2) or "").strip()
                            dt_html = embed_datatable(
                                dt_path, dt_kwargs, embed_datatable_counter[0]
                            )
                            embed_datatable_counter[0] += 1
                            target = (
                                current_subsection
                                if current_subsection
                                else current_section
                            )
                            if "content" not in target:
                                target["content"] = []
                            target["content"].append(
                                {"type": "html", "content": dt_html}
                            )
                        else:
                            new_sub = {
                                "title": title,
                                "id": None,  # assigned below after parent is known
                                "depth": heading_depth,
                                "content": [],
                                "subsections": [],
                            }

                            # Pop the stack until we find a node shallower than heading_depth
                            # (that node is the correct parent), or empty (attach to section)
                            while (
                                subsection_stack
                                and subsection_stack[-1][0] >= heading_depth
                            ):
                                subsection_stack.pop()

                            if subsection_stack:
                                parent_node = subsection_stack[-1][1]
                            else:
                                parent_node = None  # attach to section

                            parent_id = (
                                parent_node["id"]
                                if parent_node
                                else current_section["id"]
                            )
                            new_sub["id"] = f"{parent_id}-{slugify(title)}"

                            if parent_node:
                                parent_node["subsections"].append(new_sub)
                            else:
                                current_section["subsections"].append(new_sub)

                            subsection_stack.append((heading_depth, new_sub))
                            current_subsection = new_sub
                i += 1
                continue

            # Code block start
            elif line.startswith("```"):
                if current_section and include_code_blocks:
                    language = line[3:].strip() or "python"
                    code_lines = []
                    i += 1

                    # Collect code lines
                    while i < len(lines) and not lines[i].startswith("```"):
                        code_lines.append(lines[i])
                        i += 1

                    code_content = "\n".join(code_lines)

                    # Add to current subsection or section
                    target = (
                        current_subsection if current_subsection else current_section
                    )
                    if "content" not in target:
                        target["content"] = []

                    target["content"].append(
                        {"type": "code", "language": language, "content": code_content}
                    )

                    i += 1  # Skip closing ```
                    continue
                elif current_section:
                    # Skip code blocks if include_code_blocks is False
                    i += 1
                    while i < len(lines) and not lines[i].startswith("```"):
                        i += 1
                    i += 1  # Skip closing ```
                    continue

            # Reference block (tab or 4-space indented)
            elif line.startswith("    ") or line.startswith("\t"):
                if current_section:
                    ref_lines = []

                    # Collect all consecutive indented lines
                    while i < len(lines) and (
                        lines[i].startswith("    ") or lines[i].startswith("\t")
                    ):
                        # Remove indentation
                        clean_line = lines[i].lstrip(" \t")
                        ref_lines.append(clean_line)
                        i += 1

                    # Join but preserve explicit line breaks
                    ref_content = "\n".join(ref_lines)

                    # Only add reference block if it has actual content (not just whitespace)
                    if ref_content.strip():
                        target = (
                            current_subsection
                            if current_subsection
                            else current_section
                        )
                        if "content" not in target:
                            target["content"] = []

                        # Use model_output type when inside ##-pre-start/end OR heuristic fires
                        block_type = "reference"
                        if in_pre_block[0] or _is_model_output(ref_content):
                            block_type = "model_output"

                        target["content"].append(
                            {"type": block_type, "content": ref_content}
                        )
                    continue

            # HTML table detection (detect <table> tags)
            elif "<table" in line.lower() or "<div" in line.lower():
                if current_section:
                    html_lines = []
                    in_table = True

                    # Track nesting depth to handle nested div/table structures
                    depth = 0

                    # Collect HTML content respecting nesting depth
                    while i < len(lines):
                        current_line = lines[i]
                        html_lines.append(current_line)
                        cl = current_line.lower()

                        # Count opening tags (increase depth)
                        depth += cl.count("<div") + cl.count("<table")
                        # Count closing tags (decrease depth)
                        depth -= cl.count("</div>") + cl.count("</tables>")

                        i += 1

                        # When depth reaches 0, the outermost block is closed
                        if depth <= 0:
                            break

                        i += 1
                        # Safety: if next line doesn't look like HTML continuation, break
                        if i < len(lines) and not (
                            lines[i].strip().startswith("<")
                            or lines[i].strip() == ""
                            or "<" in lines[i]
                            or ">" in lines[i]
                        ):
                            break

                    html_content = "\n".join(html_lines)

                    # Parse HTML table into proper table structure if it's a table
                    if "<table" in html_content.lower():
                        # Apply column formats from format_file if loaded
                        if _global_column_formats or _override_column_formats:
                            try:
                                _soup = BeautifulSoup(html_content, "html.parser")
                                _tbl = _soup.find("table")
                                if _tbl:
                                    _ths = _tbl.find("thead")
                                    _col_names = []
                                    if _ths:
                                        # Get all <th> from the LAST header row (handles multi-row headers)
                                        _header_rows = _ths.find_all("tr")
                                        _last_hdr = _header_rows[-1] if _header_rows else _ths
                                        _all_ths = _last_hdr.find_all("th")
                                        # Pandas adds an empty <th> for the index; count <td> in body
                                        # to determine how many data columns there are
                                        _tbody = _tbl.find("tbody")
                                        _first_body_row = (_tbody.find("tr") if _tbody else None) or _tbl.find_all("tr")[1] if len(_tbl.find_all("tr")) > 1 else None
                                        _n_td = len(_first_body_row.find_all("td")) if _first_body_row else 0
                                        # Take only the last N <th> matching the number of <td> data cells
                                        if _n_td > 0 and len(_all_ths) > _n_td:
                                            _col_names = [th.get_text(strip=True) for th in _all_ths[-_n_td:]]
                                        else:
                                            _col_names = [th.get_text(strip=True) for th in _all_ths]
                                    
                                    if _col_names:
                                        _tbody_el = _tbl.find("tbody")
                                        _body_rows = _tbody_el.find_all("tr") if _tbody_el else _tbl.find_all("tr")[1:]
                                        for _tr in _body_rows:
                                            _tds = _tr.find_all("td")
                                            for _td_idx, _td in enumerate(_tds):
                                                if _td_idx >= len(_col_names):
                                                    break
                                                _cn = _col_names[_td_idx].lower()
                                                _fmt_match = _global_column_formats.get(_cn)
                                                if _fmt_match:
                                                    _raw = _td.get_text(strip=True)
                                                    try:
                                                        if _fmt_match.startswith("%"):
                                                            import pandas as _pd_fmt
                                                            _td.string = _pd_fmt.to_datetime(_raw).strftime(_fmt_match)
                                                        else:
                                                            # Try numeric conversion
                                                            try:
                                                                _num = float(_raw.replace(",", ""))
                                                            except (ValueError, AttributeError):
                                                                _num = _raw
                                                            _td.string = _fmt_match.format(_num)
                                                    except Exception:
                                                        pass
                                    html_content = str(_soup)
                            except Exception:
                                pass

                        # Clean up the HTML table
                        html_content = re.sub(r'style="[^"]*"', "", html_content)
                        html_content = html_content.replace(
                            "<th>",
                            '<th style="text-align: left; padding: 8px; border-bottom: 1px solid var(--border-color);">',
                        )
                        html_content = html_content.replace(
                            "<td>",
                            '<td style="padding: 8px; border-bottom: 1px solid var(--border-color);">',
                        )
                        html_content = html_content.replace(
                            "<thead>",
                            '<thead style="background-color: var(--sidebar-bg);">',
                        )

                    target = (
                        current_subsection if current_subsection else current_section
                    )
                    if "content" not in target:
                        target["content"] = []

                    target["content"].append({"type": "html", "content": html_content})
                    continue

            # Image reference
            elif "![" in line:
                if current_section:
                    img_pattern = r"!\[.*?\]\((.*?)\)"
                    matches = re.findall(img_pattern, line)

                    for img_path in matches:
                        full_path = base_path / img_path
                        if full_path.exists():
                            base64_img = image_to_base64(full_path)
                            if base64_img:
                                target = (
                                    current_subsection
                                    if current_subsection
                                    else current_section
                                )
                                if "content" not in target:
                                    target["content"] = []

                                target["content"].append(
                                    {
                                        "type": "image",
                                        "content": base64_img,
                                        "alt": line.split("[")[1].split("]")[0]
                                        if "[" in line
                                        else "",
                                    }
                                )
                i += 1
                continue

            # Skip orphaned closing HTML tags (leftover from nested div/table wrappers)
            elif line.strip().lower() in ("</div>", "</table>", "</span"):
                i += 1
                continue

            # Regular text
            elif line.strip() and current_section:
                # Collect text content
                text_content = line

                # Check if this is part of a multi-line string (like the ones in your examples)
                i += 1
                while i < len(lines):
                    next_line = lines[i]

                    # Check if we should stop collecting
                    if (
                        next_line.startswith("## ")
                        or next_line.startswith("### ")
                        or next_line.startswith("#### ")
                        or next_line.startswith("##### ")
                        or next_line.startswith("###### ")
                        or next_line.startswith("```")
                        or next_line.startswith("    ")
                        or next_line.startswith("\t")
                        or "<table" in next_line.lower()
                        or "<div" in next_line.lower()
                        or "![" in next_line
                    ):
                        break

                    # If line is empty, it might be a paragraph break
                    if not next_line.strip():
                        # Check if there's more text after this
                        j = i + 1
                        has_more_text = False
                        while j < len(lines):
                            if lines[j].strip() and not (
                                lines[j].startswith("## ")
                                or lines[j].startswith("### ")
                                or lines[j].startswith("#### ")
                                or lines[j].startswith("##### ")
                                or lines[j].startswith("###### ")
                                or lines[j].startswith("```")
                                or lines[j].startswith("    ")
                                or lines[j].startswith("\t")
                                or "<table" in lines[j].lower()
                                or "<div" in lines[j].lower()
                                or "![" in lines[j]
                            ):
                                has_more_text = True
                                break
                            j += 1

                        if has_more_text:
                            # Add a blank line as paragraph separator
                            text_content += "\n\n"
                            i += 1
                            continue
                        else:
                            # No more text, stop here
                            break

                    # Add the line to text content
                    text_content += "\n" + next_line
                    i += 1

                # Process the text content - fix escaped newlines and quotes
                if text_content:
                    # Fix escaped newlines
                    text_content = text_content.replace("\\n", "\n")
                    # Fix escaped quotes
                    text_content = text_content.replace('\\"', '"')
                    text_content = text_content.replace("\\'", "'")
                    # Remove surrounding quotes if they exist
                    if text_content.startswith('"') and text_content.endswith('"'):
                        text_content = text_content[1:-1]
                    if text_content.startswith("'") and text_content.endswith("'"):
                        text_content = text_content[1:-1]

                    # Split by double newlines to get paragraphs
                    paragraphs = text_content.split("\n\n")

                    for para in paragraphs:
                        if para.strip():
                            target = (
                                current_subsection
                                if current_subsection
                                else current_section
                            )
                            if "content" not in target:
                                target["content"] = []

                            target["content"].append(
                                {"type": "text", "content": para.strip()}
                            )

                continue

            # Skip blank lines
            elif not line.strip():
                i += 1
                continue

            i += 1

        # Add the last section
        if current_section:
            sections.append(current_section)

        return sections

    def process_content(content_items, section_id):
        """Process content items and return HTML"""
        html = ""
        code_counter = 0

        for item in content_items:
            content_type = item.get("type")

            if content_type == "text":
                content = item["content"]
                if content:
                    # First, clean up any remaining escaped characters
                    content = (
                        content.replace("\\n", "\n")
                        .replace("\\t", "\t")
                        .replace('\\"', '"')
                        .replace("\\'", "'")
                    )

                    # Split by newlines to create proper paragraphs
                    html += '                    <div class="text-content">\n'

                    # Handle different paragraph separators
                    if "\n\n" in content:
                        paragraphs = content.split("\n\n")
                        for para in paragraphs:
                            if para.strip():
                                # Replace single newlines within paragraphs with <br>
                                lines = [
                                    line.strip()
                                    for line in para.split("\n")
                                    if line.strip()
                                ]
                                if lines:
                                    paragraph_html = "<br>".join(lines)
                                    html += f"                        <p>{paragraph_html}</p>\n"
                    else:
                        # Single paragraph, just replace newlines with <br>
                        lines = [
                            line.strip() for line in content.split("\n") if line.strip()
                        ]
                        if lines:
                            paragraph_html = "<br>".join(lines)
                            html += f"                        <p>{paragraph_html}</p>\n"

                    html += "                    </div>\n"

            elif content_type == "reference":
                # For reference blocks, also clean up escaped characters
                if item["content"].strip():
                    content = item["content"]
                    content = (
                        content.replace("\\n", "\n")
                        .replace("\\t", "\t")
                        .replace('\\"', '"')
                        .replace("\\'", "'")
                    )

                    # Replace newlines with <br> tags
                    lines = [
                        line.strip() for line in content.split("\n") if line.strip()
                    ]
                    if lines:
                        formatted_content = "<br>".join(lines)
                        if style_output_blocks:
                            html += f'                    <div class="reference-block">{formatted_content}</div>\n'
                        else:
                            html += f'                    <div class="output-text">{formatted_content}</div>\n'

            elif content_type == "model_output":
                # Fixed-width model summary — render as <pre> to preserve alignment
                if item["content"].strip():
                    raw = item["content"]
                    # Escape HTML chars so tags inside output don't break the page
                    raw = (
                        raw.replace("&", "&amp;")
                        .replace("<", "&lt;")
                        .replace(">", "&gt;")
                    )
                    css_class = "model-output" + (
                        " styled" if style_output_blocks else ""
                    )
                    html += f'''                    <div class="model-output-container">
                        <pre class="{css_class}">{raw}</pre>
                    </div>
'''

            elif content_type == "code":
                block_id = f"{section_id}-code-{code_counter}"
                code_counter += 1
                collapsed_class = "collapsed" if code_blocks_collapsed else ""
                language = item.get("language", "python").lower()
                html += f'''                    <div class="code-container">
                            <div class="code-content {collapsed_class}" id="{block_id}">
                                <div class="code-header">
                                    <span>{language}</span>
                                    <button class="copy-code-btn" onclick="copyCode('{block_id}')" title="Copy code">copy</button>
                                </div>
                                <pre><code>{item["content"]}</code></pre>
                            </div>
                        </div>
'''

            elif content_type == "image":
                alt_text = item.get("alt", "Image")
                html += f'''                    <div class="image-container">
                            <img src="{item["content"]}" alt="{alt_text}">
                        </div>
'''

            elif content_type == "html":
                html += f"""                    <div class="html-content">
                            {item["content"]}
                        </div>
"""

        return html

    def generate_report_header_html(
        report_header, page_title, page_subtitle, page_description
    ):
        """Render the main content header block.

        If report_header dict is supplied it takes full control — page_title /
        page_subtitle / page_description are ignored.
        If report_header is None the classic title/subtitle/description trio is used.
        """
        if report_header:
            html = '<div class="report-header">\n'
            if report_header.get("title"):
                html += f'    <h1 class="report-title">{report_header["title"]}</h1>\n'
            if report_header.get("subtitle"):
                html += f'    <h2 class="report-subtitle">{report_header["subtitle"]}</h2>\n'
            if report_header.get("metadata"):
                html += '    <div class="report-metadata">\n'
                for key, value in report_header["metadata"].items():
                    html += f'        <span class="metadata-badge"><strong>{key}:</strong> {value}</span>\n'
                html += "    </div>\n"
            if report_header.get("description"):
                html += '    <div class="report-description">\n'
                for para in report_header["description"].split("\n\n"):
                    if para.strip():
                        html += f"        <p>{para.strip()}</p>\n"
                html += "    </div>\n"
            html += "</div>\n"
            return html
        else:
            # Classic fallback
            html = '<div class="report-header">\n'
            html += f'    <div class="report-title">{page_title}</div>\n'
            if page_subtitle:
                html += f'    <div class="report-subtitle">{page_subtitle}</div>\n'
            if page_description:
                html += (
                    f'    <div class="report-description">{page_description}</div>\n'
                )
            html += "</div>\n"
            return html

    def generate_report_footer_html(report_footer, footer_content, timestamp):
        """Render the footer block.

        If report_footer dict is supplied it takes full control — footer_content
        and the auto-timestamp are still appended unless report_footer explicitly
        sets "hide_timestamp": True.
        If report_footer is None the classic footer_content + timestamp is used.
        """
        html = '<div class="report-footer">\n'
        if report_footer:
            if report_footer.get("title"):
                html += f'    <h3 class="footer-title">{report_footer["title"]}</h3>\n'
            if report_footer.get("content"):
                html += '    <div class="footer-content">\n'
                for para in report_footer["content"].split("\n\n"):
                    if para.strip():
                        html += f"        <p>{para.strip()}</p>\n"
                html += "    </div>\n"
            if not report_footer.get("hide_timestamp"):
                html += f'    <div class="footer-timestamp">Generated on {timestamp}</div>\n'
        else:
            if footer_content:
                html += (
                    f'    <div class="footer-content"><p>{footer_content}</p></div>\n'
                )
            html += (
                f'    <div class="footer-timestamp">Generated on {timestamp}</div>\n'
            )
        html += "</div>\n"
        return html

    def generate_html(sections):
        """Generate complete HTML with embedded styles and content"""

        # ── Build underline_headings CSS ───────────────────────────────────────
        def _build_underline_css(uh):
            """Generate border-bottom CSS for each configured heading level."""
            if not uh:
                return ""
            css_lines = []
            for level, opts in uh.items():
                level = level.lower().strip()
                width = opts.get("line_width", "1px")
                color = opts.get("line_color", "var(--border-color)")
                style_ = opts.get("line_style", "solid")
                border = f"{width} {style_} {color}"
                if level == "h2":
                    css_lines.append(
                        f"        .section-title {{\n"
                        f"            border-bottom: {border};\n"
                        f"            padding-bottom: 16px;\n"
                        f"            margin-bottom: 24px;\n"
                        f"        }}"
                    )
                elif level in ("h3", "h4", "h5", "h6"):
                    css_lines.append(
                        f"        {level}.subsection-title {{\n"
                        f"            border-bottom: {border};\n"
                        f"            padding-bottom: 10px;\n"
                        f"            margin-bottom: 18px;\n"
                        f"        }}"
                    )
            return "\n\n".join(css_lines)

        _underline_css = _build_underline_css(underline_headings)

        # ── Nav font sizes (per heading depth) ───────────────────────────────
        _default_nav_sizes = {2: "13px", 3: "12px", 4: "11px", 5: "11px", 6: "10px"}
        _nav_sizes = {**_default_nav_sizes, **(nav_font_sizes or {})}

        def _build_nav_font_css(sizes):
            """Return CSS rules giving each depth its own font-size in the sidebar."""
            lines = []
            # depth-2 = top-level .nav-link (no subsection-link class)
            lines.append(
                f"        .nav-link:not(.subsection-link) {{\n"
                f"            font-size: {sizes[2]};\n"
                f"            font-weight: 600;\n"
                f"        }}"
            )
            for depth in range(3, 7):
                size = sizes.get(depth, sizes.get(6, "10px"))
                weight = "500" if depth == 3 else "400"
                opacity = "" if depth <= 3 else f"\n            opacity: {max(0.65, 1.0 - (depth - 3) * 0.1):.2f};"
                lines.append(
                    f"        .nav-link.subsection-link[data-depth=\"{depth}\"] {{\n"
                    f"            font-size: {size};\n"
                    f"            font-weight: {weight};{opacity}\n"
                    f"        }}"
                )
            return "\n\n".join(lines)

        _nav_font_css = _build_nav_font_css(_nav_sizes)

        # ── Heading font sizes ───────────────────────────────────────────────
        _heading_css = ""
        if heading_font_sizes:
            for level, size in heading_font_sizes.items():
                _heading_css += (
                    f"        .section-title[data-level=\"{level}\"],\n"
                    f"        h{level}.section-title {{\n"
                    f"            font-size: {size};\n"
                    f"        }}\n"
                )

        # ── Page font / font-size ────────────────────────────────────────────
        _body_font_family = page_font if page_font else (
            "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif"
        )
        _body_font_size = page_font_size if page_font_size else "14px"

        # Generate timestamp for footer
        timestamp = datetime.now().strftime("%B %d, %Y at %I:%M %p")

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{page_title}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        :root {{
            --bg-color: {light_mode_colors["bg"]};
            --text-color: {light_mode_colors["text"]};
            --sidebar-bg: {light_mode_colors["sidebar_bg"]};
            --accent-color: {light_mode_colors["accent"]};
            --border-color: {light_mode_colors["border"]};
            --hover-color: {light_mode_colors["hover"]};
            --code-bg: {light_mode_colors["code_bg"]};
            --reference-bg: {light_mode_colors["reference_bg"]};
            --reference-border: {light_mode_colors["reference_border"]};
        }}

        [data-theme="dark"] {{
            --bg-color: {dark_mode_colors["bg"]};
            --text-color: {dark_mode_colors["text"]};
            --sidebar-bg: {dark_mode_colors["sidebar_bg"]};
            --accent-color: {dark_mode_colors["accent"]};
            --border-color: {dark_mode_colors["border"]};
            --hover-color: {dark_mode_colors["hover"]};
            --code-bg: {dark_mode_colors["code_bg"]};
            --reference-bg: {dark_mode_colors["reference_bg"]};
            --reference-border: {dark_mode_colors["reference_border"]};
        }}

        body {{
            font-family: {_body_font_family};
            font-size: {_body_font_size};
            background-color: var(--bg-color);
            color: var(--text-color);
            padding-left: {'40px' if sidebar_collapsible else sidebar_width};
            transition: background-color 0.3s, color 0.3s;
            line-height: 1.6;
            margin: 0;
            box-sizing: border-box;
            overflow-x: hidden;
        }}

        /* ── Modern slim scrollbars (light + dark mode) ── */
        * {{
            scrollbar-width: thin;
            scrollbar-color: rgba(128,128,128,0.35) transparent;
        }}
        *::-webkit-scrollbar {{ width: 6px; height: 6px; }}
        *::-webkit-scrollbar-track {{ background: transparent; }}
        *::-webkit-scrollbar-thumb {{
            background: rgba(128,128,128,0.35);
            border-radius: 3px;
        }}
        *::-webkit-scrollbar-thumb:hover {{
            background: rgba(128,128,128,0.6);
        }}
        /* Dark mode: slightly brighter thumb so it's visible on dark backgrounds */
        [data-theme="dark"] *::-webkit-scrollbar-thumb {{
            background: rgba(200,200,200,0.25);
        }}
        [data-theme="dark"] *::-webkit-scrollbar-thumb:hover {{
            background: rgba(200,200,200,0.5);
        }}
        [data-theme="dark"] * {{
            scrollbar-color: rgba(200,200,200,0.25) transparent;
        }}

        /* Embedded chart wrapper: clip horizontal overflow at the wrapper level
           while keeping the inner container overflow:visible for axis labels */
        .embedded-chart {{
            overflow-x: clip;
        }}

        .sidebar {{
            position: fixed;
            top: 0;
            left: 0;
            width: {sidebar_width};
            height: 100vh;
            background-color: var(--sidebar-bg);
            padding: 15px 12px;
            overflow-y: auto;
            box-shadow: 2px 0 5px rgba(0, 0, 0, 0.1);
            z-index: 1000;
            transition: background-color 0.3s;
        }}

        /* Scrollbar appears only on hover */
        .sidebar {{
            scrollbar-width: thin;
            scrollbar-color: transparent transparent;
            transition: scrollbar-color 0.3s ease;
        }}

        .sidebar:hover {{
            scrollbar-color: rgba(120, 120, 120, 0.3) transparent;
        }}

        .sidebar::-webkit-scrollbar {{
            width: 8px;
        }}

        .sidebar::-webkit-scrollbar-track {{
            background: transparent;
        }}

        .sidebar::-webkit-scrollbar-thumb {{
            background: transparent;
            border-radius: 10px;
            transition: background 0.2s ease;
        }}

        .sidebar:hover::-webkit-scrollbar-thumb {{
            background: rgba(120, 120, 120, 0.3);
        }}

        .sidebar:hover::-webkit-scrollbar-thumb:hover {{
            background: rgba(120, 120, 120, 0.5);
        }}

        [data-theme="dark"] .sidebar:hover {{
            scrollbar-color: rgba(200, 200, 200, 0.2) transparent;
        }}

        [data-theme="dark"] .sidebar:hover::-webkit-scrollbar-thumb {{
            background: rgba(200, 200, 200, 0.2);
        }}

        [data-theme="dark"] .sidebar:hover::-webkit-scrollbar-thumb:hover {{
            background: rgba(200, 200, 200, 0.35);
        }}

        /* — collapsible sidebar: floating nav rail + popover panel — */
        .sidebar-collapsed-strip {{
            display: {'flex' if sidebar_collapsible else 'none'};
            position: fixed;
            top: 0;
            left: 0;
            width: 40px;
            height: 100vh;
            flex-direction: column;
            align-items: center;
            padding-top: 14px;
            z-index: 1001;
            pointer-events: none;
        }}

        .sidebar-collapsed-strip > * {{
            pointer-events: auto;
        }}

        .sidebar-contents-badge {{
            background: var(--sidebar-bg);
            color: var(--text-color);
            font-size: 9px;
            font-weight: 700;
            letter-spacing: 1.2px;
            text-transform: uppercase;
            padding: 6px 4px;
            border-radius: 4px;
            cursor: pointer;
            user-select: none;
            writing-mode: vertical-rl;
            text-orientation: mixed;
            transform: rotate(180deg);
            border: 1px solid var(--border-color);
            opacity: 0.85;
            transition: opacity 0.2s, background 0.2s;
            margin-bottom: 8px;
        }}

        .sidebar-contents-badge:hover {{
            opacity: 1;
            background: var(--hover-color);
        }}

        .sidebar-nav-dots {{
            display: flex;
            flex-direction: column;
            align-items: center;
            gap: 6px;
            margin-top: 4px;
        }}

        .sidebar-nav-dot {{
            width: 7px;
            height: 7px;
            border-radius: 50%;
            background: var(--border-color);
            cursor: pointer;
            transition: background 0.2s, transform 0.2s;
            position: relative;
        }}

        .sidebar-nav-dot:hover {{
            background: var(--accent-color);
            transform: scale(1.4);
        }}

        .sidebar-nav-dot.active {{
            background: var(--accent-color);
            transform: scale(1.3);
        }}

        .sidebar-nav-dot-tooltip {{
            display: none;
            position: absolute;
            left: 18px;
            top: 50%;
            transform: translateY(-50%);
            background: var(--sidebar-bg);
            color: var(--text-color);
            border: 1px solid var(--border-color);
            border-radius: 4px;
            padding: 3px 8px;
            font-size: 11px;
            white-space: nowrap;
            pointer-events: none;
            z-index: 1003;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
        }}

        .sidebar-nav-dot:hover .sidebar-nav-dot-tooltip {{
            display: block;
        }}

        .sidebar-backdrop {{
            display: none;
            position: fixed;
            inset: 0;
            z-index: 999;
        }}

        .sidebar-backdrop.active {{
            display: block;
        }}

        .sidebar.collapsible {{
            position: fixed;
            top: 10px;
            left: 44px;
            width: min({sidebar_width}, calc(100vw - 60px));
            height: auto;
            max-height: calc(100vh - 20px);
            border-radius: 10px;
            border: 1px solid var(--border-color);
            box-shadow: 0 8px 32px rgba(0,0,0,0.18);
            transform: scale(0.95);
            opacity: 0;
            pointer-events: none;
            z-index: 1002;
            transition: transform 0.2s cubic-bezier(0.4,0,0.2,1), opacity 0.2s;
            overflow-y: auto;
        }}

        .sidebar.collapsible.open {{
            transform: scale(1);
            opacity: 1;
            pointer-events: auto;
        }}
        
        .sidebar-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            padding-bottom: 12px;
            //border-bottom: 1px solid var(--border-color);
        }}

        .sidebar-title {{
            font-size: 14px;
            font-weight: bold;
            color: var(--accent-color);
        }}

        /* ── Search box with buttons below ── */
        .search-box {{
            margin-bottom: 12px;
        }}

        .search-wrap {{
            position: relative;
            display: flex;
            align-items: center;
        }}

        .search-input {{
            flex: 1;
            min-width: 0;
            width: 100%;
            padding: 7px 26px 7px 10px;
            border: 1px solid var(--border-color);
            border-radius: 4px;
            font-size: 12px;
            background-color: var(--bg-color);
            color: var(--text-color);
            transition: all 0.2s;
        }}

        .search-input:focus {{
            outline: none;
            border-color: var(--accent-color);
            box-shadow: 0 0 0 2px rgba(0, 123, 255, 0.1);
        }}

        .search-clear {{
            position: absolute;
            right: 6px;
            background: none;
            border: none;
            color: var(--text-color);
            cursor: pointer;
            font-size: 13px;
            padding: 2px 3px;
            opacity: 0;
            transition: opacity 0.15s;
            line-height: 1;
            z-index: 1;
        }}
        .search-clear.visible {{ opacity: 0.6; }}
        .search-clear:hover {{ opacity: 1 !important; }}

        /* ── Icon group: below the search input ── */
        .search-icon-group {{
            display: flex;
            align-items: center;
            gap: 2px;
            flex-shrink: 0;
            margin-top: 4px;
        }}

        .search-icon-btn {{
            background: none;
            border: none;
            color: var(--text-color);
            cursor: pointer;
            padding: 4px 5px;
            border-radius: 4px;
            opacity: 0.5;
            transition: opacity 0.15s, background 0.15s;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            line-height: 1;
        }}
        .search-icon-btn:hover {{
            opacity: 1;
            background: var(--hover-color);
        }}

        /* ── Help tooltip — opens below the search row ── */
        .help-tooltip {{
            display: none;
            background: var(--sidebar-bg);
            border: 1px solid var(--border-color);
            border-radius: 6px;
            padding: 8px 10px;
            margin-top: 6px;
            font-size: 11px;
            color: var(--text-color);
        }}
        .help-tooltip.visible {{ display: block; }}
        .help-row {{
            display: flex;
            align-items: center;
            gap: 5px;
            margin: 3px 0;
        }}
        .help-muted {{ opacity: 0.55; font-size: 10px; }}
        kbd {{
            display: inline-block;
            padding: 1px 5px;
            border: 1px solid var(--border-color);
            border-radius: 3px;
            font-size: 10px;
            font-family: monospace;
            background: var(--hover-color);
            color: var(--text-color);
            line-height: 1.5;
        }}

        .search-count {{
            font-size: 11px;
            color: var(--text-color);
            margin-top: 4px;
            display: block;
            opacity: 0.7;
        }}

        /* ── Collapsible nav groups ── */
        .nav-group {{
            display: flex;
            flex-direction: column;
        }}
        .nav-group-header {{
            display: flex;
            align-items: center;
            gap: 0;
            border-radius: 4px;
            transition: background-color 0.15s;
            cursor: pointer;
            user-select: none;
        }}
        .nav-group-header:hover {{
            background-color: var(--hover-color);
        }}
        .nav-group-children {{
            display: block;
        }}
        .nav-group-children.collapsed {{
            display: none;
        }}
        .nav-toggle-icon {{
            display: inline-flex;
            align-items: center;
            padding: 2px 3px 2px 4px;
            flex-shrink: 0;
            opacity: 0.45;
            transition: opacity 0.15s;
            fill: var(--text-color);
        }}
        .nav-group-header:hover .nav-toggle-icon {{ opacity: 0.9; }}

        .nav-links {{
            display: flex;
            flex-direction: column;
            gap: 1px;
        }}

        .nav-link {{
            color: var(--text-color);
            text-decoration: none;
            padding: 6px 10px 6px 6px;
            border-radius: 4px;
            border-left: 2px solid transparent;
            transition: color 0.15s, border-color 0.15s;
            display: block;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            flex: 1;
            min-width: 0;
        }}

        .nav-link:hover {{
            color: var(--accent-color);
        }}

        .nav-group-header:hover .nav-link {{
            color: var(--accent-color);
        }}

        .nav-link.hidden {{
            display: none;
        }}

        .nav-link.subsection-link {{
            padding-left: 20px;
            opacity: 0.9;
        }}

        .nav-link.subsection-link[data-depth="4"] {{ padding-left: 32px; }}
        .nav-link.subsection-link[data-depth="5"] {{ padding-left: 44px; }}
        .nav-link.subsection-link[data-depth="6"] {{ padding-left: 56px; }}

        __NAV_FONT_CSS_PLACEHOLDER__

        .nav-link.active {{
            color: var(--accent-color);
            font-weight: 600;
            border-left-color: var(--accent-color);
        }}

        .nav-link.subsection-link.active {{
            color: var(--accent-color);
            border-left-color: var(--accent-color);
            opacity: 1;
        }}

        .main-content {{
            padding: 0;
            max-width: 1400px;
            margin: 0 auto;
            width: 100%;
        }}

        .report-header {{
            padding: 15px 30px 30px 30px;
            margin-bottom: 30px;
            color: var(--text-color);
        }}

        .report-title {{
            font-size: 28px;
            font-weight: bold;
            margin-bottom: 8px;
            color: var(--text-color);
        }}

        .report-subtitle {{
            font-size: 18px;
            font-weight: normal;
            margin-bottom: 15px;
            color: var(--text-color);
            opacity: 0.8;
        }}

        .report-metadata {{
            display: flex;
            flex-direction: column;
            gap: 6px;
            margin-bottom: 15px;
            padding: 8px 0;
        }}

        .metadata-badge {{
            background-color: transparent;
            color: var(--text-color);
            padding: 0;
            font-size: 14px;
        }}

        .metadata-badge strong {{
            font-weight: 600;
            color: var(--text-color);
        }}

        .report-description {{
            font-size: 14px;
            line-height: 1.6;
            color: var(--text-color);
            opacity: 0.85;
            max-width: 800px;
        }}

        .report-description p {{
            margin-bottom: 12px;
        }}

        .report-description p:last-child {{
            margin-bottom: 0;
        }}

        .content-area {{
            padding: 0 30px 30px 30px;
            overflow: hidden;
        }}

        .section {{
            margin-bottom: 40px;
            scroll-margin-top: 20px;
        }}

        .section-title {{
            font-size: 24px;
            color: var(--accent-color);
            margin-bottom: 20px;
            padding-bottom: 8px;
            //border-bottom: 1px solid var(--border-color);
            position: relative;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}

        .section-title-text {{
            flex: 1;
        }}

        .code-buttons {{
            display: flex;
            gap: 4px;
            opacity: 0;
            transition: opacity 0.2s;
            margin-left: 10px;
        }}

        .code-buttons:hover {{
            opacity: 1;
        }}

        .subsection {{
            margin-bottom: 25px;
            padding-left: 15px;
            //border-left: 2px solid var(--border-color);
            padding-left: 0;
        }}

        .subsection-title {{
            font-size: 18px;
            color: var(--text-color);
            margin-bottom: 12px;
            font-weight: 600;
            position: relative;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}

        .subsection-title-text {{
            flex: 1;
        }}

        {_underline_css}

        __HEADING_CSS_PLACEHOLDER__

        .content-block {{
            margin-bottom: 15px;
        }}

        .text-content {{
            color: var(--text-color);
            margin: 12px 0;
            line-height: 1.7;
        }}

        .text-content p {{
            margin-bottom: 16px;
            padding-bottom: 8px;
        }}

        .text-content p:last-child {{
            margin-bottom: 0;
        }}

        .reference-block {{
            //background-color: var(--reference-bg);
            border-left: 3px solid var(--reference-border);
            padding: 12px 15px;
            margin: 15px 0;
            border-radius: 4px;
            font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
            font-size: 12px;
            line-height: 1.6;
            white-space: pre-wrap;
            color: var(--text-color);
        }}

        .output-text {{
            margin: 15px 0;
            font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
            font-size: 12px;
            line-height: 1.6;
            white-space: pre-wrap;
            color: var(--text-color);
        }}

        .model-output-container {{
            margin: 15px 0;
            overflow-x: auto;
        }}

        .model-output {{
            font-family: 'Courier New', 'Courier', monospace;
            font-size: 12.5px;
            line-height: 1.55;
            white-space: pre;
            letter-spacing: 0;
            word-spacing: 0;
            color: var(--text-color);
            margin: 0;
            padding: 0;
            background: transparent;
            border: none;
            tab-size: 8;
        }}

        .model-output.styled {{
            border-left: 3px solid var(--reference-border);
            padding: 12px 15px;
            border-radius: 4px;
        }}

        .code-container {{
            position: relative;
            margin: 15px 0;
        }}

        .code-content {{
            border: 1px solid var(--border-color);
            border-radius: 6px;
            overflow: hidden;
            margin-top: 8px;
            background-color: var(--code-bg);
        }}

        .code-content.collapsed {{
            display: none;
        }}

        .code-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 5px 12px;
            background-color: var(--code-bg);
            border-bottom: 1px solid var(--border-color);
            font-size: 10px;
            font-weight: 600;
            letter-spacing: 0.07em;
            text-transform: uppercase;
            opacity: 0.55;
        }}

        .copy-code-btn {{
            background: none;
            border: none;
            cursor: pointer;
            color: var(--text-color);
            opacity: 0.7;
            padding: 2px 6px;
            border-radius: 3px;
            font-size: 10px;
            letter-spacing: 0.04em;
            transition: opacity 0.2s;
        }}
        .copy-code-btn:hover {{ opacity: 1; }}

        .code-btn {{
            background: none;
            border: none;
            color: var(--text-color);
            cursor: pointer;
            padding: 4px;
            border-radius: 3px;
            font-size: 14px;
            transition: all 0.2s;
            line-height: 1;
            display: flex;
            align-items: center;
            justify-content: center;
            width: 24px;
            height: 24px;
        }}

        .code-btn:hover {{
            background-color: var(--hover-color);
        }}

        .code-btn svg {{
            width: 14px;
            height: 14px;
            fill: currentColor;
        }}

        pre {{
            margin: 0;
            padding: 12px 16px;
            overflow-x: auto;
            overflow-y: auto;
            max-height: {default_code_height};
            font-family: 'Consolas', 'SF Mono', 'Monaco', 'Courier New', monospace;
            font-size: 12px;
            line-height: 1.5;
            scrollbar-width: thin;
            scrollbar-color: rgba(120,120,120,0.3) transparent;
        }}

        pre::-webkit-scrollbar {{ width: 4px; height: 4px; }}
        pre::-webkit-scrollbar-track {{ background: transparent; }}
        pre::-webkit-scrollbar-thumb {{
            border-radius: 999px;
            background: rgba(120,120,120,0.25);
        }}
        [data-theme="dark"] pre::-webkit-scrollbar-thumb {{
            background: rgba(200,200,200,0.18);
        }}

        code {{
            color: var(--text-color);
        }}

        .image-container {{
            margin: 20px 0;
            text-align: center;
        }}

        .image-container img {{
            max-width: 100%;
            height: auto;
            border-radius: 6px;
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
            border: 1px solid var(--border-color);
        }}

        .html-content {{
            margin: 20px 0;
            overflow-x: auto;
        }}

        .html-content table {{
            border-collapse: collapse;
            width: auto;
            min-width: 40%;
            margin: 15px 0;
            font-size: 13px;
        }}

        .html-content th {{
            background-color: var(--sidebar-bg);
            color: var(--text-color);
            text-align: left;
            padding: 10px;
            border-bottom: 2px solid var(--border-color);
            font-weight: 600;
        }}

        .html-content td {{
            padding: 10px;
            border-bottom: 1px solid var(--border-color);
            color: var(--text-color);
        }}

        .html-content tr:hover {{
            background-color: var(--hover-color);
        }}

        .html-content .dataframe {{
            width: 100%;
            margin: 15px 0;
        }}

        .report-footer {{
            background-color: var(--bg-color);
            padding: 30px 30px;
            margin-top: 50px;
            border-top: 2px solid var(--border-color);
        }}

        .footer-title {{
            font-size: 18px;
            font-weight: bold;
            margin-bottom: 8px;
            color: var(--text-color);
        }}

        .footer-content {{
            font-size: 14px;
            line-height: 1.6;
            color: var(--text-color);
            opacity: 0.8;
        }}

        .footer-content p {{
            margin-bottom: 10px;
        }}

        .footer-content p:last-child {{
            margin-bottom: 0;
        }}

        .footer-timestamp {{
            opacity: 0.6;
            font-style: italic;
            font-size: 12px;
            margin-top: 10px;
            color: var(--text-color);
        }}

        @media (max-width: 768px) {{
            body {{
                padding-left: {'40px' if sidebar_collapsible else '0'};
                padding-top: {'0' if sidebar_collapsible else '60px'};
            }}

            .sidebar:not(.collapsible) {{
                width: 100%;
                height: auto;
                position: relative;
            }}

            .report-header {{
                padding: 25px 20px;
            }}

            .report-title {{
                font-size: 24px;
            }}

            .content-area {{
                padding: 0 20px 20px 20px;
            }}

            .subsection {{
                padding-left: 10px;
            }}

            .code-buttons {{
                opacity: 1; /* Always show on mobile for easier interaction */
            }}
        }}
    </style>
</head>
<body data-theme="{default_theme}">
""" + (f"""    <div class="sidebar-backdrop" id="sidebar-backdrop" onclick="closeSidebar()"></div>
    <div class="sidebar-collapsed-strip" id="sidebar-strip">
        <span class="sidebar-contents-badge" onclick="toggleSidebar()">Index</span>
        <div class="sidebar-nav-dots" id="sidebar-nav-dots"></div>
    </div>
""" if sidebar_collapsible else "") + f"""    <div class="sidebar{' collapsible' if sidebar_collapsible else ''}">
        <div class="sidebar-header">
            <div class="sidebar-title">{navigation_title}</div>""" + (f"""
            <button class="sidebar-hamburger" onclick="closeSidebar()" title="Close navigation (S)" style="margin-left:auto;background:none;border:none;color:var(--text-color);cursor:pointer;padding:4px 6px;border-radius:4px;opacity:0.5;transition:opacity 0.15s;">
                <svg width="16" height="16" viewBox="0 0 24 24" fill="currentColor"><path d="M19 6.41L17.59 5 12 10.59 6.41 5 5 6.41 10.59 12 5 17.59 6.41 19 12 13.41 17.59 19 19 17.59 13.41 12z"/></svg>
            </button>""" if sidebar_collapsible else "") + """
        </div>
        <div class="search-box">
            <div class="search-wrap">
                <input type="text"
                       class="search-input"
                       id="search-input"
                       placeholder="Search… (F)"
                       title="Press F to focus · Esc to clear · ↑↓ navigate"
                       onkeyup="filterSections()"
                       oninput="toggleClearButton()">
                <button class="search-clear" id="search-clear" onclick="clearSearch()" title="Clear search (Esc)">×</button>
            </div>
            <div class="search-icon-group">
                <button class="search-icon-btn" title="Expand all" onclick="expandAll()">
                    <svg width="14" height="14" viewBox="0 0 24 24" fill="currentColor">
                        <path d="M19 13h-6v6h-2v-6H5v-2h6V5h2v6h6v2z"/>
                    </svg>
                </button>
                <button class="search-icon-btn" title="Collapse all" onclick="collapseAll()">
                    <svg width="14" height="14" viewBox="0 0 24 24" fill="currentColor">
                        <path d="M19 13H5v-2h14v2z"/>
                    </svg>
                </button>
                <button class="search-icon-btn" id="themeToggleBtn" title="Toggle light / dark theme (T)" onclick="toggleTheme()">
                    <svg id="theme-icon" width="14" height="14" viewBox="0 0 24 24" fill="currentColor">
                        <path d="M12 7a5 5 0 0 1 5 5 5 5 0 0 1-5 5 5 5 0 0 1-5-5 5 5 0 0 1 5-5m0-2a7 7 0 0 0-7 7 7 7 0 0 0 7 7 7 7 0 0 0 7-7 7 7 0 0 0-7-7M2 11h2v2H2v-2m18 0h2v2h-2v-2M11 2h2v2h-2V2m0 18h2v2h-2v-2M4.22 3.93l1.42 1.42-1.42 1.41-1.41-1.41 1.41-1.42m15.14 13.3 1.41 1.41-1.41 1.42-1.42-1.42 1.42-1.41M4.22 19.07l-1.41-1.42 1.41-1.41 1.42 1.41-1.42 1.42M19.36 5.36l-1.42-1.42 1.42-1.41 1.41 1.42-1.41 1.41z"/>
                    </svg>
                </button>
                <button class="search-icon-btn" id="helpBtn" title="Keyboard shortcuts" onclick="toggleHelp()">
                    <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                        <path d="M8.5 8.5A3.5 3.5 0 1 1 12 12v1"/>
                        <path d="M12 17h.01"/>
                    </svg>
                </button>
            </div>
            <div class="help-tooltip" id="helpTooltip">
                <div class="help-row"><kbd>F</kbd> Focus search box</div>
                <div class="help-row"><kbd>↑</kbd><kbd>↓</kbd> Navigate sidebar links</div>
                <div class="help-row"><kbd>T</kbd> Toggle light / dark</div>
                <div class="help-row"><kbd>Esc</kbd> Clear search / close</div>
                <div class="help-row"><kbd>S</kbd> Toggle sidebar open/close</div>
                <div class="help-row"><kbd>N</kbd> Focus navigation panel</div>
                <div class="help-row"><kbd>M</kbd> Focus main content</div>
            </div>
            <span class="search-count" id="search-count"></span>
        </div>

        <div class="nav-links" id="nav-links">
"""

        # icon total footprint: span padding-left(4) + svg(10) + span padding-right(3) = 17px
        # nav-link inside the row has padding-left:6px, so leaf offset = 17 + 6 = 23px
        _nav_icon_offset = 23

        def render_nav_links(subsections, depth_offset=0):
            nav_html = ""
            for subsection in subsections:
                subsection_id = subsection.get("id", "")
                depth = subsection.get("depth", 3)
                base_pad = 6 + depth_offset * 12
                children = subsection.get("subsections", [])
                if children:
                    grp_id = f"navgrp-{subsection_id}"
                    nav_html += (
                        f'            <div class="nav-group" id="ng-{subsection_id}">\n'
                        f'              <div class="nav-group-header" style="padding-left:{base_pad}px" onclick="handleNavRowClick(event, \'{grp_id}\')">\n'
                        f'                <span class="nav-toggle-icon"><svg id="ngi-{grp_id}" viewBox="0 0 24 24" width="10" height="10" fill="currentColor"><path d="M19 13H5v-2h14v2z"/></svg></span>\n'
                        f'                <a href="#{subsection_id}" class="nav-link subsection-link" data-section-id="{subsection_id}" data-depth="{depth}">{subsection["title"]}</a>\n'
                        f'              </div>\n'
                        f'              <div class="nav-group-children" id="{grp_id}">\n'
                    )
                    nav_html += render_nav_links(children, depth_offset + 1)
                    nav_html += '              </div>\n            </div>\n'
                else:
                    leaf_pad = base_pad + _nav_icon_offset
                    nav_html += f'            <a href="#{subsection_id}" class="nav-link subsection-link" style="padding-left:{leaf_pad}px" data-section-id="{subsection_id}" data-depth="{depth}">{subsection["title"]}</a>\n'
            return nav_html

        for section in sections:
            sec_id = section["id"]
            subsections = section.get("subsections", [])
            if subsections:
                grp_id = f"navgrp-{sec_id}"
                html += (
                    f'            <div class="nav-group" id="ng-{sec_id}">\n'
                    f'              <div class="nav-group-header" onclick="handleNavRowClick(event, \'{grp_id}\')">\n'
                    f'                <span class="nav-toggle-icon"><svg id="ngi-{grp_id}" viewBox="0 0 24 24" width="10" height="10" fill="currentColor"><path d="M19 13H5v-2h14v2z"/></svg></span>\n'
                    f'                <a href="#{sec_id}" class="nav-link" data-section-id="{sec_id}">{section["title"]}</a>\n'
                    f'              </div>\n'
                    f'              <div class="nav-group-children" id="{grp_id}">\n'
                )
                html += render_nav_links(subsections)
                html += '              </div>\n            </div>\n'
            else:
                html += f'            <a href="#{sec_id}" class="nav-link" style="padding-left:{_nav_icon_offset}px" data-section-id="{sec_id}">{section["title"]}</a>\n'

        html += """        </div>
    </div>

    <div class="main-content" id="main-content-area" tabindex="-1" style="outline:none;">
"""

        html += generate_report_header_html(
            report_header, page_title, page_subtitle, page_description
        )

        html += """
        <div class="content-area">
"""

        def render_subsections(subsections):
            """Recursively render subsections at any depth."""
            sub_html = ""
            for subsection in subsections:
                subsection_id = subsection.get("id", "")
                depth = subsection.get("depth", 3)  # 3=###, 4=####, 5=#####, 6=######
                # h3 for depth-3 (###), h4 for depth-4, etc., capped at h6
                htag = f"h{min(depth, 6)}"

                # Collect code blocks for toggle buttons
                subsection_code_blocks = [
                    item
                    for item in subsection.get("content", [])
                    if item.get("type") == "code"
                ]
                sub_html += f'                <div class="subsection" id="{subsection_id}" data-section-id="{subsection_id}">\n'
                if include_code_blocks and subsection_code_blocks:
                    buttons_html = '<div class="code-buttons">'
                    for idx in range(len(subsection_code_blocks)):
                        block_id = f"{subsection_id}-code-{idx}"
                        buttons_html += f"""
                                <button class="code-btn code-toggle" onclick="toggleCode('{block_id}')" title="Toggle code">
                                    <svg viewBox="0 0 24 24">
                                        <path d="M19 13H5c-0.6 0-1-0.4-1-1s0.4-1 1-1h14c0.6 0 1 0.4 1 1S19.6 13 19 13z"/>
                                    </svg>
                                </button>
                            """
                    buttons_html += "</div>"
                    sub_html += f'                    <{htag} class="subsection-title"><span class="subsection-title-text">{subsection["title"]}</span>{buttons_html}</{htag}>\n'
                else:
                    sub_html += f'                    <{htag} class="subsection-title">{subsection["title"]}</{htag}>\n'

                if subsection.get("content"):
                    sub_html += process_content(subsection["content"], subsection_id)

                # Recurse into any deeper headings
                if subsection.get("subsections"):
                    sub_html += render_subsections(subsection["subsections"])

                sub_html += "                </div>\n"
            return sub_html

        for section in sections:
            html += f'            <div class="section" id="{section["id"]}" data-section-id="{section["id"]}">\n'

            # Generate section title with code buttons for each code block in this section
            section_code_blocks = [
                item
                for item in section.get("content", [])
                if item.get("type") == "code"
            ]
            if include_code_blocks and section_code_blocks:
                # Only the toggle button remains in the title — copy is now inside each block
                buttons_html = '<div class="code-buttons">'
                for idx in range(len(section_code_blocks)):
                    block_id = f"{section['id']}-code-{idx}"
                    buttons_html += f"""
                        <button class="code-btn code-toggle" onclick="toggleCode('{block_id}')" title="Toggle code">
                            <svg viewBox="0 0 24 24">
                                <path d="M19 13H5c-0.6 0-1-0.4-1-1s0.4-1 1-1h14c0.6 0 1 0.4 1 1S19.6 13 19 13z"/>
                            </svg>
                        </button>
                    """
                buttons_html += "</div>"
                html += f'                <h2 class="section-title"><span class="section-title-text">{section["title"]}</span>{buttons_html}</h2>\n'
            else:
                html += f'                <h2 class="section-title">{section["title"]}</h2>\n'

            if section.get("subsections"):
                html += render_subsections(section["subsections"])

            if section.get("content"):
                html += process_content(section["content"], section["id"])

            html += "            </div>\n\n"

        html += """        </div>

"""
        html += generate_report_footer_html(report_footer, footer_content, timestamp)
        html += """    </div>

    <script>
        function openSidebar() {
            var sb = document.querySelector('.sidebar.collapsible');
            var bd = document.getElementById('sidebar-backdrop');
            if (sb) sb.classList.add('open');
            if (bd) bd.classList.add('active');
        }

        function closeSidebar() {
            var sb = document.querySelector('.sidebar.collapsible');
            var bd = document.getElementById('sidebar-backdrop');
            if (sb) sb.classList.remove('open');
            if (bd) bd.classList.remove('active');
        }

        function toggleSidebar() {
            var sb = document.querySelector('.sidebar.collapsible');
            if (!sb) return;
            if (sb.classList.contains('open')) closeSidebar(); else openSidebar();
        }

        // Build nav dots from H2 sections
        (function() {
            var dotsContainer = document.getElementById('sidebar-nav-dots');
            if (!dotsContainer) return;
            var sections = document.querySelectorAll('.section[data-section-id]');
            sections.forEach(function(sec) {
                var id = sec.getAttribute('data-section-id');
                var title = sec.querySelector('.section-title');
                var label = title ? (title.querySelector('.section-title-text') || title).textContent.trim() : id;
                var dot = document.createElement('div');
                dot.className = 'sidebar-nav-dot';
                dot.setAttribute('data-dot-section', id);
                dot.title = label;
                dot.innerHTML = '<span class="sidebar-nav-dot-tooltip">' + label + '</span>';
                dot.addEventListener('click', function() {
                    sec.scrollIntoView({ behavior: 'smooth', block: 'start' });
                });
                dotsContainer.appendChild(dot);
            });
        })();
    
        function toggleTheme() {
            const body = document.body;
            const currentTheme = body.getAttribute('data-theme');
            const newTheme = currentTheme === 'light' ? 'dark' : 'light';
            body.setAttribute('data-theme', newTheme);
            localStorage.setItem('report-theme', newTheme);
            updateThemeIcon(newTheme);

            const isDark = newTheme === 'dark';
            (window._chartThemeFns || []).forEach(fn => { try { fn(isDark); } catch(e) {} });
        }

        function updateThemeIcon(theme) {
            const icon = document.getElementById('theme-icon');
            if (!icon) return;
            if (theme === 'dark') {
                icon.innerHTML = '<path d="M12 3a9 9 0 1 0 9 9c0-.46-.04-.92-.1-1.36a5.389 5.389 0 0 1-4.4 2.26 5.403 5.403 0 0 1-3.14-9.8c-.44-.06-.9-.1-1.36-.1z"/>';
            } else {
                icon.innerHTML = '<path d="M12 7a5 5 0 0 1 5 5 5 5 0 0 1-5 5 5 5 0 0 1-5-5 5 5 0 0 1 5-5m0-2a7 7 0 0 0-7 7 7 7 0 0 0 7 7 7 7 0 0 0 7-7 7 7 0 0 0-7-7M2 11h2v2H2v-2m18 0h2v2h-2v-2M11 2h2v2h-2V2m0 18h2v2h-2v-2M4.22 3.93l1.42 1.42-1.42 1.41-1.41-1.41 1.41-1.42m15.14 13.3 1.41 1.41-1.41 1.42-1.42-1.42 1.42-1.41M4.22 19.07l-1.41-1.42 1.41-1.41 1.42 1.41-1.42 1.42M19.36 5.36l-1.42-1.42 1.42-1.41 1.41 1.42-1.41 1.41z"/>';
            }
        }

        function toggleHelp() {
            document.getElementById('helpTooltip').classList.toggle('visible');
        }

        // Close help tooltip when clicking outside
        document.addEventListener('click', function(e) {
            const btn = document.getElementById('helpBtn');
            const tt  = document.getElementById('helpTooltip');
            if (btn && tt && !btn.contains(e.target) && !tt.contains(e.target)) {
                tt.classList.remove('visible');
            }
        });

        function toggleNavGroup(grpId) {
            const el  = document.getElementById(grpId);
            const ico = document.getElementById('ngi-' + grpId);
            if (!el) return;
            const collapsed = el.classList.toggle('collapsed');
            if (ico) {
                ico.innerHTML = collapsed
                    ? '<path d="M19 13h-6v6h-2v-6H5v-2h6V5h2v6h6v2z"/>'
                    : '<path d="M19 13H5v-2h14v2z"/>';
            }
        }

        // Row click: if click landed on the <a> nav-link, let it navigate normally.
        // Otherwise (clicked icon or row padding) toggle the group.
        function handleNavRowClick(event, grpId) {
            if (event.target.closest('a.nav-link')) {
                return;
            }
            event.preventDefault();
            toggleNavGroup(grpId);
        }

        function collapseAll() {
            document.querySelectorAll('.nav-group-children').forEach(el => {
                el.classList.add('collapsed');
                const ico = document.getElementById('ngi-' + el.id);
                if (ico) ico.innerHTML = '<path d="M19 13h-6v6h-2v-6H5v-2h6V5h2v6h6v2z"/>';
            });
        }

        function expandAll() {
            document.querySelectorAll('.nav-group-children').forEach(el => {
                el.classList.remove('collapsed');
                const ico = document.getElementById('ngi-' + el.id);
                if (ico) ico.innerHTML = '<path d="M19 13H5v-2h14v2z"/>';
            });
        }

        // ── Keyboard shortcuts ────────────────────────────────────────────
        document.addEventListener('keydown', function(event) {
            const tag     = document.activeElement.tagName;
            const inInput = tag === 'INPUT' || tag === 'TEXTAREA';

            // Esc — clear search or close any open modal overlay
            if (event.key === 'Escape') {
                const overlay = document.getElementById('cellModalOverlay');
                if (overlay && overlay.classList.contains('active')) {
                    overlay.classList.remove('active');
                } else if (inInput) {
                    clearSearch();
                    document.activeElement.blur();
                    document.getElementById('main-content-area').focus();
                }
                return;
            }

            // F — focus search box only; Esc exits
            if (event.key === 'f' || event.key === 'F') {
                if (!inInput) {
                    event.preventDefault();
                    const inp = document.getElementById('search-input');
                    inp.focus(); inp.select();
                }
                // When already in the search box, let F type normally
                return;
            }

            // T — toggle theme (only when not typing)
            if ((event.key === 't' || event.key === 'T') && !inInput) {
                event.preventDefault();
                toggleTheme();
                return;
            }

            // S — toggle collapsible sidebar
            if ((event.key === 's' || event.key === 'S') && !inInput) {
                event.preventDefault();
                toggleSidebar();
                return;
            }

            // N — focus navigation sidebar
            if ((event.key === 'n' || event.key === 'N') && !inInput) {
                event.preventDefault();
                const links = Array.from(
                    document.querySelectorAll('.nav-link:not(.hidden)')
                );
                if (links.length) links[0].focus();
                return;
            }

            // M — focus main content (for keyboard scrolling)
            if ((event.key === 'm' || event.key === 'M') && !inInput) {
                event.preventDefault();
                document.getElementById('main-content-area').focus();
                return;
            }

            // ↑ / ↓ — navigate visible nav links ONLY when a nav-link is already focused
            const focusedIsNavLink = document.activeElement.classList.contains('nav-link');
            if (focusedIsNavLink && (event.key === 'ArrowDown' || event.key === 'ArrowUp')) {
                event.preventDefault();
                const links = Array.from(
                    document.querySelectorAll('.nav-link:not(.hidden)')
                );
                if (!links.length) return;
                const idx = links.indexOf(document.activeElement);
                if (event.key === 'ArrowDown') {
                    links[idx + 1 < links.length ? idx + 1 : 0].focus();
                } else {
                    links[idx - 1 >= 0 ? idx - 1 : links.length - 1].focus();
                }
            }
        });

        const savedTheme = localStorage.getItem('report-theme');
        if (savedTheme) {
            document.body.setAttribute('data-theme', savedTheme);
            updateThemeIcon(savedTheme);
            window.addEventListener('load', function() {
                const isDark = savedTheme === 'dark';
                (window._chartThemeFns || []).forEach(fn => { try { fn(isDark); } catch(e) {} });
            });
        } else {
            updateThemeIcon(document.body.getAttribute('data-theme') || 'light');
        }

        function toggleCode(blockId) {
            const codeContent = document.getElementById(blockId);
            const toggleBtn = event.target.closest('.code-toggle');

            if (codeContent.classList.contains('collapsed')) {
                codeContent.classList.remove('collapsed');
                // Change SVG to minus/hide icon
                toggleBtn.innerHTML = `<svg viewBox="0 0 24 24">
                    <path d="M19 13H5c-0.6 0-1-0.4-1-1s0.4-1 1-1h14c0.6 0 1 0.4 1 1S19.6 13 19 13z"/>
                </svg>`;
            } else {
                codeContent.classList.add('collapsed');
                // Change SVG to plus/show icon
                toggleBtn.innerHTML = `<svg viewBox="0 0 24 24">
                    <path d="M19 13h-6v6c0 0.6-0.4 1-1 1s-1-0.4-1-1v-6H5c-0.6 0-1-0.4-1-1s0.4-1 1-1h6V5c0-0.6 0.4-1 1-1s1 0.4 1 1v6h6c0.6 0 1 0.4 1 1S19.6 13 19 13z"/>
                </svg>`;
            }
        }

        function copyCode(blockId) {
            const codeElement = document.querySelector(`#${blockId} code`);
            if (!codeElement) return;
            const text = codeElement.textContent;
            navigator.clipboard.writeText(text).then(() => {
                // Find the copy button inside this specific block
                const btn = document.querySelector(`#${blockId} .copy-code-btn`);
                if (!btn) return;
                const orig = btn.textContent;
                btn.textContent = 'copied!';
                btn.style.opacity = '1';
                setTimeout(() => {
                    btn.textContent = orig;
                    btn.style.opacity = '';
                }, 2000);
            });
        }

        function toggleClearButton() {
            const searchInput = document.getElementById('search-input');
            const clearButton = document.getElementById('search-clear');

            if (searchInput.value.trim() !== '') {
                clearButton.classList.add('visible');
            } else {
                clearButton.classList.remove('visible');
            }
        }

        function clearSearch() {
            const searchInput = document.getElementById('search-input');
            searchInput.value = '';
            toggleClearButton();
            // Strip any stale scroll-spy active highlights left from before the search
            document.querySelectorAll('.nav-link.active').forEach(link => {
                link.classList.remove('active');
            });
            filterSections();
        }

        function filterSections() {
            const searchInput = document.getElementById('search-input');
            const searchTerm = searchInput.value.toLowerCase().trim();
            const navLinks = document.querySelectorAll('.nav-link');
            const sections = document.querySelectorAll('.section');
            const searchCount = document.getElementById('search-count');

            let visibleCount = 0;
            const totalCount = navLinks.length;

            // When there's no search term, show everything
            if (searchTerm === '') {
                navLinks.forEach(link => link.classList.remove('hidden'));
                sections.forEach(section => section.classList.remove('hidden'));
                searchCount.textContent = '';
                return;
            }

            // Reset visibility for all sections (they remain visible by default)
            sections.forEach(section => section.classList.remove('hidden'));

            // Filter navigation links only
            navLinks.forEach(link => {
                const linkText = link.textContent.toLowerCase();
                if (linkText.includes(searchTerm)) {
                    link.classList.remove('hidden');
                    visibleCount++;
                } else {
                    link.classList.add('hidden');
                }
            });

            searchCount.textContent = `Showing ${visibleCount} of ${totalCount} navigation items`;
        }

        // Initialize navigation links with event listeners
        function initNavigation() {
            document.querySelectorAll('.nav-link').forEach(link => {
                link.addEventListener('click', function(e) {
                    e.preventDefault();
                    const targetId = this.getAttribute('href');
                    const targetElement = document.querySelector(targetId);
                    if (targetElement) {
                        targetElement.scrollIntoView({
                            behavior: 'smooth',
                            block: 'start'
                        });

                        // Close collapsible sidebar after navigation
                        closeSidebar();

                        // Highlight the section briefly
                        targetElement.style.backgroundColor = 'var(--hover-color)';
                        setTimeout(() => {
                            targetElement.style.backgroundColor = '';
                        }, 1000);
                    }
                });
            });
        }

        // Initialize code buttons with correct icons based on collapsed state
        function initCodeButtons() {
            document.querySelectorAll('.code-toggle').forEach(btn => {
                const blockId = btn.getAttribute('onclick').match(/toggleCode\('([^']+)'\)/)[1];
                const codeContent = document.getElementById(blockId);
                if (codeContent && codeContent.classList.contains('collapsed')) {
                    // Set to plus icon for collapsed state
                    btn.innerHTML = `<svg viewBox="0 0 24 24">
                        <path d="M19 13h-6v6c0 0.6-0.4 1-1 1s-1-0.4-1-1v-6H5c-0.6 0-1-0.4-1-1s0.4-1 1-1h6V5c0-0.6 0.4-1 1-1s1 0.4 1 1v6h6c0.6 0 1 0.4 1 1S19.6 13 19 13z"/>
                    </svg>`;
                }
            });
        }

        function initScrollSpy() {
            const sections = document.querySelectorAll('[data-section-id]');
            const navLinks = document.querySelectorAll('.nav-link');
            const navDots = document.querySelectorAll('.sidebar-nav-dot');

            const observer = new IntersectionObserver(
                (entries) => {
                    entries.forEach(entry => {
                        const id = entry.target.getAttribute('data-section-id');

                        if (entry.isIntersecting) {
                            // Don't apply active highlight while user is searching
                            const searchTerm = document.getElementById('search-input').value.trim();
                            if (searchTerm) return;

                            // Remove active from all
                            navLinks.forEach(link => link.classList.remove('active'));

                            // Add active to matching nav
                            const activeLink = document.querySelector(
                                `.nav-link[data-section-id="${id}"]`
                            );
                            if (activeLink) {
                                activeLink.classList.add('active');

                                // Auto-scroll sidebar to keep active link visible
                                activeLink.scrollIntoView({
                                    block: 'nearest',
                                    behavior: 'smooth'
                                });
                            }

                            // Update nav dots - find the parent H2 section id
                            var dotId = id;
                            var el = entry.target;
                            var parentSec = el.closest('.section[data-section-id]');
                            if (parentSec) dotId = parentSec.getAttribute('data-section-id');
                            navDots.forEach(function(d) { d.classList.remove('active'); });
                            var activeDot = document.querySelector('.sidebar-nav-dot[data-dot-section="' + dotId + '"]');
                            if (activeDot) activeDot.classList.add('active');
                        }
                    });
                },
                {
                    root: null,
                    rootMargin: '-30% 0px -60% 0px',
                    threshold: 0
                }
            );

            sections.forEach(section => observer.observe(section));
        }

        // Run after page load
        window.addEventListener('DOMContentLoaded', initScrollSpy);

        document.addEventListener('DOMContentLoaded', function() {
            initNavigation();
            initCodeButtons();
            toggleClearButton();

            // Default keyboard scroll focus on main content area
            const mainArea = document.getElementById('main-content-area');
            if (mainArea) mainArea.focus();
        });

        // ── Datatable: row search ─────────────────────────────────────────
        function dtSearchRows(queryId, maxRows) {
            var box    = document.querySelector('#dt-wrap-' + queryId + ' input[placeholder^="\\u25bc"]');
            var term   = box ? box.value.toLowerCase().trim() : '';
            var rcount = document.getElementById('dt-rcount-' + queryId);
            var tbody  = document.getElementById('dt-body-' + queryId);
            var full   = window['dtFullData_' + queryId];
            var cols   = window['dtFullCols_' + queryId];
            var total  = window['dtTotalRows_' + queryId] || 0;

            // Determine visible columns (respects column filter)
            function getVisibleCols() {
                var table = document.getElementById('dt-' + queryId);
                if (!table) return cols;
                var visHeaders = Array.from(table.querySelectorAll('thead th:not(.dt-col-hidden)'));
                if (visHeaders.length === 0 || visHeaders.length === cols.length) return cols;
                return visHeaders.map(function(th){return th.getAttribute('data-col) || '';});
            }

            if (full && cols) {
                if (!term) {
                    tbody.innerHTML = '';
                    var show = Math.min(maxRows, full.length);
                    for (var i = 0; i < full.length; i++) {
                        tbody.appendChild(dtBuildRow(full[i], cols, i, i >= show));
                    }
                    rcount.textContent = 'Showing ' + show.toLocaleString() + ' of ' + total.toLocaleString() + ' rows';
                    return;
                }
                var filtered;
                if (term.includes(':')) {
                    // Column-specific search e.g. "radio:245"
                    var parts = term.split(':');
                    var colSearch = parts[0].trim();
                    var valSearch = parts.slice(1).join(':').trim();
                    var matchedCol = cols.find(function(c){return c.toLowerCase() === colSearch;});
                    if (!matchedCol) matchedCol = cols.find(function(c){return c.toLowerCase().includes(colSearch);});
                    if (!matchedCol) {
                        tbody.innerHTML = '<tr><td colspan="' + cols.length + '" style="text-align:center;padding:16px;opacity:0.5;">Column \u201c' + colSearch + '\u201d not found</td></tr>';
                        rcount.textContent = 'No results';
                        return;
                    }
                    filtered = full.filter(function(row) {
                        return String(row[matchedCol] !== null && row[matchedCol] !== undefined ? row[matchedCol] : '').toLowerCase().includes(valSearch);
                    });
                } else {
                    // General search - multi-word (all words must match across visible columns)
                    var searchCols = getVisibleCols();
                    var words = term.split(/\s+/).filter(function(w){return w;});
                    filtered = full.filter(function(row) {
                        return words.every(function(word) {
                            return searchCols.some(function(c) {
                                return String(row[c] !== null && row[c] !== undefined ? row[c] : '').toLowerCase().includes(word);
                            });
                        });
                    });
                }

                tbody.innerHTML = '';
                if (!filtered.length) {
                    tbody.innerHTML = '<tr><td colspan="' + cols.length + '" style="text-align:center;padding:16px;opacity:0.5;">No results for \u201c' + term + '\u201d</td></tr>';
                    rcount.textContent = 'No results';
                    return;
                }
                var lim = Math.min(filtered.length, 1000);
                for (var i = 0; i < lim; i++) tbody.appendChild(dtBuildRow(filtered[i], cols, i, false));
                var note = filtered.length > 1000 ? ' (showing first 1,000)' : '';
                rcount.textContent = 'Found ' + filtered.length.toLocaleString() + ' results' + note;
            } else {
                // Fallback: search only in rendered rows
                var rows = tbody ? Array.from(tbody.querySelectorAll('tr')) : [];
                var vis = 0;
                rows.forEach(function(r) {
                    var match = !term || Array.from(r.querySelectorAll('td')).some(function(td) {
                        return td.textContent.toLowerCase().includes(term);
                    });
                    r.classList.toggle('dt-hidden', !match);
                    if (match) vis++;
                });
                if (rcount) rcount.textContent = 'Showing ' + vis.toLocaleString() + ' rows';
            }
        }

        function dtBuildRow(rowData, cols, idx, hidden) {
            var tr = document.createElement('tr');
            tr.setAttribute('data-row-id', idx);
            if (hidden) tr.classList.add('dt-hidden');
            cols.forEach(function(c) {
                var td = document.createElement('td');
                td.setAttribute('data-col', c.toLowerCase());
                var raw = rowData[c];
                var display;
                if (raw === null || raw === undefined) {
                    display = '';
                } else if (typeof raw === 'number' && !Number.isInteger(raw)) {
                    // Trim floats to 4 decimal places, remove trailing zeros
                    display = parseFloat(raw.toFixed(4)).toString();
                } else {
                    display = String(raw);
                }
                td.textContent = display;
                tr.appendChild(td);
            });
            return tr;
        }

        // ── Datatable: column filter ──────────────────────────────────────
        function dtFilterCols(queryId) {
            var box     = document.querySelector('#dt-wrap-' + queryId + ' input[placeholder^="\\uD83D\\uDD0D"]');
            var rawTerm    = box ? box.value.trim().toLowerCase() : '';
            var ccount  = document.getElementById('dt-ccount-' + queryId);
            var table   = document.getElementById('dt-' + queryId);
            if (!table) return;
            var headers = Array.from(table.querySelectorAll('thead th'));
            var total   = headers.length;
            var vis     = 0;
            // Support comma-separated multi-term filter
            var terms = rawTerm ? rawTerm.split(',').map(function(t){return t.trim();}).filter(function(t){return t;}) : [];
            headers.forEach(function(th, i) {
                var col   = (th.getAttribute('data-col') || '').toLowerCase();
                var match = terms.length === 0 || terms.some(function(t){return col.includes(t);});
                th.classList.toggle('dt-col-hidden', !match);
                if (match) vis++;
                Array.from(table.querySelectorAll('tbody tr')).forEach(function(row) {
                    var cells = row.querySelectorAll('td');
                    if (cells[i]) cells[i].classList.toggle('dt-col-hidden', !match);
                });
            });
            if (ccount) ccount.textContent = 'Showing ' + vis + ' of ' + total + ' cols';
        }

        // ── Datatable: cell expand (reuses the existing report modal) ─────
        function dtExpand(td) {
            var full = td.getAttribute('data-full') || td.textContent;
            var col  = td.getAttribute('data-col') || '';
            var overlay = document.getElementById('cellModalOverlay');
            if (overlay) {
                document.getElementById('cellModalColumn').textContent  = col;
                document.getElementById('cellModalContent').textContent = full;
                overlay.classList.add('active');
            } else {
                // Fallback if report modal is absent: browser alert
                alert(col + ': ' + full);
            }
        }
    </script>
</body>
</html>"""

        # ── Post-process: substitute placeholders that cannot live inside the
        # f-string (they contain literal { } which Python would misinterpret)
        html = (
            html
            .replace("__NAV_FONT_CSS_PLACEHOLDER__", _nav_font_css)
            .replace("__HEADING_CSS_PLACEHOLDER__",  _heading_css)
        )

        return html

    try:
        md_path = Path(markdown_file)
        if not md_path.exists():
            print(f"Error: Markdown file '{markdown_file}' not found!")
            return None

        print(f"📖 Reading markdown file: {markdown_file}")
        with open(md_path, "r", encoding="utf-8") as f:
            md_content = f.read()

        print("🔍 Parsing markdown content...")
        sections = parse_markdown(md_content, md_path.parent)

        if not sections:
            print("No sections found in markdown file!")
            return None

        print(f"✅ Found {len(sections)} sections")

        def _count_content(nodes):
            code = images = refs = html_blocks = 0
            for node in nodes:
                for item in node.get("content", []):
                    t = item.get("type")
                    if t == "code":
                        code += 1
                    elif t == "image":
                        images += 1
                    elif t == "reference":
                        refs += 1
                    elif t == "html":
                        html_blocks += 1
                c2, i2, r2, h2 = _count_content(node.get("subsections", []))
                code += c2
                images += i2
                refs += r2
                html_blocks += h2
            return code, images, refs, html_blocks

        total_code = total_images = total_refs = total_html = 0
        for section in sections:
            for item in section.get("content", []):
                t = item.get("type")
                if t == "code":
                    total_code += 1
                elif t == "image":
                    total_images += 1
                elif t == "reference":
                    total_refs += 1
                elif t == "html":
                    total_html += 1
            c2, i2, r2, h2 = _count_content(section.get("subsections", []))
            total_code += c2
            total_images += i2
            total_refs += r2
            total_html += h2

        print("🌐 Generating HTML report...")
        html_content = generate_html(sections)

        output_path = md_path.parent / output_file
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_content)

        print("✅ HTML report created successfully!")
        print(f"📁 Output: {output_path}")
        print(f"📊 Sections: {len(sections)}")
        if include_code_blocks:
            print(
                f"💻 Code blocks: {total_code} ({'collapsed' if code_blocks_collapsed else 'expanded'})"
            )
            print(f"🎨 Code buttons: Integrated into section titles (hover to reveal)")
        print(f"🖼️  Images embedded: {total_images}")
        print(
            f"📝 Reference blocks: {total_refs} ({'styled' if style_output_blocks else 'plain'})"
        )
        print(f"📋 HTML tables: {total_html}")

        return str(output_path)

    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback

        traceback.print_exc()
        return None


# if __name__ == "__main__":
#     convert_markdown_to_html(
#         # markdown_file=r"C:\my_disk\edupunk\all_docs\docs\analytics_express\machine_learning\model_selection\roc_curve.md",
#         # markdown_file=r"C:\my_disk\edupunk\all_docs\docs\analytics_express\machine_learning\model_selection\demystify_ml_model_selection.md",
#         markdown_file=r"C:\my_disk\edupunk\all_docs\docs\analytics_express\machine_learning\regression\house_prices.md",
#         # markdown_file=r"C:\my_disk\edupunk\all_docs\docs\analytics_express\machine_learning\regression\marketing_effectiveness.md",
#         output_file=r"C:\my_disk\projects\visual_library\report\py_to_ipynb_to_html.html",
#         page_title="Python Diagrams Documentation",
#         page_subtitle="Comprehensive Guide to Creating Visual Diagrams",
#         page_description="This report demonstrates...",
#         footer_content="Created with Python • Link",
#         default_theme="dark",
#         navigation_title="Report Juice!",
#         include_code_blocks=True,
#         code_blocks_collapsed=True,
#         style_output_blocks=False,  # Set to False for plain text output without border/background
#     )