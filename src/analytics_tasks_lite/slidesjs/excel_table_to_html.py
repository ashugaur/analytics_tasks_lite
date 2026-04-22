# %% excel_table_to_html

## Dependencies
"""
Excel Table to SlideJS Converter
Converts Excel tables to HTML with advanced formatting, multi-level headers, and flexible heatmaps
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import numpy as np


# ============================================================================
# THEME DEFAULTS
# ============================================================================

TABLE_THEME_DEFAULTS = {
    "light": {
        "header_bg":     "#2A918B",
        "header_text":   "white",
        "subheader_bg":  "#2A918B",
        "category_bg":   "#FFFFFF",
        "category_text": "#2A918B",
        "total_bg":      "#2A918B",
        "total_text":    "white",
        "data_text":     "#2A918B",
        "stripe_a":      "#f9f9f9",
        "stripe_b":      "white",
        "border":        "#dddddd",
    },
    "dark": {
        "header_bg":     "#1a4a47",
        "header_text":   "#e0e0e0",
        "subheader_bg":  "#1a4a47",
        "category_bg":   "#1e1e1e",
        "category_text": "#7ecfca",
        "total_bg":      "#1a4a47",
        "total_text":    "#e0e0e0",
        "data_text":     "#a0cece",
        "stripe_a":      "#2a2a2a",
        "stripe_b":      "#1e1e1e",
        "border":        "#444444",
    },
}


def _table_read_color_mapping(color_file_path, sheet_name=None):
    if color_file_path is None:
        return None
    if isinstance(sheet_name, dict):
        sheet_name = sheet_name.get("name")
    try:
        return pd.read_excel(color_file_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"  ⚠️  Could not read color file: {e}")
        return None


def _table_resolve_color(param_value, color_df, topic, chart_element,
                         element_name, theme, fallback_key):
    """Priority: explicit param > color file > TABLE_THEME_DEFAULTS."""
    if param_value is not None:
        return param_value
    if color_df is not None and topic is not None:
        try:
            hex_col = "light_hex" if theme == "light" else "dark_hex"
            match = color_df[
                (color_df["topic"]           == topic)
                & (color_df["chart_type"]    == "table")
                & (color_df["chart_element"] == chart_element)
                & (color_df["element_name"]  == element_name)
            ]
            if not match.empty:
                return str(match.iloc[0][hex_col])
        except Exception:
            pass
    defaults = TABLE_THEME_DEFAULTS.get(theme, TABLE_THEME_DEFAULTS["light"])
    return defaults.get(fallback_key, "#333333")


def _build_table_html(df, start_row, start_col, end_col, header_rows,
                      category_row_set, last_row_is_total, merged_cells,
                      multi_level_headers, merged_cell_parents,
                      col_config_by_idx, heatmap_data,
                      header_bg, header_color, subheader_bg,
                      category_bg, category_color, total_row_bg, total_row_color,
                      data_text_color, row_colors, border_color,
                      cell_padding, font_family, font_size,
                      text_align, number_align, bold_first_col,
                      indent_child_rows, child_indent):
    """Build a single <table>…</table> HTML string for one theme."""
    html = (f'<table style="width:100%; border-collapse:collapse; '
            f'font-family:{font_family}; font-size:{font_size};">\n')
    skip_cells = set()

    for row_idx in range(len(df)):
        excel_row = start_row + row_idx
        is_header    = row_idx < header_rows
        is_category  = row_idx in category_row_set
        is_total_row = (row_idx == len(df) - 1) and last_row_is_total
        is_child_row = not is_header and not is_category and not is_total_row

        if is_header:
            if row_idx == 0:
                html += "  <thead>\n"
            bg = header_bg if row_idx == 0 else subheader_bg
            html += f'    <tr style="background:{bg}; color:{header_color}; font-weight:bold;">\n'
        elif is_category:
            html += f'    <tr style="background:{category_bg}; color:{category_color}; font-weight:bold;">\n'
        elif is_total_row:
            html += f'    <tr style="background:{total_row_bg}; color:{total_row_color}; font-weight:bold;">\n'
        else:
            bg_color = row_colors[row_idx % 2]
            html += f'    <tr style="background:{bg_color}; color:{data_text_color};">\n'

        for col_idx in range(len(df.columns)):
            excel_col = start_col + col_idx
            if (excel_row, excel_col) in skip_cells:
                continue

            cell_value = df.iloc[row_idx, col_idx]
            if pd.isna(cell_value):
                cell_value = ""

            merge_info = merged_cells.get((excel_row, excel_col), {"colspan": 1, "rowspan": 1})
            colspan = merge_info["colspan"]
            rowspan = merge_info["rowspan"]

            for r in range(excel_row, excel_row + rowspan):
                for c in range(excel_col, excel_col + colspan):
                    if (r, c) != (excel_row, excel_col):
                        skip_cells.add((r, c))

            is_merged_header_continuation = (
                multi_level_headers and is_header
                and (row_idx, col_idx) in merged_cell_parents
            )

            if is_header:
                if colspan > 1:
                    cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:none;"
                    if col_idx == 0:
                        cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:1px solid {border_color}; border-right:none;"
                    if col_idx + colspan == len(df.columns):
                        cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:1px solid {border_color};"
                elif is_merged_header_continuation:
                    cell_style = f"padding:{cell_padding}; border:none;"
                else:
                    cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:none;"
                    if col_idx == 0:
                        cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:1px solid {border_color}; border-right:none;"
                    if col_idx == len(df.columns) - 1:
                        cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:1px solid {border_color};"
            else:
                cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:none;"

            if multi_level_headers and is_header and row_idx > 0 and cell_value == "":
                parent_bg = header_bg
                for (mer_row, mer_col), mer_info in merged_cells.items():
                    if (mer_row < excel_row
                            and mer_col <= excel_col < mer_col + mer_info["colspan"]
                            and mer_row + mer_info["rowspan"] > excel_row):
                        parent_bg = header_bg
                        break
                cell_style += f" background:{parent_bg};"

            if is_child_row and indent_child_rows and col_idx == 0:
                cell_style += f" padding-left:{child_indent};"

            cell_style += " text-align:center;" if colspan > 1 else f" text-align:{text_align};"
            if rowspan > 1:
                cell_style += " vertical-align:middle;"
            if bold_first_col and col_idx == 0 and not is_header and not is_category:
                cell_style += " font-weight:bold;"

            col_config = col_config_by_idx.get(col_idx, {})

            if (col_config.get("heatmap") and col_idx in heatmap_data
                    and not is_header and not is_category and not is_total_row):
                if isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                    for range_config in col_config.get("ranges", []):
                        min_pct, max_pct = range_config.get("percentile", (0, 100))
                        min_val = np.percentile(heatmap_data[col_idx]["values"], min_pct)
                        max_val = np.percentile(heatmap_data[col_idx]["values"], max_pct)
                        if min_val <= cell_value <= max_val:
                            bg_col = range_config.get("color", "#cccccc")
                            cell_style += f" background:{bg_col}; color:{data_text_color}; font-weight:bold;"
                            break

            number_format  = col_config.get("number_format", "auto")
            decimal_places = col_config.get("decimal_places", 2)

            if isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                if number_format == "percentage":
                    cell_value = f"{cell_value * 100:.{decimal_places}f}%" if cell_value < 1 else f"{cell_value:.{decimal_places}f}%"
                elif number_format == "integer":
                    cell_value = f"{int(cell_value):,}"
                elif number_format == "decimal":
                    cell_value = f"{cell_value:,.{decimal_places}f}"
                else:
                    cell_value = f"{int(cell_value):,}" if cell_value == int(cell_value) else f"{cell_value:,.{decimal_places}f}"
                if colspan == 1:
                    cell_style += f" text-align:{number_align};"

            cell_attrs = ""
            if colspan > 1: cell_attrs += f'colspan="{colspan}" '
            if rowspan > 1: cell_attrs += f'rowspan="{rowspan}" '
            cell_attrs += f'style="{cell_style}"'

            tag = "th" if is_header else "td"
            html += f"      <{tag} {cell_attrs}>{cell_value}</{tag}>\n"

        html += "    </tr>\n"
        if is_header and row_idx == header_rows - 1:
            html += "  </thead>\n  <tbody>\n"

    html += "  </tbody>\n</table>"
    return html


def excel_table_to_html(
    excel_file,
    sheet_name,
    output_file=None,
    start_row=1,
    end_row=None,
    start_col=1,
    end_col=None,
    # Row formatting
    header_rows=1,
    category_rows=None,  # List of row indices (1-indexed) that are category headers
    indent_child_rows=True,
    child_indent="20px",
    # Multi-level headers
    multi_level_headers=False,
    # Column configuration
    column_config=None,  # Dict with column name -> formatting config
    # Compact styling
    compact_mode="compact",  # 'standard', 'compact', 'ultra-compact'
    font_size=None,  # Override compact mode
    cell_padding=None,  # Override compact mode
    # Color scheme — None = resolve from color file, then TABLE_THEME_DEFAULTS
    header_bg=None,
    header_color=None,
    subheader_bg=None,
    category_bg=None,
    category_color=None,
    row_colors=None,
    total_row_bg=None,
    total_row_color=None,
    data_text_color=None,
    # Table styling
    show_borders=True,
    border_color=None,
    font_family="Segoe UI, Helvetica Neue, Arial, sans-serif",
    # Legacy parameters (for backward compatibility)
    apply_heatmap=False,
    heatmap_cols=None,
    heatmap_color="blue",
    text_align="left",
    number_align="right",
    bold_first_col=False,
    # ── Theme & color mapping ─────────────────────────────────────────────────
    theme="light",
    color_file_path=None,
    color_sheet_name=None,
    color_topic=None,
):
    """
    Convert Excel table to HTML with advanced formatting capabilities.

    Parameters:
    ---
    excel_file : str or Path
        Path to Excel file
    sheet_name : str
        Sheet name to convert
    output_file : str or Path, optional
        Output file path (e.g., 'tables/summary.htmltable')
    start_row/end_row/start_col/end_col : int
        Table range (1-indexed)

    Row Formatting:
    ---
    header_rows : int
        Number of header rows (default=1)
    category_rows : list of int
        Row indices (1-indexed, relative to start_row) that are category headers
        Example: [1, 5] for Drug A and Drug B rows
    indent_child_rows : bool
        Add left padding to non-category data rows (default=True)
    child_indent : str
        Indentation for child rows (default='20px')

    Multi-Level Headers:
    ---
    multi_level_headers : bool
        Enable special handling for multi-level headers with blank continuation cells

    Column Configuration:
    ---
    column_config : dict
        Column-specific formatting. Example:

        'Drug A LOYALIST': {
            'heatmap': True,
            'ranges': [
                {'percentile': (67, 100), 'color': '#28a745'}, # Top 33%
                {'percentile': (33, 67), 'color': '#ffc107'}, # Middle 33%
                {'percentile': (0, 33), 'color': '#dc3545' } # Bottom 33%
            ]
            'number_format': 'percentage', # 'percentage', 'decimal', 'integer'
            'decimal_places': 1
        }

    Compact Styling:
    ---
    compact_mode : str
        'standard': padding=10px, font=14px
        'compact': padding=6px, font=12px (default)
        'ultra-compact': padding=4px, font=11px
    font_size : str
        Override font size (e.g., '11px')
    cell_padding : str
        Override cell padding (e.g., '5px')

    Colors:
    ---
    data_text_color : str
        Text color for data values (default='#2A918B')

    Returns:
    ---
    str : Path to saved file or HTML string

    Examples:
    ---
    # Simple table with category rows
    excel_table_to_html(
        excel_file='data.xlsx',
        sheet_name='Pivot',
        output_file='tables/pivot.htmltable',
        category_rows=[1, 5, 9], # Drug A, Drug B, NON-LOYALIST
        column_config={
            'Drug A LOYALIST': {
                'heatmap': True,
                'ranges': [
                    {'percentile': (67, 100), 'color': '#28a745'},
                    {'percentile': (33, 67), 'color': '#ffc107'},
                    {'percentile': (0, 33), 'color': '#dc3545'}]
            },
            'number_format': 'percentage'
        }
    }

    # Multi-level header table
    excel_table_to_html(
        excel_file='data.xlsx',
        sheet_name='AgeBreakdown',
        output_file='tables/age_breakdown.htmltable',
        header_rows=2,
        multi_level_headers=True,
        category_rows=[1, 6, 11]
    )
    """

    print("\n🔄 Converting Excel table to HTML...")
    print(f"   File: {excel_file}")
    print(f"   Sheet: {sheet_name}")

    # Set compact mode defaults
    compact_presets = {
        "standard":      {"padding": "10px", "font": "14px"},
        "compact":       {"padding": "6px",  "font": "12px"},
        "ultra-compact": {"padding": "4px",  "font": "11px"},
    }
    preset = compact_presets.get(compact_mode, compact_presets["compact"])
    if font_size is None:
        font_size = preset["font"]
    if cell_padding is None:
        cell_padding = preset["padding"]

    # ── Resolve colours for both themes ──────────────────────────────────────
    _color_df = _table_read_color_mapping(color_file_path, sheet_name=color_sheet_name)

    def _rc(param_val, chart_element, element_name, fallback_key, t):
        return _table_resolve_color(param_val, _color_df, color_topic,
                                    chart_element, element_name, t, fallback_key)

    def _resolve_theme_colors(t):
        is_active = (t == theme)
        pv = lambda v: v if is_active else None
        sa = _rc(None, "data_row", "stripe_a", "stripe_a", t)
        sb = _rc(None, "data_row", "stripe_b", "stripe_b", t)
        rc = (row_colors if (row_colors is not None and is_active) else (sa, sb))
        return dict(
            header_bg      = _rc(pv(header_bg),       "header",       "background", "header_bg",    t),
            header_color   = _rc(pv(header_color),    "header",       "text",       "header_text",  t),
            subheader_bg   = _rc(pv(subheader_bg),    "subheader",    "background", "subheader_bg", t),
            category_bg    = _rc(pv(category_bg),     "category_row", "background", "category_bg",  t),
            category_color = _rc(pv(category_color),  "category_row", "text",       "category_text",t),
            total_row_bg   = _rc(pv(total_row_bg),    "total_row",    "background", "total_bg",     t),
            total_row_color= _rc(pv(total_row_color), "total_row",    "text",       "total_text",   t),
            data_text_color= _rc(pv(data_text_color), "data_row",     "text",       "data_text",    t),
            border_color   = _rc(pv(border_color),    "border",       "main",       "border",       t),
            row_colors     = rc,
        )
    # ─────────────────────────────────────────────────────────────────────────

    # Load workbook
    wb = load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]

    merged_cells = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        merged_cells[(min_row, min_col)] = {
            "colspan": max_col - min_col + 1,
            "rowspan": max_row - min_row + 1,
        }
    if merged_cells:
        print(f"   ✓ Found {len(merged_cells)} merged cell range(s)")

    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

    if end_row is None:
        end_row = len(df)
        while end_row > start_row and df.iloc[end_row - 1].isna().all():
            end_row -= 1
    if end_col is None:
        end_col = len(df.columns)

    df = df.iloc[start_row - 1:end_row, start_col - 1:end_col]
    print(f"   ✓ Table size: {len(df)} rows x {len(df.columns)} columns")

    col_headers = []
    for col_idx in range(len(df.columns)):
        header_val = df.iloc[0, col_idx]
        if pd.isna(header_val):
            excel_col = start_col + col_idx
            parent_header = None
            for (mer_row, mer_col), mer_info in merged_cells.items():
                if mer_row == start_row and mer_col <= excel_col < mer_col + mer_info["colspan"]:
                    parent_header = ws.cell(mer_row, mer_col).value
                    break
            col_headers.append(parent_header if parent_header else f"Column_{col_idx}")
        else:
            col_headers.append(str(header_val))

    col_config_by_idx = {}
    if column_config:
        for col_identifier, config in column_config.items():
            if isinstance(col_identifier, int):
                col_config_by_idx[col_identifier] = config
            else:
                try:
                    col_config_by_idx[col_headers.index(col_identifier)] = config
                except ValueError:
                    print(f"   ⚠️ Warning: Column '{col_identifier}' not found in headers")

    category_row_set = set()
    if category_rows:
        category_row_set = {r - 1 for r in category_rows}
        print(f"   ✓ Category rows: {category_rows}")

    last_row_is_total = False
    if len(df) > 0:
        first_cell = str(df.iloc[-1, 0]).lower()
        if "total" in first_cell or "sum" in first_cell or "grand" in first_cell:
            last_row_is_total = True

    heatmap_data = {}
    for col_idx, config in col_config_by_idx.items():
        if config.get("heatmap"):
            values = [
                df.iloc[r, col_idx]
                for r in range(header_rows, len(df))
                if r not in category_row_set
                and not (r == len(df) - 1 and last_row_is_total)
                and isinstance(df.iloc[r, col_idx], (int, float))
                and not pd.isna(df.iloc[r, col_idx])
            ]
            if values:
                heatmap_data[col_idx] = {
                    "values": values,
                    "percentiles": np.percentile(values, [0, 33, 67, 100]),
                }

    merged_cell_parents = {}
    if multi_level_headers:
        for (mer_row, mer_col), mer_info in merged_cells.items():
            mer_row_idx = mer_row - start_row
            mer_col_idx = mer_col - start_col
            if mer_row_idx < header_rows:
                for r in range(mer_row_idx, mer_row_idx + mer_info["rowspan"]):
                    for c in range(mer_col_idx, mer_col_idx + mer_info["colspan"]):
                        if (r, c) != (mer_row_idx, mer_col_idx):
                            merged_cell_parents[(r, c)] = (mer_row_idx, mer_col_idx)

    common = dict(
        df=df, start_row=start_row, start_col=start_col, end_col=end_col,
        header_rows=header_rows, category_row_set=category_row_set,
        last_row_is_total=last_row_is_total, merged_cells=merged_cells,
        multi_level_headers=multi_level_headers,
        merged_cell_parents=merged_cell_parents,
        col_config_by_idx=col_config_by_idx, heatmap_data=heatmap_data,
        cell_padding=cell_padding, font_family=font_family, font_size=font_size,
        text_align=text_align, number_align=number_align,
        bold_first_col=bold_first_col,
        indent_child_rows=indent_child_rows, child_indent=child_indent,
    )

    # Build both theme tables
    html_light = _build_table_html(**common, **_resolve_theme_colors("light"))
    html_dark  = _build_table_html(**common, **_resolve_theme_colors("dark"))

    # ── Stable wrapper ID for the setChartTheme DOM lookup ───────────────────
    # We use a stable ID derived from the output filename (or a uuid fallback)
    # so the IIFE can find the wrapper after slidejs re-injects the script.
    # document.currentScript is null when a script is re-injected by slidejs,
    # so we embed the wrapper ID directly into the script at generation time.
    import hashlib as _hl
    _id_seed = str(output_file) if output_file else str(id(html_light))
    static_uid = "dtw_" + _hl.md5(_id_seed.encode()).hexdigest()[:10]

    html = f"""<div class="dual-theme-table-wrapper" data-dtw-uid="{static_uid}" style="width:100%;height:100%;overflow:auto;" data-chart-ready="true">
  <div class="table-light" style="display:{'block' if theme == 'light' else 'none'};">
{html_light}
  </div>
  <div class="table-dark" style="display:{'none' if theme == 'light' else 'block'};">
{html_dark}
  </div>
</div>
<script>
(function() {{
  var uid = '{static_uid}';
  window._dtwSeq = (window._dtwSeq || 0) + 1;
  var runtimeId = 'dtw_r_' + window._dtwSeq + '_' + Date.now();
  var allWrappers = document.querySelectorAll('[data-dtw-uid="' + uid + '"]:not([id])');
  var w = allWrappers[allWrappers.length - 1];
  if (w) w.id = runtimeId;

  var attempts = 0;
  function register() {{
    var el = document.getElementById(runtimeId);
    if (!el) {{
      if (attempts++ < 20) {{ setTimeout(register, 50); }}
      return;
    }}
    var isDark = !!window.isDarkMode;
    var lightDiv = el.querySelector('.table-light');
    var darkDiv  = el.querySelector('.table-dark');
    if (lightDiv) lightDiv.style.display = isDark ? 'none'  : 'block';
    if (darkDiv)  darkDiv.style.display  = isDark ? 'block' : 'none';
    window.setChartTheme = function(isDark) {{
      if (lightDiv) lightDiv.style.display = isDark ? 'none'  : 'block';
      if (darkDiv)  darkDiv.style.display  = isDark ? 'block' : 'none';
    }};
  }}
  register();
}})();
</script>"""

    if output_file:
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"   ✅ Table saved to: {output_path}")
        print(f"   📦 File size: {len(html):,} characters (dual-theme)")
        return str(output_path)
    else:
        print(f"   ✅ HTML generated: {len(html):,} characters (dual-theme)")
        return f"TEXT:{html}"


# ===========================================
# CONVENIENCE WRAPPER FUNCTIONS

    # Determine range
    if end_row is None:
        end_row = len(df)
        while end_row > start_row and df.iloc[end_row - 1].isna().all():
            end_row -= 1

    if end_col is None:
        end_col = len(df.columns)

    # Slice dataframe
    start_row_idx = start_row - 1
    end_row_idx = end_row
    start_col_idx = start_col - 1
    end_col_idx = end_col

    df = df.iloc[start_row_idx:end_row_idx, start_col_idx:end_col_idx]

    print(f"   ✓ Table size: {len(df)} rows x {len(df.columns)} columns")

    # Get column headers (from first header row)
    col_headers = []
    for col_idx in range(len(df.columns)):
        header_val = df.iloc[0, col_idx]
        if pd.isna(header_val):
            # For multi-level headers, look for merged cell above
            excel_col = start_col + col_idx
            excel_row = start_row
            # Find if this is part of a merged header
            parent_header = None
            for (mer_row, mer_col), mer_info in merged_cells.items():
                if mer_row == excel_row:
                    if mer_col <= excel_col < mer_col + mer_info["colspan"]:
                        parent_cell = ws.cell(mer_row, mer_col).value
                        parent_header = parent_cell
                        break
            col_headers.append(parent_header if parent_header else f"Column_{col_idx}")
        else:
            col_headers.append(str(header_val))

    # Parse column_config
    col_config_by_idx = {}
    if column_config:
        for col_identifier, config in column_config.items():
            if isinstance(col_identifier, int):
                # Direct index provided
                col_config_by_idx[col_identifier] = config
            else:
                # Column name provided - find index
                try:
                    col_idx = col_headers.index(col_identifier)
                    col_config_by_idx[col_idx] = config
                except ValueError:
                    print(f" ⚠️ Warning: Column '{col_identifier}' not found in headers")

    # Prepare category rows (convert to 0-indexed relative to df)
    category_row_set = set()
    if category_rows:
        category_row_set = {r - 1 for r in category_rows}  # Convert to 0-indexed
        print(f"   ✓ Category rows: {category_rows}")

    # Check for total row
    last_row_is_total = False
    if len(df) > 0:
        first_cell = str(df.iloc[-1, 0]).lower()
        if "total" in first_cell or "sum" in first_cell or "grand" in first_cell:
            last_row_is_total = True

    # Calculate heatmap percentiles per column
    heatmap_data = {}
    for col_idx, config in col_config_by_idx.items():
        if config.get("heatmap"):
            # Get all numeric values in this column (excluding headers and category rows)
            values = []
            for row_idx in range(header_rows, len(df)):
                if row_idx in category_row_set:
                    continue
                if row_idx == len(df) - 1 and last_row_is_total:
                    continue
                val = df.iloc[row_idx, col_idx]
                if isinstance(val, (int, float)) and not pd.isna(val):
                    values.append(val)

            if values:
                heatmap_data[col_idx] = {
                    "values": values,
                    "percentiles": np.percentile(values, [0, 33, 67, 100]),
                }

    # Build merged cell tracking for multi-level headers
    # Map: (row_idx, col_idx) -> parent merged cell info
    merged_cell_parents = {}
    if multi_level_headers:
        for (mer_row, mer_col), mer_info in merged_cells.items():
            mer_row_idx = mer_row - start_row
            mer_col_idx = mer_col - start_col

            if mer_row_idx < header_rows:
                for r in range(mer_row_idx, mer_row_idx + mer_info["rowspan"]):
                    for c in range(mer_col_idx, mer_col_idx + mer_info["colspan"]):
                        if (r, c) != (mer_row_idx, mer_col_idx):
                            merged_cell_parents[(r, c)] = (mer_row_idx, mer_col_idx)

    # Start HTML
    html = f'<table style="width:100%; border-collapse:collapse; font-family:{font_family}; font-size:{font_size};">\n'

    # Track cells to skip (merged cells)
    skip_cells = set()

    # Generate rows
    for row_idx in range(len(df)):
        excel_row = start_row + row_idx

        # Determine row type
        is_header = row_idx < header_rows
        is_category = row_idx in category_row_set
        is_total_row = (row_idx == len(df) - 1) and last_row_is_total
        is_child_row = not is_header and not is_category and not is_total_row

        # Start row
        if is_header:
            if row_idx == 0:
                html += "  <thead>\n"

            bg = header_bg if row_idx == 0 else subheader_bg
            html += f'    <tr style="background:{bg}; color:{header_color}; font-weight:bold;">\n'
        elif is_category:
            html += f'    <tr style="background:{category_bg}; color:{category_color}; font-weight:bold;">\n'
        elif is_total_row:
            html += f'    <tr style="background:{total_row_bg}; color:{total_row_color}; font-weight:bold;">\n'
        else:
            bg_color = row_colors[row_idx % 2]
            html += f'    <tr style="background:{bg_color};">\n'

        for col_idx in range(len(df.columns)):
            excel_col = start_col + col_idx

            # Skip if part of merged cell
            if (excel_row, excel_col) in skip_cells:
                continue

            # Get cell value
            cell_value = df.iloc[row_idx, col_idx]
            if pd.isna(cell_value):
                cell_value = ""

            # Check for merged cell
            merge_info = merged_cells.get(
                (excel_row, excel_col), {"colspan": 1, "rowspan": 1}
            )
            colspan = merge_info["colspan"]
            rowspan = merge_info["rowspan"]

            # Mark cells to skip
            for r in range(excel_row, excel_row + rowspan):
                for c in range(excel_col, excel_col + colspan):
                    if (r, c) != (excel_row, excel_col):
                        skip_cells.add((r, c))

            # MINIMAL HORIZONTAL-ONLY BORDERS FOR ALL ROW TYPES
            # Check if this cell is part of a merged header parent (for seamless look)
            is_merged_header_continuation = False
            if (
                multi_level_headers
                and is_header
                and (row_idx, col_idx) in merged_cell_parents
            ):
                is_merged_header_continuation = True

            # Build minimal border style
            if is_header:
                if colspan > 1:
                    # This is a merged header cell (like Drug A) - no internal borders
                    cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:none;"
                    # Add left border only for first column
                    if col_idx == 0:
                        cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:1px solid {border_color}; border-right:none;"
                    # Add right border only for last column in this merged range
                    if col_idx + colspan == len(df.columns):
                        cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:1px solid {border_color};"
                elif is_merged_header_continuation:
                    # This cell is under a merged parent - no borders for seamless look
                    cell_style = f"padding:{cell_padding}; border:none;"
                else:
                    # Regular header cell - horizontal borders only
                    cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:none;"
                    # Add left border for first column
                    if col_idx == 0:
                        cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:1px solid {border_color}; border-right:none;"
                    # Add right border for last column
                    if col_idx == len(df.columns) - 1:
                        cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:1px solid {border_color};"

            elif is_category:
                # Category rows: NO INTERNAL BORDERS - seamless horizontal bar
                # Only top and bottom borders, no left/right borders at all
                cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:none;"

            elif is_total_row:
                # Total rows: similar to category rows but with different background
                cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:none;"
            else:
                # Data rows: horizontal borders only, no vertical borders
                cell_style = f"padding:{cell_padding}; border-top:1px solid {border_color}; border-bottom:1px solid {border_color}; border-left:none; border-right:none;"

            # Track merged header parents for multi-level headers
            if multi_level_headers and is_header and row_idx > 0 and cell_value == "":
                parent_bg = header_bg if row_idx == 1 else subheader_bg
                for (mer_row, mer_col), mer_info in merged_cells.items():
                    if (
                        mer_row < excel_row
                        and mer_col <= excel_col < mer_col + mer_info["colspan"]
                    ):
                        if mer_row + mer_info["rowspan"] > excel_row:
                            parent_bg = header_bg
                            break

                cell_style += f" background:{parent_bg};"

            # Indentation for child rows
            if is_child_row and indent_child_rows and col_idx == 0:
                cell_style += f" padding-left:{child_indent};"

            # Alignment
            if colspan > 1:
                cell_style += " text-align:center;"
            else:
                cell_style += f" text-align:{text_align};"

            if rowspan > 1:
                cell_style += " vertical-align:middle;"

            # Bold first column (legacy)
            if bold_first_col and col_idx == 0 and not is_header and not is_category:
                cell_style += " font-weight:bold;"

            # Get column config
            col_config = col_config_by_idx.get(col_idx, {})

            # Apply heatmap - BACKGROUND ONLY, keep text color consistent
            if (
                col_config.get("heatmap")
                and col_idx in heatmap_data
                and not is_header
                and not is_category
                and not is_total_row
            ):
                if isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                    ranges = col_config.get("ranges", [])
                    percentiles = heatmap_data[col_idx]["percentiles"]

                    # Determine which range this value falls into
                    for range_config in ranges:
                        pct_range = range_config.get("percentile", (0, 100))
                        min_pct, max_pct = pct_range

                        # Get actual values at these percentiles
                        min_val = np.percentile(
                            heatmap_data[col_idx]["values"], min_pct
                        )
                        max_val = np.percentile(
                            heatmap_data[col_idx]["values"], max_pct
                        )

                        if min_val <= cell_value <= max_val:
                            bg_color = range_config.get("color", "#cccccc")
                            # USE CONSISTENT TEXT COLOR - no dynamic text_color from config
                            cell_style += f" background:{bg_color}; color:{data_text_color}; font-weight:bold;"
                            break

            # Number formatting
            number_format = col_config.get("number_format", "auto")
            decimal_places = col_config.get("decimal_places", 2)

            if isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                if number_format == "percentage":
                    if cell_value < 1:  # 0.123 format
                        cell_value = f"{cell_value * 100:.{decimal_places}f}%"
                    else:  # Already in percentage (12.3 format)
                        cell_value = f"{cell_value:.{decimal_places}f}%"
                elif number_format == "integer":
                    cell_value = f"{int(cell_value):,}"
                elif number_format == "decimal":
                    cell_value = f"{cell_value:,.{decimal_places}f}"
                else:  # auto
                    if cell_value == int(cell_value):
                        cell_value = f"{int(cell_value):,}"
                    else:
                        cell_value = f"{cell_value:,.{decimal_places}f}"

                # Right align numbers (unless merged)
                if colspan == 1:
                    cell_style += f" text-align:{number_align};"

            # Build cell attributes
            cell_attrs = ""
            if colspan > 1:
                cell_attrs += f'colspan="{colspan}" '
            if rowspan > 1:
                cell_attrs += f'rowspan="{rowspan}" '
            cell_attrs += f'style="{cell_style}"'

            tag = "th" if is_header else "td"

            html += f"      <{tag} {cell_attrs}>{cell_value}</{tag}>\n"

        html += "    </tr>\n"

        # Close thead
        if is_header and row_idx == header_rows - 1:
            html += "  </thead>\n  <tbody>\n"

    html += "  </tbody>\n</table>"

    # Output
    if output_file:
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)

        print(f"   ✅ Table saved to: {output_path}")
        print(f"   📦 File size: {len(html):,} characters")
        return str(output_path)
    else:
        print(f"   ✅ HTML generated: {len(html):,} characters")
        return f"TEXT:{html}"


# ===========================================
# CONVENIENCE WRAPPER FUNCTIONS
# ===========================================


def create_pivot_table(
    excel_file,
    sheet_name,
    output_file,
    category_rows,
    heatmap_columns=None,
    column_config=None,  # Allow passing additional column configs
    **kwargs,
):
    """
    Create a pivot table with category rows and optional heatmaps.

    Example:
    ---
    create_pivot_table(
        excel_file='data.xlsx',
        sheet_name='Patient_Volume',
        output_file='tables/pivot.htmltable',
        category_rows=[1, 5, 9],
        heatmap_columns=['Drug A LOYALIST', 'Drug B LOYALIST', 'NON-LOYALIST'],
        column_config={'Grand Total': {'number_format': 'percentage', 'decimal_places': 1}}
    )
    """

    # Build column config with default 3-tier heatmap (background only, no text_color)
    final_column_config = column_config or {}  # Start with user-provided config

    if heatmap_columns:
        for col_name in heatmap_columns:
            # Only add if not already configured by user
            if col_name not in final_column_config:
                final_column_config[col_name] = {
                    "heatmap": True,
                    "ranges": [
                        {"percentile": (33, 100), "color": "#2F2A9190"},  # Top/Mid
                        {"percentile": (0, 33), "color": "#2F2A9190"},  # Bottom
                    ],
                    "number_format": "percentage",
                    "decimal_places": 1,
                }

    # Call main function
    return excel_table_to_html(
        excel_file=excel_file,
        sheet_name=sheet_name,
        output_file=output_file,
        category_rows=category_rows,
        column_config=final_column_config,
        indent_child_rows=True,
        **kwargs,
    )


def create_multilevel_table(
    excel_file,
    sheet_name,
    output_file,
    category_rows,
    header_rows=2,
    heatmap_col_ranges=None,
    exclude_cols=None,
    heatmap_ranges=None,
    column_config=None,
    **kwargs,
):
    """
    Create a multi-level header table with heatmaps applied to column ranges.

    Parameters:
    ---
    excel_file : str
        Path to Excel file
    sheet_name : str
        Sheet name
    output_file : str
        Output file path
    category_rows : list
        List of row indices (1-indexed) that are category headers
    header_rows : int
        Number of header rows (default=2 for multi-level)
    heatmap_col_ranges : list of tuples
        List of (start_col, end_col) ranges to apply heatmap (1-indexed)
    exclude_cols : list
        List of column indices to exclude from heatmap (1-indexed)
    heatmap_ranges : list
        Custom heatmap percentile ranges (optional)
    column_config : dict
        Additional column-specific configurations

    Example:
    ---
    create_multilevel_table(
        excel_file='data.xlsx',
        sheet_name='Age_Breakdown',
        output_file='tables/age_breakdown.htmltable',
        category_rows=[3, 8, 13],
        header_rows=2,
        heatmap_col_ranges=[(2, 8), (10, 16)], # Apply heatmap to columns 2-8 and 10-16
        exclude_cols=[9, 17, 18], # Exclude Drug A Total, Drug B Total, Grand Total
        column_config={
            9: {'number_format': 'percentage', 'decimal_places': 1}, # Drug A Total
            17: {'number_format': 'percentage', 'decimal_places': 1}, # Drug B Total
            18: {'number_format': 'percentage', 'decimal_places': 1} # Grand Total
        }
    )

    Or with custom heatmap ranges:
    create_multilevel_table(
        excel_file='data.xlsx',
        sheet_name='Age_Breakdown',
        output_file='tables/age_breakdown.htmltable',
        category_rows=[3, 8, 13],
        header_rows=2,
        heatmap_col_ranges=[(2, 8), (10, 16)],
        heatmap_ranges=[
            {"percentile": (67, 100), "color": "#28a745"}, # Custom colors
            {"percentile": (33, 67), "color": "#ffc107"},
            {"percentile": (0, 33), "color": "#dc3545"}
        ],
        column_config={
            9: {'number_format': 'percentage', 'decimal_places': 1},
            17: {'number_format': 'percentage', 'decimal_places': 1},
            18: {'number_format': 'percentage', 'decimal_places': 1}
        }
    )
    """

    # Start with user-provided config (if any)
    final_column_config = column_config.copy() if column_config else {}

    # Default heatmap ranges if not provided
    if heatmap_ranges is None:
        heatmap_ranges = [
            {"percentile": (33, 100), "color": "#2F2A9190"},  # Top/Mid
            {"percentile": (0, 33), "color": "#2F2A9190"},  # Bottom
        ]

    # Apply heatmap to specified column ranges
    if heatmap_col_ranges:
        # First, determine all column headers by reading the Excel file
        # df_temp = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, nrows=header_rows)

        # Convert to 0-indexed for internal use
        # exclude_set = {col - 1 for col in (exclude_cols or [])}

        # Apply heatmap to each range
        for start_col, end_col in heatmap_col_ranges:
            for col_idx in range(start_col - 1, end_col):  # Convert to 0-indexed
                if exclude_cols and (col_idx + 1) in exclude_cols:
                    continue

                # if col_idx not in exclude_set:
                # If column not already configured
                if col_idx not in final_column_config:
                    final_column_config[col_idx] = {
                        "heatmap": True,
                        "ranges": heatmap_ranges,
                        "number_format": "percentage",
                        "decimal_places": 1,
                    }
                    # if heatmap_ranges:
                    #     final_column_config[col_idx]["ranges"] = heatmap_ranges

    # Call main function
    return excel_table_to_html(
        excel_file=excel_file,
        sheet_name=sheet_name,
        output_file=output_file,
        category_rows=category_rows,
        column_config=final_column_config,
        multi_level_headers=True,
        indent_child_rows=True,
        **kwargs,
    )


if __name__ == "__main__":
    create_pivot_table(
        excel_file=r"C:\my_disk\projects\visual_library\slidejs\data\excel_pivots.xlsx",
        sheet_name="data",
        output_file=r"C:\my_disk\projects\visual_library\slidejs\data\excel_pivots_data.htmltable",
        category_rows=[2, 6],
        heatmap_columns=["A", "B", "C"],
        column_config={
            "Grand Total": {"number_format": "percentage", "decimal_places": 1}
        },
        compact_mode="ultra-compact",
        border_color="#e0e0e0",
        row_colors=("#fafafa", "white"),
        font_size="8px",
    )

    create_multilevel_table(
        excel_file="data/patient_volume.xlsx",
        sheet_name="Sheet1",
        output_file="tables/age_breakdown.htmltable",
        category_rows=[3, 8, 13],
        header_rows=2,
        heatmap_col_ranges=[
            (2, 8),  # Drug A age columns
            (10, 16),  # Drug B age columns
        ],
        exclude_cols=[9, 17, 18],
        column_config={
            "Drug A Total": {
                "number_format": "percentage",
                "decimal_places": 1,
            },  # Drug A® Total
            "Drug B Total": {
                "number_format": "percentage",
                "decimal_places": 1,
            },  # Drug B Total
            "Grand Total": {
                "number_format": "percentage",
                "decimal_places": 1,
            },  # Grand Total
        },
        compact_mode="ultra-compact",
        border_color="#e0e0e0",
        row_colors=("#fafafa", "white"),
        font_size="9px",
    )