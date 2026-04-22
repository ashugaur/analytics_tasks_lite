import re
import pandas as pd


# ── Optional: install coloraide if not present ────────────────────────────────
# pip install coloraide

try:
    import coloraide
    from coloraide import Color
    COLORAIDE_AVAILABLE = True
except ImportError:
    COLORAIDE_AVAILABLE = False
    print("⚠ coloraide not installed — run: pip install coloraide")
    print("  Falling back to storing original values unchanged.")


# ── Color conversion ──────────────────────────────────────────────────────────

def to_hex(color_str: str) -> str:
    """
    Convert any CSS color string (oklch, hsl, rgb, hex, named) to the
    closest sRGB hex value.

    oklch colors outside the sRGB gamut are clipped to the nearest
    in-gamut color using coloraide's CSS gamut mapping (LCH chroma
    reduction), which preserves perceived lightness and hue better
    than simple channel clamping.

    Returns the original string unchanged if:
      - coloraide is not installed
      - the string is already a valid hex
      - the string is empty or cannot be parsed
    """
    if not color_str or not isinstance(color_str, str):
        return color_str

    color_str = color_str.strip()

    # Already hex — nothing to do
    if re.match(r'^#[0-9a-fA-F]{3,8}$', color_str):
        return color_str

    # rgba(...) strings — ECharts understands these natively, keep as-is
    if color_str.startswith("rgba("):
        return color_str

    if not COLORAIDE_AVAILABLE:
        return color_str

    try:
        c = Color(color_str)

        # Clip to sRGB gamut if needed (oklch can exceed sRGB)
        if not c.in_gamut("srgb"):
            c = c.convert("srgb").clip()
        else:
            c = c.convert("srgb")

        # Format as #rrggbb
        r = round(c.get("red")   * 255)
        g = round(c.get("green") * 255)
        b = round(c.get("blue")  * 255)

        # Clamp channels to 0-255 (safety after gamut clip)
        r, g, b = max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b))

        return f"#{r:02x}{g:02x}{b:02x}"

    except Exception as e:
        print(f"  ⚠ Could not convert '{color_str}': {e} — keeping original")
        return color_str


# ── Tier 1 chrome mapping ─────────────────────────────────────────────────────
# (chart_element, element_name) → (light_token, dark_token)
# element_name of None means "match any element_name for this chart_element"

CHROME_MAPPING = {
    # Chart titles
    ("chart_title",    None):           ("text",         "text"),
    ("chart_subtitle", None):           ("text_muted",   "text_muted"),

    # Axis chrome
    ("axis_title",     None):           ("text",         "text"),
    ("axis_label",     None):           ("text_muted",   "text_muted"),
    ("axis_line",      None):           ("border",       "border"),
    ("axis_tick",      None):           ("border",       "border"),

    # Grid
    ("gridline",       None):           ("border_muted", "border_muted"),
    ("gridline",       "major"):        ("border_muted", "border_muted"),
    ("gridline",       "main"):         ("border_muted", "border_muted"),

    # Legend (both naming conventions in your file)
    ("legend_text",    None):           ("text",         "text"),
    ("legend",         "text"):         ("text",         "text"),

    # Annotations
    ("bar_label",      None):           ("text",         "text_muted"),
    ("bar_label",      "bar_label"):    ("text",         "text_muted"),
    ("bar_label",      "total"):        ("text",         "text_muted"),
    ("xaxis",          None):           ("text_muted",   "text_muted"),
    ("xaxis",          "xaxis"):        ("text_muted",   "text_muted"),
    ("stage_total",    None):           ("text",         "text_muted"),
    ("stage_total",    "stage_total"):  ("text",         "text_muted"),

    # Table (embedded)
    ("table",          "text"):         ("text",         "text"),
    ("table",          "header"):       ("text",         "text"),

    # Tooltips — bar_chart style (chart_element carries the full role)
    ("tooltip_bg",     None):           ("bg_light",     "bg"),
    ("tooltip_bg",     "main"):         ("bg_light",     "bg"),
    ("tooltip_border", None):           ("border_muted", "border"),
    ("tooltip_border", "main"):         ("border_muted", "border"),
    ("tooltip_text",   None):           ("text",         "text"),
    ("tooltip_text",   "main"):         ("text",         "text"),

    # Tooltips — line_chart style (element_name carries the role)
    ("tooltip",        "bg"):           ("bg_light",     "bg"),
    ("tooltip",        "border"):       ("border_muted", "border"),
    ("tooltip",        "text"):         ("text",         "text"),
}

# Tier 2: data/brand colors — never modified
TIER2_ELEMENTS = {"series", "category"}


def get_chrome_mapping(chart_element: str, element_name: str) -> tuple | None:
    """
    Look up (light_token, dark_token) for a chart_element / element_name pair.
    Tries exact match first, then falls back to wildcard (None) on element_name.
    """
    el = str(chart_element).strip().lower() if pd.notna(chart_element) else ""
    en = str(element_name).strip().lower()  if pd.notna(element_name)  else ""

    if (el, en) in CHROME_MAPPING:
        return CHROME_MAPPING[(el, en)]
    if (el, None) in CHROME_MAPPING:
        return CHROME_MAPPING[(el, None)]
    return None


# ── CSS parser (shared with parse_ui_colors_to_theme_df) ─────────────────────

_VAR_MAP = {
    "bg-dark": "bg_dark", "bg": "bg", "bg-light": "bg_light",
    "text": "text", "text-muted": "text_muted", "highlight": "highlight",
    "border": "border", "border-muted": "border_muted",
    "primary": "primary", "secondary": "secondary",
    "danger": "danger", "warning": "warning",
    "success": "success", "info": "info",
}


def _parse_vars(block: str, prefer: str = "oklch") -> dict:
    hsl_vals, oklch_vals = {}, {}
    for m in re.finditer(r'--([a-zA-Z0-9-]+)\s*:\s*([^;]+);', block):
        name  = m.group(1).strip()
        value = m.group(2).strip()
        if name not in _VAR_MAP:
            continue
        col = _VAR_MAP[name]
        if value.startswith("oklch"):
            oklch_vals[col] = value
        elif value.startswith("hsl"):
            hsl_vals[col]   = value
    return {**hsl_vals, **oklch_vals} if prefer == "oklch" else {**oklch_vals, **hsl_vals}


def _extract_block(text: str, selector: str) -> str | None:
    m = re.search(
        re.escape(selector) + r'\s*\{([^}]*(?:\{[^}]*\}[^}]*)*)\}',
        text, re.DOTALL
    )
    return m.group(1) if m else None


def _parse_both_themes(css_text: str) -> tuple[dict, dict]:
    """
    Returns (light_vars, dark_vars).
    Site convention: :root = dark theme, body.light = light theme.
    """
    root_block  = _extract_block(css_text, ":root")
    light_block = _extract_block(css_text, "body.light")

    if root_block and light_block:
        return _parse_vars(light_block), _parse_vars(root_block)
    elif root_block:
        all_vars = _parse_vars(root_block)
        return all_vars, all_vars
    else:
        all_vars = _parse_vars(css_text)
        return all_vars, all_vars


# ── Main function ─────────────────────────────────────────────────────────────

def apply_theme_to_colors(
    colors_path,
    sheet_name: str,
    css_text: str,
    output_sheet_suffix: str = None,
    save: bool = True,
    convert_to_hex: bool = True,
) -> pd.DataFrame:
    """
    Reads the colors Excel file, applies a new theme to all Tier 1 (chrome)
    rows converting all color values to hex, and returns the updated DataFrame.

    Tier 2 rows (series / category) are never touched.

    Parameters
    ----------
    colors_path : str or Path
        Path to colors.xlsm
    sheet_name : str
        Source sheet, e.g. 'colorsEcharts'
    css_text : str
        Full CSS pasted from iamsajid.com/ui-colors/
    output_sheet_suffix : str, optional
        New sheet saved as '{sheet_name}_{suffix}', e.g. 'colorsEcharts_green'
    save : bool
        Write result back to the Excel file
    convert_to_hex : bool
        Convert oklch/hsl values to hex before storing (default True)

    Returns
    -------
    pd.DataFrame with updated light_hex / dark_hex columns
    """
    from openpyxl import load_workbook

    print("=" * 60)
    print("🎨 apply_theme_to_colors")
    print("=" * 60)

    # ── Step 1: parse CSS tokens ──────────────────────────────────────────────
    light_vars, dark_vars = _parse_both_themes(css_text)
    print(f"\n📋 Tokens parsed:")
    print(f"  Light: {len(light_vars)} — {list(light_vars.keys())}")
    print(f"  Dark:  {len(dark_vars)} — {list(dark_vars.keys())}")

    # Pre-convert all token values to hex once (avoids repeated conversions)
    if convert_to_hex:
        light_vars_hex = {k: to_hex(v) for k, v in light_vars.items()}
        dark_vars_hex  = {k: to_hex(v) for k, v in dark_vars.items()}
        print(f"\n🔄 Token → hex preview (light):")
        for k, v in light_vars.items():
            print(f"  {k:15s}  {v:30s}  →  {light_vars_hex[k]}")
        print(f"\n🔄 Token → hex preview (dark):")
        for k, v in dark_vars.items():
            print(f"  {k:15s}  {v:30s}  →  {dark_vars_hex[k]}")
    else:
        light_vars_hex = light_vars
        dark_vars_hex  = dark_vars

    # ── Step 2: load the colors sheet ────────────────────────────────────────
    df = pd.read_excel(colors_path, sheet_name=sheet_name, dtype=str)
    df = df.fillna("")
    print(f"\n📊 Loaded '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")

    # ── Step 3: apply mapping row by row ─────────────────────────────────────
    print(f"\n⚙️  Applying theme...")
    updated_count = 0
    skipped_tier2 = 0
    unmatched     = []

    for idx, row in df.iterrows():
        chart_element = str(row.get("chart_element", "")).strip().lower()
        element_name  = str(row.get("element_name",  "")).strip().lower()

        # Tier 2 — data/brand colors, never touch
        if chart_element in TIER2_ELEMENTS:
            skipped_tier2 += 1
            continue

        mapping = get_chrome_mapping(chart_element, element_name)

        if mapping is None:
            unmatched.append(
                f"  row {idx+2:>3}: chart_element='{chart_element}'"
                f"  element_name='{element_name}'"
            )
            continue

        light_token, dark_token = mapping

        new_light = light_vars_hex.get(light_token, "")
        new_dark  = dark_vars_hex.get(dark_token,   "")

        if new_light:
            df.at[idx, "light_hex"] = new_light
        if new_dark:
            df.at[idx, "dark_hex"] = new_dark

        updated_count += 1

    print(f"\n📈 Results:")
    print(f"  ✓ Chrome rows updated : {updated_count}")
    print(f"  ✓ Tier 2 rows kept    : {skipped_tier2} (untouched)")
    if unmatched:
        print(f"  ⚠ Unmatched rows      : {len(unmatched)} (kept original)")
        for u in unmatched:
            print(u)
    else:
        print(f"  ✓ Unmatched rows      : 0")

    # ── Step 4: save to new sheet ─────────────────────────────────────────────
    if save:
        suffix       = output_sheet_suffix or "updated"
        target_sheet = f"{sheet_name}_{suffix}"

        wb = load_workbook(colors_path, keep_vba=True)

        if target_sheet in wb.sheetnames:
            del wb[target_sheet]
            print(f"\n  ♻️  Removed existing sheet '{target_sheet}'")

        ws = wb.create_sheet(target_sheet)
        ws.append(list(df.columns))
        for _, data_row in df.iterrows():
            ws.append([data_row[c] for c in df.columns])

        wb.save(colors_path)
        print(f"  ✓ Saved to sheet '{target_sheet}'")
        print(f"  ✓ File: {colors_path}")

    print("\n" + "=" * 60)
    return df


if __name__ == "__main__":
    # ── Usage ─────────────────────────────────────────────────────────────────────

    # _colors = r"C:/my_disk/projects/visual_library/____settings/colors.xlsm"

    css_green = """
    :root {
    --bg-dark: hsl(152 100% 0%);
    --bg: hsl(155 100% 1%);
    --bg-light: hsl(160 100% 3%);
    --text: hsl(157 100% 86%);
    --text-muted: hsl(158 44% 60%);
    --highlight: hsl(167 100% 12%);
    --border: hsl(165 100% 7%);
    --border-muted: hsl(162 100% 3%);
    --primary: hsl(162 60% 51%);
    --secondary: hsl(328 79% 74%);
    --danger: hsl(8 74% 66%);
    --warning: hsl(52 76% 40%);
    --success: hsl(154 54% 47%);
    --info: hsl(217 87% 69%);
    /* oklch */
    --bg-dark: oklch(0.1 0.07 168);
    --bg: oklch(0.15 0.07 168);
    --bg-light: oklch(0.2 0.07 168);
    --text: oklch(0.96 0.1 168);
    --text-muted: oklch(0.76 0.1 168);
    --highlight: oklch(0.5 0.14 168);
    --border: oklch(0.4 0.14 168);
    --border-muted: oklch(0.3 0.14 168);
    --primary: oklch(0.76 0.14 168);
    --secondary: oklch(0.76 0.14 348);
    --danger: oklch(0.7 0.14 30);
    --warning: oklch(0.7 0.14 100);
    --success: oklch(0.7 0.14 160);
    --info: oklch(0.7 0.14 260);
    }
    body.light {
    /* hsl (fallback color) */
    --bg-dark: hsl(156 74% 84%);
    --bg: hsl(156 100% 89%);
    --bg-light: hsl(156 100% 94%);
    --text: hsl(151 100% 0%);
    --text-muted: hsl(165 100% 7%);
    --highlight: hsl(159 100% 85%);
    --border: hsl(168 100% 21%);
    --border-muted: hsl(165 87% 39%);
    --primary: hsl(165 100% 7%);
    --secondary: hsl(323 63% 29%);
    --danger: hsl(7 55% 41%);
    --warning: hsl(53 100% 14%);
    --success: hsl(161 100% 15%);
    --info: hsl(217 58% 44%);
    /* oklch */
    --bg-dark: oklch(0.92 0.07 168);
    --bg: oklch(0.96 0.07 168);
    --bg-light: oklch(1 0.07 168);
    --text: oklch(0.15 0.14 168);
    --text-muted: oklch(0.4 0.14 168);
    --highlight: oklch(1 0.14 168);
    --border: oklch(0.6 0.14 168);
    --border-muted: oklch(0.7 0.14 168);
    --primary: oklch(0.4 0.14 168);
    --secondary: oklch(0.4 0.14 348);
    --danger: oklch(0.5 0.14 30);
    --warning: oklch(0.5 0.14 100);
    --success: oklch(0.5 0.14 160);
    --info: oklch(0.5 0.14 260);
    }
    """

    df_result = apply_theme_to_colors(
        colors_path         = _colors,
        sheet_name          = "colorsEcharts",
        css_text            = css_green,
        output_sheet_suffix = "green",
        save                = True,
        convert_to_hex      = True,
    )

    # Preview result
    with pd.option_context("display.max_columns", None, "display.width", 200):
        print(df_result[["topic", "chart_type", "chart_element",
                        "element_name", "light_hex", "dark_hex"]].to_string(index=False))
