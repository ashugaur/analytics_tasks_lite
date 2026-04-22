import pandas as pd
import json
import gzip
import base64
from pathlib import Path


def generate_data_table(
    df,
    output_file="data_table.html",
    title="Data Table",
    # ── Formatting ────────────────────────────────────────────────
    global_column_formats=None,
    table_font_family="'SF Mono', Monaco, 'Inconsolata', 'Fira Code', monospace",
    table_font_size="11px",
    row_density="normal",           # "normal" | "compact" | "ultracompact"
    max_cell_chars=100,
    truncate_cell_text=True,
    enable_cell_expansion=True,
    freeze_columns=None,
    # ── Row / search behaviour ────────────────────────────────────
    max_rows=50,                    # rows visible on initial load
    stored_rows=500,                # rows embedded in HTML for in-page search
    include_row_search=True,
    row_search_threshold=10,
    include_column_search=True,
    column_search_threshold=5,
    enable_multiword_search=False,
    max_table_height="600px",
    table_scroll_threshold=20,
    # ── Full-data embedding (no pako dependency) ──────────────────
    embed_full_data=True,           # embed all rows as compressed JSON
    # ── Theme ─────────────────────────────────────────────────────
    default_theme="light",          # "light" | "dark"
    light_mode_colors=None,
    dark_mode_colors=None,
    theme_config_file=None,         # path to .xlsx with Theme_Config sheet
    theme=None,                     # Test_ID value in Theme_Config sheet
):
    """
    Export a pandas (or polars) DataFrame to a fully standalone HTML data table.

    Inherits the modern styling engine from data_explorer:
      - Light / dark theme toggle
      - Sortable columns
      - Column filter and row search (searches all embedded rows)
      - Cell truncation with click-to-expand modal
      - Freeze-column highlight
      - Row-density control
      - Column format strings (same syntax as data_explorer's global_column_formats)
      - Theme loading from an Excel Theme_Config sheet

    Parameters
    ----------
    df : pd.DataFrame or pl.DataFrame
        Data to render. Polars DataFrames are converted automatically.
    output_file : str
        Path for the generated HTML file.
    title : str
        Page title and table heading.
    global_column_formats : dict
        {column_name: format_string} — same syntax as data_explorer.
        Examples:
            {"revenue": "{:,.2f}", "date": "%Y-%m-%d", "pct": "{:.1%}"}
    table_font_family : str
        CSS font-family for table cells.
    table_font_size : str
        CSS font-size for table cells (e.g. "11px", "13px").
    row_density : str
        "normal" | "compact" | "ultracompact"
    max_cell_chars : int
        Characters shown before truncation.
    truncate_cell_text : bool
        Whether to truncate long cell values.
    enable_cell_expansion : bool
        Click truncated cell to see full value in a modal.
    freeze_columns : list
        Column names to visually highlight (accent header + tinted cells).
    max_rows : int
        Rows visible on first load (others hidden, revealed by search).
    stored_rows : int
        Rows embedded in the HTML for in-page search.
        Ignored when embed_full_data=True (all rows are embedded).
    include_row_search : bool
        Show a row search bar.
    row_search_threshold : int
        Minimum rows before the row search bar appears.
    include_column_search : bool
        Show a column filter bar.
    column_search_threshold : int
        Minimum columns before the column filter bar appears.
    enable_multiword_search : bool
        When True, space-separated words must ALL match (AND logic).
    max_table_height : str
        CSS max-height for the scrollable table wrapper.
    table_scroll_threshold : int
        Row count above which the height cap is applied.
    embed_full_data : bool
        When True, compress and embed the entire DataFrame so the row search
        works across all rows (not just stored_rows).
    default_theme : str
        Starting theme: "light" or "dark".
    light_mode_colors : dict
        Override any of: bg, text, sidebar_bg, table_header, accent, border, hover.
    dark_mode_colors : dict
        Same keys as light_mode_colors for dark mode.
    theme_config_file : str
        Path to an .xlsx workbook that has a Theme_Config sheet
        (same format as data_explorer).
    theme : str
        Test_ID value from the Theme_Config sheet to load.

    Returns
    -------
    str
        Absolute path to the generated HTML file.

    Examples
    --------
    # Minimal usage
    generate_data_table(df, "output/wfd.html", title="Word Frequency")

    # With formatting + dark default
    generate_data_table(
        df,
        "output/wfd.html",
        title="Word Frequency",
        global_column_formats={"frequency %": "{:.2f}", "frequency": "{:,}"},
        default_theme="dark",
        freeze_columns=["word"],
        row_density="compact",
    )

    # Load colours from an Excel theme file
    generate_data_table(
        df,
        "output/wfd.html",
        theme_config_file="themes.xlsx",
        theme="test_1",
    )
    """

    # ── 0. Polars compat ──────────────────────────────────────────────────────
    try:
        import polars as _pl
        if isinstance(df, _pl.DataFrame):
            df = df.to_pandas()
    except ImportError:
        pass

    if not isinstance(df, pd.DataFrame):
        raise TypeError("df must be a pandas or polars DataFrame")

    # ── 1. Defaults ───────────────────────────────────────────────────────────
    if global_column_formats is None:
        global_column_formats = {}
    if freeze_columns is None:
        freeze_columns = []

    density_settings = {
        "normal":       {"th_padding": "12px 8px",  "td_padding": "10px 12px"},
        "compact":      {"th_padding": "8px 6px",   "td_padding": "4px 6px"},
        "ultracompact": {"th_padding": "4px 4px",   "td_padding": "2px 4px"},
    }
    density = density_settings.get(row_density.lower(), density_settings["normal"])

    # ── 2. Theme loading ──────────────────────────────────────────────────────
    def _load_theme_config(config_file, theme_id):
        try:
            import io as _io
            import openpyxl as _xl
        except ImportError:
            print("❌ theme_config_file requires openpyxl — pip install openpyxl")
            return None, None
        try:
            _path = Path(config_file)
            if not _path.exists():
                print(f"❌ Theme file not found: {_path}")
                return None, None
            with open(_path, "rb") as _fh:
                _buf = _io.BytesIO(_fh.read())
            _wb = _xl.load_workbook(_buf, read_only=True, data_only=True, keep_vba=False)
            _sheet = "Theme_Config"
            if _sheet not in _wb.sheetnames:
                print(f"❌ Sheet '{_sheet}' not found in workbook")
                return None, None
            _ws = _wb[_sheet]
            _rows = list(_ws.iter_rows(values_only=True))
            if not _rows:
                return None, None
            _hdr = [str(c).strip().lower() if c is not None else "" for c in _rows[0]]
            def _col(n):
                k = n.strip().lower()
                if k not in _hdr:
                    raise KeyError(f"Column '{n}' missing in Theme_Config")
                return _hdr.index(k)
            _idx = {k: _col(k) for k in (
                "test_id", "theme_name", "primary", "text", "muted",
                "content_bg", "slide_bg", "bg", "bg_light", "border", "border_muted"
            )}
            _light = _dark = None
            for _r in _rows[1:]:
                _tid = str(_r[_idx["test_id"]]).strip() if _r[_idx["test_id"]] is not None else ""
                _tname = str(_r[_idx["theme_name"]]).strip().lower() if _r[_idx["theme_name"]] is not None else ""
                if _tid == str(theme_id).strip():
                    def _v(row, key):
                        val = row[_idx[key]]
                        return str(val).strip() if val is not None else ""
                    _colors = {
                        "bg":           _v(_r, "bg"),
                        "text":         _v(_r, "text"),
                        "sidebar_bg":   _v(_r, "slide_bg"),
                        "table_header": _v(_r, "content_bg"),
                        "accent":       _v(_r, "primary"),
                        "border":       _v(_r, "border"),
                        "hover":        _v(_r, "bg_light"),
                    }
                    if _tname == "light":
                        _light = _colors
                    elif _tname == "dark":
                        _dark = _colors
            if _light is None and _dark is None:
                print(f"❌ theme_id '{theme_id}' not found in Theme_Config")
                return None, None
            print(f"✅ Theme loaded: '{theme_id}'")
            return _light, _dark
        except Exception as _e:
            print(f"❌ load_theme_config error: {_e}")
            return None, None

    if theme is not None and theme_config_file is not None:
        _lc, _dc = _load_theme_config(theme_config_file, theme)
        if _lc is not None:
            light_mode_colors = _lc
        if _dc is not None:
            dark_mode_colors = _dc
    elif theme is not None:
        print("⚠️  theme= provided but theme_config_file= is missing — using defaults")

    _light_defaults = {
        "bg": "#ffffff", "text": "#333333", "sidebar_bg": "#f8f9fa",
        "table_header": "#e9ecef", "accent": "#007bff",
        "border": "#dee2e6", "hover": "#f1f3f5",
    }
    _dark_defaults = {
        "bg": "#1e1e1e", "text": "#e0e0e0", "sidebar_bg": "#2d2d2d",
        "table_header": "#3d3d3d", "accent": "#4a9eff",
        "border": "#444444", "hover": "#383838",
    }
    if light_mode_colors is None:
        light_mode_colors = _light_defaults
    else:
        for k, v in _light_defaults.items():
            light_mode_colors.setdefault(k, v)
    if dark_mode_colors is None:
        dark_mode_colors = _dark_defaults
    else:
        for k, v in _dark_defaults.items():
            dark_mode_colors.setdefault(k, v)

    # ── 3. Format helpers ─────────────────────────────────────────────────────
    def _format_value(value, fmt):
        if pd.isna(value):
            return ""
        try:
            if fmt.startswith("%"):
                return pd.to_datetime(value).strftime(fmt)
            return fmt.format(value)
        except Exception:
            return str(value)

    def _format_df(src_df):
        out = src_df.copy()
        fmts_lower = {k.lower(): v for k, v in global_column_formats.items()}
        for col in out.columns:
            col_lower = str(col).lower()
            if col_lower in fmts_lower:
                out[col] = out[col].apply(lambda x: _format_value(x, fmts_lower[col_lower]))
        return out

    # ── 4. Full-data compression ──────────────────────────────────────────────
    def _compress_df(src_df):
        """Compress full DataFrame to base64-gzipped JSON (pako-free)."""
        payload = {
            "columns": [str(c) for c in src_df.columns],
            "data": [
                [None if (isinstance(v, float) and pd.isna(v)) else
                 (v.item() if hasattr(v, "item") else v)
                 for v in row]
                for row in src_df.itertuples(index=False)
            ],
        }
        raw = json.dumps(payload, default=str).encode("utf-8")
        compressed = gzip.compress(raw, compresslevel=6)
        return base64.b64encode(compressed).decode("ascii")

    # ── 5. Build table HTML ───────────────────────────────────────────────────
    total_rows = len(df)
    total_cols = len(df.columns)
    actual_stored = total_rows if embed_full_data else min(stored_rows, total_rows)
    actual_max_rows = min(max_rows, actual_stored)

    df_display = df if embed_full_data else df.head(actual_stored)
    df_fmt = _format_df(df_display)

    query_id = "tbl_0"
    freeze_set = set(c.lower() for c in freeze_columns)

    table_html = ""

    # Compressed data block — decoded by vanilla JS (no pako needed)
    if embed_full_data and total_rows > 0:
        compressed_b64 = _compress_df(df)
        cols_js = "[" + ", ".join(f'"{c}"' for c in df.columns) + "]"
        table_html += f"""
<script>
window.fullDataCompressed_{query_id} = '{compressed_b64}';
window.fullDataColumns_{query_id} = {cols_js};
window.fullDataRows_{query_id} = {total_rows};
</script>"""

    # Search bars
    show_col_search = include_column_search and total_cols > column_search_threshold
    show_row_search = include_row_search and actual_stored > row_search_threshold

    if show_col_search or show_row_search:
        table_html += f'<div class="search-container" id="search-container-{query_id}">\n'
        if show_col_search:
            table_html += f"""  <div class="column-search-box">
    <input type="text" class="search-input" id="col-search-{query_id}"
           placeholder="🔍 Filter columns..."
           onkeyup="filterColumns('{query_id}')">
    <span class="search-count" id="col-count-{query_id}">Showing {total_cols} of {total_cols} columns</span>
  </div>"""
        if show_row_search:
            search_info = f"all {total_rows:,}" if embed_full_data else f"{actual_stored:,}"
            row_info = f"Showing {actual_max_rows:,} of {total_rows:,} rows"
            table_html += f"""  <div class="row-search-box" id="row-search-{query_id}">
    <input type="text" class="search-input"
           placeholder="▼ Search {search_info} rows..."
           onkeyup="filterRows('{query_id}', {actual_max_rows})">
    <span class="search-count" id="row-count-{query_id}">{row_info}</span>
  </div>"""
        table_html += "</div>\n"

    # Table wrapper
    apply_height_cap = max_table_height is not None and actual_stored > table_scroll_threshold
    wrapper_style = f' style="max-height:{max_table_height}; overflow-y:auto;"' if apply_height_cap else ""

    table_html += f"""<div class="table-wrapper"{wrapper_style} id="table-wrapper-{query_id}">
<table class="data-table {row_density}-density" id="table-{query_id}"
       style="font-family:{table_font_family};font-size:{table_font_size};">
<thead><tr>"""

    for col in df_fmt.columns:
        freeze_attr = ' data-freeze-column="true"' if str(col).lower() in freeze_set else ""
        table_html += f'<th data-column="{str(col).lower()}"{freeze_attr}>{col}<span class="sort-indicator">⇅</span></th>'

    table_html += "</tr></thead><tbody>"

    for idx, row in df_fmt.iterrows():
        row_class = ' class="initially-hidden"' if idx >= actual_max_rows else ""
        table_html += f'<tr data-row-id="{idx}"{row_class}>\n'
        for col, val in zip(df_fmt.columns, row):
            col_lower = str(col).lower()
            freeze_attr = ' data-freeze-column="true"' if col_lower in freeze_set else ""
            cell_str = str(val) if val is not None and not (isinstance(val, float) and pd.isna(val)) else ""
            if truncate_cell_text and len(cell_str) > max_cell_chars:
                truncated = cell_str[:max_cell_chars]
                if enable_cell_expansion:
                    escaped = cell_str.replace("&", "&amp;").replace('"', "&quot;").replace("<", "&lt;").replace(">", "&gt;")
                    table_html += (
                        f'<td data-column="{col_lower}"{freeze_attr} class="cell-truncated" '
                        f'data-full-text="{escaped}" onclick="expandCell(this)" title="Click to expand">'
                        f'{truncated}<span class="truncation-indicator">…</span></td>\n'
                    )
                else:
                    table_html += f'<td data-column="{col_lower}"{freeze_attr}>{truncated}<span class="truncation-indicator">…</span></td>\n'
            else:
                table_html += f'<td data-column="{col_lower}"{freeze_attr}>{cell_str}</td>\n'
        table_html += "</tr>\n"

    table_html += "</tbody></table></div>"

    # ── 6. Assemble full HTML ─────────────────────────────────────────────────
    lc = light_mode_colors
    dc = dark_mode_colors

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
:root {{
    --bg-color: {lc["bg"]};
    --text-color: {lc["text"]};
    --sidebar-bg: {lc["sidebar_bg"]};
    --table-header-bg: {lc["table_header"]};
    --accent-color: {lc["accent"]};
    --border-color: {lc["border"]};
    --hover-color: {lc["hover"]};
}}
[data-theme="dark"] {{
    --bg-color: {dc["bg"]};
    --text-color: {dc["text"]};
    --sidebar-bg: {dc["sidebar_bg"]};
    --table-header-bg: {dc["table_header"]};
    --accent-color: {dc["accent"]};
    --border-color: {dc["border"]};
    --hover-color: {dc["hover"]};
}}
body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background-color: var(--bg-color);
    color: var(--text-color);
    padding: 30px;
    min-height: 100vh;
    transition: background-color 0.3s, color 0.3s;
}}
.page-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 20px;
    padding-bottom: 14px;
    /*border-bottom: 2px solid var(--accent-color);*/
}}
.page-title {{
    font-size: 22px;
    font-weight: 700;
    color: var(--accent-color);
}}
.page-meta {{
    font-size: 12px;
    color: var(--text-color);
    opacity: 0.55;
    margin-top: 4px;
}}
.theme-toggle {{
    background: none;
    border: 1px solid var(--border-color);
    color: var(--text-color);
    cursor: pointer;
    padding: 6px 10px;
    border-radius: 6px;
    font-size: 14px;
    transition: all 0.2s;
    display: flex;
    align-items: center;
    gap: 6px;
}}
.theme-toggle:hover {{ background-color: var(--hover-color); }}
.theme-toggle svg {{ width:16px; height:16px; fill:currentColor; }}
/* Search bars */
.search-container {{
    display: flex;
    flex-direction: column;
    gap: 4px;
    margin-bottom: 10px;
}}
.column-search-box, .row-search-box {{
    display: flex;
    align-items: center;
    gap: 8px;
}}
.search-input {{
    flex: 1;
    padding: 5px 10px;
    border: none;
    border-bottom: 1px solid var(--border-color);
    border-radius: 0;
    font-size: 12px;
    background: transparent;
    color: var(--text-color);
    transition: border-color 0.2s;
}}
.search-input:focus {{ outline: none; border-bottom-color: var(--accent-color); }}
.search-count {{ font-size: 11px; white-space: nowrap; opacity: 0.55; }}
/* Table wrapper */
.table-wrapper {{
    overflow-x: auto;
    overflow-y: auto;
    border: 1px solid var(--border-color);
    border-radius: 6px;
    scrollbar-width: thin;
    scrollbar-color: rgba(120,120,120,0.3) transparent;
}}
.table-wrapper::-webkit-scrollbar {{ width:6px; height:6px; }}
.table-wrapper::-webkit-scrollbar-track {{ background: transparent; }}
.table-wrapper::-webkit-scrollbar-thumb {{ border-radius:999px; background:rgba(120,120,120,0.25); }}
.table-wrapper::-webkit-scrollbar-thumb:hover {{ background:rgba(120,120,120,0.5); }}
[data-theme="dark"] .table-wrapper::-webkit-scrollbar-thumb {{ background:rgba(200,200,200,0.18); }}
/* Table */
.data-table {{
    width: 100%;
    border-collapse: collapse;
    white-space: nowrap;
}}
.data-table thead {{
    background-color: var(--table-header-bg);
    position: sticky;
    top: 0;
    z-index: 10;
}}
.data-table th {{
    text-align: left;
    font-weight: 600;
    border-bottom: 2px solid var(--border-color);
    cursor: pointer;
    user-select: none;
    background-color: var(--table-header-bg);
}}
.data-table th:hover {{ background-color: var(--hover-color); }}
.sort-indicator {{ font-size:11px; margin-left:5px; opacity:0.3; }}
.data-table td {{ border-bottom: 1px solid var(--border-color); }}
.data-table tbody tr:hover {{ background-color: var(--hover-color); }}
.data-table th.hidden-column, .data-table td.hidden-column {{ display:none; }}
.data-table tr.hidden-row {{ display:none; }}
.data-table tr.initially-hidden {{ display:none; }}
.data-table tr.initially-hidden.search-match {{ display:table-row; }}
/* Row density */
.data-table.normal-density th       {{ padding: {density_settings["normal"]["th_padding"]}; }}
.data-table.normal-density td       {{ padding: {density_settings["normal"]["td_padding"]}; }}
.data-table.compact-density th      {{ padding: {density_settings["compact"]["th_padding"]}; }}
.data-table.compact-density td      {{ padding: {density_settings["compact"]["td_padding"]}; }}
.data-table.ultracompact-density th {{ padding: {density_settings["ultracompact"]["th_padding"]}; }}
.data-table.ultracompact-density td {{ padding: {density_settings["ultracompact"]["td_padding"]}; }}
/* Freeze columns */
.data-table th[data-freeze-column="true"] {{
    background-color: var(--accent-color);
    color: #fff;
    border-bottom-color: var(--accent-color);
}}
.data-table td[data-freeze-column="true"] {{
    background-color: var(--hover-color);
    font-weight: 500;
}}
.data-table tbody tr:hover td[data-freeze-column="true"] {{
    background-color: var(--border-color);
}}
/* Cell truncation */
.truncation-indicator {{ color:var(--accent-color); font-weight:bold; font-size:12px; }}
td.cell-truncated {{ cursor:pointer; max-width:300px; }}
td.cell-truncated:hover {{ background-color:var(--hover-color); outline:1px solid var(--accent-color); }}
/* Cell modal */
.cell-modal-overlay {{
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.45);
    z-index: 9000;
    align-items: center;
    justify-content: center;
}}
.cell-modal-overlay.active {{ display:flex; }}
.cell-modal {{
    background: var(--bg-color);
    color: var(--text-color);
    border: 1px solid var(--border-color);
    border-radius: 8px;
    padding: 24px;
    max-width: min(700px, 90vw);
    max-height: 70vh;
    overflow-y: auto;
    box-shadow: 0 8px 32px rgba(0,0,0,0.25);
    word-break: break-word;
    white-space: pre-wrap;
    font-size: 14px;
    line-height: 1.6;
}}
.cell-modal-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 16px;
    padding-bottom: 10px;
    border-bottom: 1px solid var(--border-color);
}}
.cell-modal-column {{
    font-weight: 600;
    font-size: 13px;
    color: var(--accent-color);
    text-transform: uppercase;
    letter-spacing: 0.04em;
}}
.cell-modal-close {{
    background: none;
    border: 1px solid var(--border-color);
    color: var(--text-color);
    cursor: pointer;
    padding: 4px 10px;
    border-radius: 4px;
    font-size: 16px;
}}
.cell-modal-close:hover {{ background-color: var(--hover-color); }}
</style>
</head>
<body data-theme="{default_theme}">

<!-- Cell expansion modal -->
<div class="cell-modal-overlay" id="cellModalOverlay" onclick="closeCellModal(event)">
  <div class="cell-modal">
    <div class="cell-modal-header">
      <span class="cell-modal-column" id="cellModalColumn"></span>
      <button class="cell-modal-close" onclick="closeCellModalDirect()">✕</button>
    </div>
    <div id="cellModalContent"></div>
  </div>
</div>

<div class="page-header">
  <div>
    <div class="page-title">{title}</div>
    <div class="page-meta">{total_rows:,} rows &nbsp;·&nbsp; {total_cols} columns</div>
  </div>
  <button class="theme-toggle" onclick="toggleTheme()" title="Toggle theme">
    <svg id="themeIconSun" viewBox="0 0 24 24" style="display:none">
      <circle cx="12" cy="12" r="4"/>
      <path d="M12 2v2M12 20v2M4.22 4.22l1.42 1.42M18.36 18.36l1.42 1.42M2 12h2M20 12h2M4.22 19.78l1.42-1.42M18.36 5.64l1.42-1.42"
            stroke="currentColor" stroke-width="2" stroke-linecap="round" fill="none"/>
    </svg>
    <svg id="themeIconMoon" viewBox="0 0 24 24">
      <path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z" fill="currentColor"/>
    </svg>
    Theme
  </button>
</div>

{table_html}

<script>
// ── Decompress full dataset (vanilla JS, no pako) ─────────────────────────
document.addEventListener('DOMContentLoaded', function() {{
    const datasetKeys = Object.keys(window).filter(k => k.startsWith('fullDataCompressed_'));
    datasetKeys.forEach(function(key) {{
        const queryId = key.replace('fullDataCompressed_', '');
        const b64 = window[key];
        if (!b64) return;
        try {{
            // Decode base64 → gzip bytes → JSON string
            const binStr = atob(b64);
            const bytes = new Uint8Array(binStr.length);
            for (let i = 0; i < binStr.length; i++) bytes[i] = binStr.charCodeAt(i);

            // Use DecompressionStream (supported in all modern browsers)
            const ds = new DecompressionStream('gzip');
            const writer = ds.writable.getWriter();
            writer.write(bytes);
            writer.close();

            new Response(ds.readable).text().then(function(text) {{
                const parsed = JSON.parse(text);
                const cols = parsed.columns;
                const fullData = parsed.data.map(function(row) {{
                    const obj = {{}};
                    for (let i = 0; i < cols.length; i++) obj[cols[i]] = row[i];
                    return obj;
                }});
                window['searchData_' + queryId] = fullData;
                window['searchColumns_' + queryId] = cols;
                console.log('✅ Ready:', queryId, fullData.length.toLocaleString(), 'rows');
            }});
        }} catch(e) {{
            console.error('Decompress failed:', e);
        }}
    }});

    updateThemeIcon(document.body.getAttribute('data-theme') || 'light');
    makeSortable();
}});

// ── Theme ─────────────────────────────────────────────────────────────────
function toggleTheme() {{
    const body = document.body;
    const next = body.getAttribute('data-theme') === 'light' ? 'dark' : 'light';
    body.setAttribute('data-theme', next);
    updateThemeIcon(next);
}}
function updateThemeIcon(theme) {{
    const sun  = document.getElementById('themeIconSun');
    const moon = document.getElementById('themeIconMoon');
    if (!sun || !moon) return;
    sun.style.display  = theme === 'dark' ? 'block' : 'none';
    moon.style.display = theme === 'dark' ? 'none'  : 'block';
}}

// ── Cell modal ────────────────────────────────────────────────────────────
function expandCell(td) {{
    const fullText = td.getAttribute('data-full-text') || td.textContent;
    const colName  = td.getAttribute('data-column') || '';
    document.getElementById('cellModalColumn').textContent  = colName;
    document.getElementById('cellModalContent').textContent = fullText;
    document.getElementById('cellModalOverlay').classList.add('active');
}}
function closeCellModal(e) {{
    if (e.target === document.getElementById('cellModalOverlay'))
        document.getElementById('cellModalOverlay').classList.remove('active');
}}
function closeCellModalDirect() {{
    document.getElementById('cellModalOverlay').classList.remove('active');
}}
document.addEventListener('keydown', function(e) {{
    if (e.key === 'Escape')
        document.getElementById('cellModalOverlay').classList.remove('active');
}});

// ── Column filter ─────────────────────────────────────────────────────────
function filterColumns(queryId) {{
    const input  = document.getElementById('col-search-' + queryId);
    const table  = document.getElementById('table-' + queryId);
    const count  = document.getElementById('col-count-' + queryId);
    const headers = table.querySelectorAll('th[data-column]');
    const rows    = table.querySelectorAll('tbody tr');
    const text    = input.value.trim().toLowerCase();
    const total   = headers.length;

    if (!text) {{
        headers.forEach(h => h.classList.remove('hidden-column'));
        rows.forEach(r => r.querySelectorAll('td').forEach(c => c.classList.remove('hidden-column')));
        count.textContent = 'Showing ' + total + ' of ' + total + ' columns';
        return;
    }}

    const terms = text.split(',').map(t => t.trim()).filter(Boolean);
    let visible = 0;
    headers.forEach((h, i) => {{
        const col = h.getAttribute('data-column');
        const match = terms.some(t => col.includes(t));
        match ? h.classList.remove('hidden-column') : h.classList.add('hidden-column');
        if (match) visible++;
        rows.forEach(row => {{
            const cells = row.querySelectorAll('td');
            if (cells[i])
                match ? cells[i].classList.remove('hidden-column') : cells[i].classList.add('hidden-column');
        }});
    }});
    count.textContent = 'Showing ' + visible + ' of ' + total + ' columns';
}}

// ── Row search ────────────────────────────────────────────────────────────
function filterRows(queryId, maxRows) {{
    const box   = document.getElementById('row-search-' + queryId);
    const input = box.querySelector('input');
    const table = document.getElementById('table-' + queryId);
    const count = document.getElementById('row-count-' + queryId);
    const term  = input.value.toLowerCase().trim();

    const fullData = window['searchData_' + queryId];
    const columns  = window['searchColumns_' + queryId];

    if (fullData && columns) {{
        filterRowsFullData(queryId, term, table, count, fullData, columns,
                           {str(enable_multiword_search).lower()}, maxRows);
    }} else {{
        filterRowsTableOnly(queryId, term, table, count, maxRows);
    }}
}}

function filterRowsFullData(queryId, term, table, countEl, fullData, columns,
                             enableMultiWord, maxRows) {{
    const tbody    = table.querySelector('tbody');
    const totalRows = window['fullDataRows_' + queryId] || fullData.length;
    const storedRows = Math.min({actual_stored}, totalRows);
    const visHeaders = Array.from(table.querySelectorAll('thead th:not(.hidden-column)'));
    const visCols    = visHeaders.map(h => h.getAttribute('data-column'));

    if (!term) {{
        tbody.innerHTML = '';
        const show = Math.min(maxRows, storedRows);
        for (let i = 0; i < storedRows && i < fullData.length; i++) {{
            tbody.appendChild(buildRow(fullData[i], columns, visCols, i, i >= show));
        }}
        countEl.textContent = 'Showing ' + show.toLocaleString() + ' of ' + totalRows.toLocaleString() + ' rows';
        return;
    }}

    const colsToSearch = (visCols.length === 0 || visCols.length === columns.length)
        ? columns
        : columns.filter(c => visCols.includes(c.toLowerCase()));

    let filtered;
    if (enableMultiWord) {{
        const words = term.split(/\s+/).filter(Boolean);
        filtered = fullData.filter(row =>
            words.every(w => colsToSearch.some(c => String(row[c] ?? '').toLowerCase().includes(w)))
        );
    }} else if (term.includes(':')) {{
        const [colPart, valPart] = term.split(':');
        const matchCol = columns.find(c => c.toLowerCase() === colPart.trim()) ||
                         columns.find(c => c.toLowerCase().includes(colPart.trim()));
        if (!matchCol) {{ countEl.textContent = '⚠️ Column not found'; return; }}
        filtered = fullData.filter(row => String(row[matchCol] ?? '').toLowerCase().includes(valPart.trim()));
    }} else {{
        filtered = fullData.filter(row =>
            colsToSearch.some(c => String(row[c] ?? '').toLowerCase().includes(term))
        );
    }}

    tbody.innerHTML = '';
    if (!filtered.length) {{
        tbody.innerHTML = '<tr><td colspan="' + columns.length + '" ' +
            'style="text-align:center;padding:20px;opacity:0.5;">No results for "' + term + '"</td></tr>';
        countEl.textContent = 'No results';
        return;
    }}

    const limit = Math.min(filtered.length, 1000);
    for (let i = 0; i < limit; i++) tbody.appendChild(buildRow(filtered[i], columns, visCols, i, false));
    const note = filtered.length > 1000 ? ' (showing first 1,000)' : '';
    countEl.textContent = 'Found ' + filtered.length.toLocaleString() + ' results' + note;
}}

function buildRow(rowData, columns, visCols, idx, hidden) {{
    const tr = document.createElement('tr');
    tr.setAttribute('data-row-id', idx);
    if (hidden) tr.classList.add('initially-hidden');
    columns.forEach(function(col) {{
        const td = document.createElement('td');
        td.setAttribute('data-column', col.toLowerCase());
        td.textContent = rowData[col] !== null && rowData[col] !== undefined ? String(rowData[col]) : '';
        if (visCols.length > 0 && visCols.length < columns.length && !visCols.includes(col.toLowerCase()))
            td.classList.add('hidden-column');
        tr.appendChild(td);
    }});
    return tr;
}}

function filterRowsTableOnly(queryId, term, table, countEl, maxRows) {{
    const rows  = table.querySelectorAll('tbody tr[data-row-id]');
    const total = rows.length;
    if (!term) {{
        rows.forEach(r => {{ r.classList.remove('hidden-row', 'search-match'); }});
        countEl.textContent = 'Showing ' + maxRows.toLocaleString() + ' of ' + total.toLocaleString() + ' rows';
        return;
    }}
    const visHeaders = Array.from(table.querySelectorAll('thead th:not(.hidden-column)'));
    const visCols    = visHeaders.map(h => h.getAttribute('data-column'));
    let visible = 0;
    rows.forEach(function(row) {{
        const cells = Array.from(row.querySelectorAll('td')).filter(c => visCols.includes(c.getAttribute('data-column')));
        const match = cells.some(c => c.textContent.toLowerCase().includes(term));
        match ? (row.classList.remove('hidden-row'), row.classList.add('search-match'), visible++)
              : (row.classList.add('hidden-row'),    row.classList.remove('search-match'));
    }});
    countEl.textContent = 'Showing ' + visible.toLocaleString() + ' of ' + total.toLocaleString() + ' rows';
}}

// ── Sortable columns ──────────────────────────────────────────────────────
function makeSortable() {{
    document.querySelectorAll('.data-table').forEach(function(table) {{
        const headers = table.querySelectorAll('thead th[data-column]');
        headers.forEach(function(header, colIdx) {{
            header.addEventListener('click', function() {{
                const tbody = table.querySelector('tbody');
                const rows  = Array.from(tbody.querySelectorAll('tr'));
                const asc   = this.getAttribute('data-sort') !== 'asc';
                headers.forEach(h => h.removeAttribute('data-sort'));
                this.setAttribute('data-sort', asc ? 'asc' : 'desc');
                const ind = this.querySelector('.sort-indicator');
                if (ind) {{ ind.textContent = asc ? '⮝' : '⮟'; ind.style.opacity = '1'; }}
                rows.sort(function(a, b) {{
                    const ca = a.children[colIdx] ? a.children[colIdx].textContent.trim() : '';
                    const cb = b.children[colIdx] ? b.children[colIdx].textContent.trim() : '';
                    const na = parseFloat(ca.replace(/[^0-9.-]/g, ''));
                    const nb = parseFloat(cb.replace(/[^0-9.-]/g, ''));
                    if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
                    return asc ? ca.localeCompare(cb) : cb.localeCompare(ca);
                }});
                rows.forEach(r => tbody.appendChild(r));
            }});
        }});
    }});
}}
</script>
</body>
</html>"""

    # ── 7. Write file ─────────────────────────────────────────────────────────
    out_path = Path(output_file)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")
    size_kb = out_path.stat().st_size / 1024
    print(f"✅ Table exported: {out_path.absolute()}  ({size_kb:.1f} KB, {total_rows:,} rows)")
    return str(out_path.absolute())


# ── Quick usage example ───────────────────────────────────────────────────────
if __name__ == "__main__":
    import pandas as pd
    from collections import Counter

    # Simulate a wfd result
    words = "the quick brown fox jumps over the lazy dog the fox".split()
    freq = Counter(words)
    total = sum(freq.values())
    df = pd.DataFrame({
        "word": list(freq.keys()),
        "frequency": list(freq.values()),
        "frequency %": [v / total * 100 for v in freq.values()],
        "total words": [total] * len(freq),
    }).sort_values("frequency", ascending=False).reset_index(drop=True)

    generate_data_table(
        df,
        output_file="data_table_example.html",
        title="Word Frequency Distribution",
        global_column_formats={
            "frequency %": "{:.2f}",
            "frequency": "{:,}",
            "total words": "{:,}",
        },
        freeze_columns=["word"],
        row_density="compact",
        default_theme="light",
        max_rows=25,
        embed_full_data=True,
    )
